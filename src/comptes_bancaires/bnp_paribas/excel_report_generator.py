from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils.dataframe import dataframe_to_rows

from comptes_bancaires.bnp_paribas.report_data_handler import ReportDataHandler
from database.compte_titre import CompteTireBdd

class ExcelReportGenerator(CompteTireBdd):
    """
    La classe `ExcelReportGenerator` hérite de `CompteTireBdd` et fournit des outils pour générer des rapports Excel
    à partir des opérations financières catégorisées. Elle permet de :

    - Créer des dossiers annuels pour stocker les rapports.
    - Filtrer les opérations par revenus et dépenses.
    - Générer des bilans par catégorie et sous-catégorie pour chaque année.
    - Calculer les totaux et pourcentages pour chaque mois et catégorie.
    - Formater automatiquement les fichiers Excel avec styles, bordures et formats monétaires.
    """

    def __init__(self, db_path: str, root_path: str):
        """
        Initialise le générateur et prépare l'arborescence des dossiers.

        Args:
            - db_path (str) : Chemin vers la base de données SQLite.
            - root_path (str) : Dossier racine pour le stockage des rapports.
        """
        super().__init__(db_path)
        self.__root_path = root_path
        
        ReportDataHandler._create_annual_folders(self.__root_path, self._get_categorized_operations_df())


    # --- [ Flux Principal ] ---
    def generate_all_reports(self, two_last_year_only: bool):
        """
        Génère les bilans financiers annuels pour toutes les années présentes dans les opérations catégorisées,
        crée un fichier Excel pour chaque année et y ajoute les feuilles de bilan.

        Arguments :
        - two_last_year_only (bool) : si True, génère les bilans pour les deux dernières années.
        """
        years_operations_categorisees = self._get_categorized_operations_by_year()

        # Créez les graphiques uniquement pour les 2 dernières années
        if two_last_year_only:
            two_last_years = list(years_operations_categorisees.keys())[-2:]
            years_operations_categorisees = {year: years_operations_categorisees[year] for year in two_last_years}

        # Regroupe toutes les opérations pour faire le bilan des différentes années
        all_operation_categorisees = pd.DataFrame()

        for year, operation_categorisees in years_operations_categorisees.items():
            self.__output_file = f"{self.__root_path}{year}/Bilan {year}.xlsx"
            self.__wb = self.__create_excel_file(operation_categorisees)
            self.__add_to_excel_file(operation_categorisees)
            all_operation_categorisees = pd.concat([all_operation_categorisees, operation_categorisees], ignore_index=True)
            
        # Bilan de toutes les années
        annees = list(years_operations_categorisees.keys())
        self.__output_file = f"{self.__root_path}/Bilan {annees[0]}-{annees[-1]}.xlsx"
        self.__wb = self.__create_excel_file(all_operation_categorisees)
        self.__add_to_excel_file(all_operation_categorisees)


    # --- [ Formatage Excel ] ---
    def __create_excel_file(self, df: pd.DataFrame) -> load_workbook:
        """
        Crée un fichier Excel à partir d'un DataFrame unique.
        
        Args:
            df (pd.DataFrame): Le DataFrame à sauvegarder.
            sheet_name (str): Le nom de la feuille Excel.
            
        Returns:
            Workbook: le fichier Excel ouvert avec openpyxl.
        """
        # S'assurer que toutes les colonnes sont des chaînes
        df.columns = [str(col) for col in df.columns]

        # Création du fichier Excel
        with pd.ExcelWriter(self.__output_file, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name="Données", index=False)

        # Charger le fichier Excel pour manipulation ultérieure
        return load_workbook(self.__output_file)

    def __add_to_excel_file(self, operation_categorisees: pd.DataFrame):
        """
        Met à jour un fichier Excel avec les données des opérations catégorisées.

        Args :
            operation_categorisees (pd.DataFrame) : DataFrame contenant les opérations catégorisées.

        Actions :
            - Crée des feuilles de bilan pour les revenus et dépenses.
            - Formate les cellules (alignement, format de date).
            - Crée des tableaux Excel avec style et ajustement automatique des colonnes.
            - Sauvegarde le fichier Excel final.
        """
        self.__sheet_bilan(operation_categorisees)
        center_alignment = Alignment(horizontal='center', vertical='center')

        for sheet_name in self.__wb.sheetnames:
            if sheet_name not in ["Bilan Revenus", "Bilan Dépenses"]:
                ws = self.__wb[sheet_name]
                min_col, min_row, max_col, max_row = 1, 1, ws.max_column, ws.max_row
                table_style = TableStyleInfo(
                    name='TableStyleMedium2', showFirstColumn=False,
                    showLastColumn=False, showRowStripes=True, showColumnStripes=True
                )
                safe_table_name = f'Table_{sheet_name}'.replace(" ", "_")
                table = Table(displayName=safe_table_name, ref=f'A1:{chr(64 + max_col)}{max_row}')
                table.tableStyleInfo = table_style
                ws.add_table(table)

                for row in ws.iter_rows(min_row=2, max_row=max_row):
                    for cell in row:
                        if isinstance(cell.value, datetime):
                            cell.number_format = 'DD/MM/YYYY'
                        cell.alignment = center_alignment

                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except Exception:
                            pass
                    adjusted_width = max_length + 10
                    ws.column_dimensions[column].width = adjusted_width

        self.__wb.save(self.__output_file)

    def __add_dataframe_to_sheet(self, df: pd.DataFrame, sheet_name: str, start_row: int, spacing: int = 5):
        """
        Ajoute un DataFrame à une feuille Excel avec un formatage spécifique.

        Args:
            df (pd.DataFrame): DataFrame à ajouter à la feuille.
            sheet_name (str): Nom de la feuille Excel.
            start_row (int): Ligne de départ pour ajouter les données.
            spacing (int): Espacement après le dernier bloc de données.
        """
        assert isinstance(df, pd.DataFrame), f"df doit être un DataFrame, mais c'est {type(df)}."
        assert isinstance(sheet_name, str), f"sheet_name doit être une chaîne de caractères, mais c'est {type(sheet_name)}."
        assert isinstance(start_row, int), f"start_row doit être un entier, mais c'est {type(start_row)}."
        assert isinstance(spacing, int), f"spacing doit être un entier, mais c'est {type(spacing)}."

        # Remplacer les zéros par une chaîne vide
        df.replace(0, '', inplace=True)

        if sheet_name in self.__wb.sheetnames:
            ws = self.__wb[sheet_name]
            start_row -= 1
        else:
            ws = self.__wb.create_sheet(title=sheet_name)

        for r in dataframe_to_rows(df, index=True, header=True):
            ws.append(r)

        # --- Styles ---
        font_month = Font(color="FFFFFF", bold=True)
        fill_month = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        alignment_center = Alignment(horizontal='center', vertical='center')
        font_revenue = Font(color="000000", bold=True)
        fill_revenue = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

        border_bold = Border(right=Side(style='thin'), top=Side(style='medium'), bottom=Side(style='medium'))
        border_bold_lr = Border(left=Side(style='medium'), right=Side(style='medium'))
        border_bold_right = Border(right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))
        border_standard = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Ligne des mois
        for row in ws.iter_rows(min_row=start_row, max_row=start_row, min_col=2, max_col=ws.max_column):
            for idx, cell in enumerate(row):
                if idx != ws.max_column-4:
                    cell.font = font_month
                    cell.fill = fill_month
                    cell.alignment = alignment_center
                    cell.border = border_bold

        normal = False
        # Formatage des cellules
        for idx_row, row in enumerate(ws.iter_rows(min_row=start_row+2, max_row=ws.max_row, min_col=0, max_col=ws.max_column)):
            if normal:
                for idx_col, cell in enumerate(row):
                    if (idx_col != ws.max_column-3) and (idx_col != ws.max_column-4) and (idx_col != ws.max_column-1):
                        if idx_col == 0:
                            cell.border = Border(right=Side(style='medium'), bottom=Side(style='thin'))
                        else:
                            cell.border = border_standard
                    elif (idx_col == ws.max_column-4) or (idx_col == ws.max_column-1):
                        cell.border = Border(right=Side(style='medium'), bottom=Side(style='thin'))

                    if (idx_row == ws.max_row-start_row-2) and (idx_col != ws.max_column-2):
                        if idx_col == 0:
                            cell.border = Border(right=Side(style='medium'), bottom=Side(style='medium'))
                        elif (idx_col == ws.max_column-4) or (idx_col == ws.max_column-1):
                            cell.border = Border(right=Side(style='medium'), bottom=Side(style='medium'))
                        else:
                            cell.border = Border(right=Side(style='thin'), bottom=Side(style='medium'))

                    if idx_col == ws.max_column-3:
                        cell.border = Border(right=Side(style='medium'))

            else:
                for idx_col, cell in enumerate(row):
                    if idx_col == 0:
                        cell.font = font_revenue
                        cell.fill = fill_revenue
                        cell.border = Border(right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))
                    elif idx_col != ws.max_column-3:
                        cell.font = font_revenue
                        cell.fill = fill_revenue
                        cell.border = border_bold
                    else:
                        cell.border = border_bold_lr

                row[-1].border = border_bold_right
                normal = True

        # Largeurs colonnes
        ws.column_dimensions['A'].width = 250 / 7
        ws.column_dimensions['P'].width = 100 / 7
        columns = [chr(i) for i in range(ord('B'), ord('M') + 1)] + ['O']
        for col_letter in columns:
            ws.column_dimensions[col_letter].width = 69 / 7

        for cell in ws['P']:
            cell.alignment = alignment_center

        # Espacement
        ws.append([''] * ws.max_column)
        for _ in range(spacing-1):
            ws.append([''] * ws.max_column)

        # Bordure cellule spécifique
        last_row_idx = ws.max_row - spacing
        second_last_col_idx = ws.max_column - 2
        specific_cell = ws.cell(row=last_row_idx, column=second_last_col_idx + 1)
        specific_cell.border = Border(right=Side(style='thin'), bottom=Side(style='medium'))

        # Format monétaire et pourcentage
        money_format = '#,##0.00 €'
        percentage_format = '0.00%'

        for col_num, col_name in enumerate(df.columns):
            col_name = col_name.strip()
            if col_name in ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec', 'Total']:
                for row in ws.iter_rows(min_row=start_row+1, max_row=ws.max_row, min_col=col_num+2, max_col=col_num+2):
                    for cell in row:
                        cell.number_format = money_format

        if 'Pourcentage' in df.columns:
            pct_col = df.columns.get_loc('Pourcentage')
            for row in ws.iter_rows(min_row=start_row+1, max_row=ws.max_row, min_col=pct_col+2, max_col=pct_col+2):
                for cell in row:
                    cell.number_format = percentage_format

    def __sheet_bilan(self, operation_categorisees: pd.DataFrame):
        """
        Prépare et ajoute les bilans des revenus et dépenses dans les feuilles Excel correspondantes.

        Args :
            operation_categorisees (pd.DataFrame) : DataFrame contenant les opérations catégorisées.

        Actions :
            - Crée un bilan par sous-catégories pour les revenus et l'ajoute à la feuille "Bilan Revenus".
            - Crée un bilan par catégories et sous-catégories pour les dépenses et l'ajoute à la feuille "Bilan Dépenses".
        """
        df_revenus = ReportDataHandler._get_income_df(operation_categorisees)
        df_revenus_save = self.__bilan_subcategories(df_revenus, "Revenus")
        self.__add_dataframe_to_sheet(df_revenus_save, "Bilan Revenus", 1)

        df_depenses = ReportDataHandler._get_expense_df(operation_categorisees)
        df_depenses_save = self.__bilan_categories(df_depenses, "Dépenses")
        self.__add_dataframe_to_sheet(df_depenses_save, "Bilan Dépenses", 1)

        df_depenses_save = self.__bilan_subcategories(df_depenses, "Dépenses")
        self.__add_dataframe_to_sheet(df_depenses_save, "Bilan Dépenses", self.__wb["Bilan Dépenses"].max_row + 2)


    # --- [ Traitement des Données ] ---
    def __bilan_categories(self, df_revenus: pd.DataFrame, sheet_name: str) -> pd.DataFrame:
        """
        Agrège les montants par catégorie et par mois, calcule les totaux et pourcentages, 
        et prépare un DataFrame prêt à être exporté vers Excel.

        Args:
            df_revenus (pd.DataFrame) : DataFrame contenant les opérations catégorisées.
            sheet_name (str) : nom de la feuille ou ligne de total à ajouter en tête.

        Returns:
            pd.DataFrame : DataFrame pivoté avec les totaux et pourcentages par catégorie.
        """
        df_revenus["mois"] = df_revenus["operation_date"].dt.strftime("%b")
        df_resultat = (
            df_revenus.groupby(["mois", "category"])['amount']
            .sum()
            .reset_index()
            .sort_values(["mois", "category"])
        )

        months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
        df_pivot = df_resultat.pivot_table(
            index="category",
            columns="mois",
            values='amount',
            aggfunc="sum",
            fill_value=0
        )
        df_pivot = df_pivot.reindex(columns=months, fill_value=0)

        # Calculer Total et Pourcentage pour chaque sous-catégorie
        df_pivot["Total"] = df_pivot[months].sum(axis=1)
        total_general = df_pivot["Total"].sum()
        df_pivot["Pourcentage"] = (df_pivot["Total"] / total_general).round(4)

        # Créer la ligne sheet_name avec les totaux par mois uniquement
        revenus_total = df_pivot[months].sum()            # somme uniquement des mois
        revenus_total[''] = ' '                           # colonne vide
        revenus_total['Total'] = df_pivot["Total"].sum()  # total général
        revenus_total['Pourcentage'] = 1.0               # 100%
        revenus_total.name = sheet_name

        # Ajouter la ligne sheet_name au-dessus
        df_pivot = pd.concat([pd.DataFrame([revenus_total]), df_pivot])

        # Trier les sous-catégories par Total ou Pourcentage (exclure sheet_name)
        df_sorted = pd.concat(
            [df_pivot.loc[[sheet_name]], df_pivot.drop(sheet_name).sort_values(by='Pourcentage', ascending=False)]
        )

        return df_sorted

    def __bilan_subcategories(self, df_revenus: pd.DataFrame, sheet_name: str) -> pd.DataFrame:
        """
        Agrège les montants par sous-catégorie et par mois, calcule les totaux et pourcentages, 
        et prépare un DataFrame prêt à être exporté vers Excel.

        Args:
            df_revenus (pd.DataFrame) : DataFrame contenant les opérations catégorisées.
            sheet_name (str) : nom de la feuille ou ligne de total à ajouter en tête.

        Returns:
            pd.DataFrame : DataFrame pivoté avec les totaux et pourcentages par sous-catégorie.
        """
        df_revenus["mois"] = df_revenus["operation_date"].dt.strftime("%b")
        df_resultat = (
            df_revenus.groupby(["mois", "sub_category"])['amount']
            .sum()
            .reset_index()
            .sort_values(["mois", "sub_category"])
        )

        months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
        df_pivot = df_resultat.pivot_table(
            index="sub_category",
            columns="mois",
            values='amount',
            aggfunc="sum",
            fill_value=0
        )
        df_pivot = df_pivot.reindex(columns=months, fill_value=0)

        # Calculer Total et Pourcentage pour chaque sous-catégorie
        df_pivot["Total"] = df_pivot[months].sum(axis=1)
        total_general = df_pivot["Total"].sum()
        df_pivot["Pourcentage"] = (df_pivot["Total"] / total_general).round(4)

        # Créer la ligne sheet_name avec les totaux par mois uniquement
        revenus_total = df_pivot[months].sum()            # somme uniquement des mois
        revenus_total[''] = ' '                           # colonne vide
        revenus_total['Total'] = df_pivot["Total"].sum()  # total général
        revenus_total['Pourcentage'] = 1.0               # 100%
        revenus_total.name = sheet_name

        # Ajouter la ligne sheet_name au-dessus
        df_pivot = pd.concat([pd.DataFrame([revenus_total]), df_pivot])

        # Trier les sous-catégories par Total ou Pourcentage (exclure sheet_name)
        df_sorted = pd.concat(
            [df_pivot.loc[[sheet_name]], df_pivot.drop(sheet_name).sort_values(by='Pourcentage', ascending=False)]
        )

        return df_sorted
