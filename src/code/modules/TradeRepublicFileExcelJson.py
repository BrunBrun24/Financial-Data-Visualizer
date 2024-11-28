import pandas as pd
import os
import re
from datetime import datetime, date
from openpyxl import Workbook
import pdfplumber
import json
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


class TradeRepublicFileExcelJson:
    """
    La classe TradeRepublicFileExcelJson est conçue pour analyser les données du courtier Trade République via les fichiers PDF téléchargés,
    et créer un fichier Excel et des fichiers JSON avec ces données.

    Attributs :
        - directoryData : Chemin du répertoire de données contenant les fichiers PDF.
        - tickerMapping : Dictionnaire pour mapper les tickers avec des informations supplémentaires.
    """

    def __init__(self, directoryData: str, tickerMapping: dict):
        """
        Initialise une instance de TradeRepublicFileExcelJson avec un répertoire de données.

        Args:
            directoryData (str): Chemin du répertoire de données contenant les fichiers PDF.
            tickerMapping (dict): Un dictionnaire pour mapper les tickers avec des informations supplémentaires.

        Raises:
            AssertionError: Si directoryData n'est pas une chaîne, n'existe pas ou n'est pas un répertoire.
        """
        assert isinstance(directoryData, str), f"directoryData doit être une chaîne: {type(directoryData)}."
        assert os.path.exists(directoryData), f"Le fichier ou le dossier '{directoryData}' n'existe pas."
        assert os.path.isdir(directoryData), f"'{directoryData}' n'est pas un dossier."
        assert isinstance(tickerMapping, dict), f"mappingTickers doit être un dictionnaire: ({type(tickerMapping)})"

        self.directoryData = directoryData
        self.tickerMapping = tickerMapping

        self.ProcessPdf()



    #################### MAIN ####################
    def DownloadDataAndCreateFileExcel(self, cheminEnregistrementFichier: str):
        """
        Crée un récapitulatif des gains en générant un classeur Excel avec plusieurs feuilles pour différents types de données financières.
        Les feuilles sont créées à partir de fichiers PDF spécifiques dans le dossier spécifié.

        Args:
            cheminEnregistrementFichier (str): Chemin où le fichier Excel sera enregistré.
        """
        assert isinstance(cheminEnregistrementFichier, str), \
            f"cheminEnregistrementFichier doit être une chaîne: {type(cheminEnregistrementFichier)}."

        buySell = []
        # Créez un classeur Excel
        workbook = Workbook()
        # Supprime la feuille par défaut "Sheet"
        sheetToDelete = workbook["Sheet"]
        workbook.remove(sheetToDelete)

        # L'ordre dans le dictionnaire définit la continuité des feuilles dans le fichier Excel
        operations = [
            {"nomFichier": "Ordres d'achats", "data": self.DataOrdresAchats, "fileJson": "Bilan/Archives/Bourse/Argents investis.json"},
            {"nomFichier": "Ordres de ventes/FacturesVentes", "data": self.DataOrdresVentes, "fileJson": "Bilan/Archives/Bourse/Argents vendus.json"},
            {"nomFichier": "Dépôts d'argents", "data": self.DataDepotArgent, "fileJson": "Bilan/Archives/Bourse/Dépôts d'espèces.json"},
            {"nomFichier": "Dividendes", "data": self.DataDividendes, "fileJson": "Bilan/Archives/Bourse/Dividendes.json"},
            {"nomFichier": "Interets", "data": self.DataInteret, "fileJson": "Bilan/Archives/Bourse/Interêts.json"},
            {"nomFichier": "Retraits d'argents", "data": self.DataRetraitArgent, "fileJson": "Bilan/Archives/Bourse/Retraits.json"},
        ]

        # Boucle pour traiter chaque type de donnée et créer les feuilles Excel correspondantes
        for operation in operations:
            nomFichier = self.directoryData + operation["nomFichier"]
            filesPdf = self.GetPdfFilesFromFolder(nomFichier)

            dataFrame, nomFeuille, colonnesEuros, colonnesPourcentages, colonnesDates, appliquerTableau = operation["data"](filesPdf, nomFichier)
            dataFrame.columns = [col.capitalize() for col in dataFrame.columns]
            self.PutDataFrameSheetExcel(workbook=workbook, nomFeuille=nomFeuille, df=dataFrame, colonnesEuros=colonnesEuros, colonnesPourcentages=colonnesPourcentages, colonnesDates=colonnesDates, appliquerTableau=appliquerTableau)
            self.SaveDataFrameAsJson(dataFrame, operation["fileJson"])

            if operation["nomFichier"] in ["Ordres d'achats", "Ordres de ventes/FacturesVentes"]:
                buySell.append(dataFrame)


        self.SaveTransactionsToJson(buySell, "Bilan/Archives/Bourse/Transactions.json")
        # Enregistrez le fichier Excel
        workbook.save(cheminEnregistrementFichier)
    ##############################################



    #################### JSON ####################
    def SaveDataFrameAsJson(self, dataFrame: pd.DataFrame, pathJson: str):
        """
        Enregistre un DataFrame sous forme de fichier JSON et applique un mapping des tickers si une colonne 'ticker' existe.

        Args:
            dataFrame (pd.DataFrame): Le DataFrame à sauvegarder.
            pathJson (str): Le chemin où sauvegarder le fichier JSON.
        """
        assert isinstance(dataFrame, pd.DataFrame), f"dataFrame doit être un pd.DataFrame: {type(dataFrame)}."
        assert isinstance(pathJson, str) and pathJson.endswith(".json"), \
            f"pathJson doit être une chaîne se terminant par '.json': {type(pathJson)}."

        # Créer une copie du DataFrame pour éviter les modifications de l'original
        dataFrameCopy = dataFrame.copy()

        mappingTickers = self.tickerMapping.copy()
        dataFrameCopy = self.ConvertDateColumnsFromFirstRow(dataFrameCopy, "%Y-%m-%d")

        # Normalisation des clés du dictionnaire de mapping en minuscules
        mappingTickers = {key.lower(): value for key, value in mappingTickers.items()}

        # Ne pas modifier le nom des colonnes mais les comparer en minuscules
        columnsLower = [col.lower() for col in dataFrameCopy.columns]

        # Vérifier si une colonne 'ticker' existe (insensible à la casse)
        if 'ticker' in columnsLower:
            tickerCol = dataFrameCopy.columns[columnsLower.index('ticker')]  # Récupérer le nom original de la colonne 'ticker'

            # Création d'un motif pour rechercher les noms d'entreprises dans les tickers
            pattern = re.compile('|'.join(re.escape(key) for key in mappingTickers.keys()), re.IGNORECASE)

            # Appliquer le mappage des tickers à la colonne 'ticker'
            dataFrameCopy[tickerCol] = dataFrameCopy[tickerCol].apply(
                lambda ticker: self.MapTickerValue(ticker, mappingTickers, pattern)
            )

        # Convertir le DataFrame en JSON
        jsonData = dataFrameCopy.to_json(orient="records", force_ascii=False, indent=4)

        # Enregistrer le JSON dans le fichier
        with open(pathJson, 'w', encoding="utf-8") as f:
            f.write(jsonData)

    def SaveTransactionsToJson(self, dataFrames: list, pathJson: str):
        """
        Enregistre les transactions d'achats et de ventes dans un fichier JSON structuré, avec correction des tickers.

        Args:
            dataFrames (list): Liste contenant deux DataFrames, l'un pour les achats, l'autre pour les ventes.
            pathJson (str): Chemin où sauvegarder le fichier JSON.
        """
        # Vérifications des types des arguments
        assert isinstance(dataFrames, list) and len(dataFrames) == 2, \
            f"dataFrames doit être une liste contenant deux DataFrames: {type(dataFrames)}."
        assert isinstance(pathJson, str) and pathJson.endswith(".json"), \
            f"pathJson doit être une chaîne se terminant par '.json': {type(pathJson)}."
        
        tickerMapping = self.tickerMapping.copy()
        achats, ventes = dataFrames

        # Normalisation des clés du dictionnaire de mapping en minuscules
        tickerMapping = {key.lower(): value for key, value in tickerMapping.items()}

        # Création d'un motif pour rechercher les noms d'entreprises dans les tickers
        pattern = re.compile('|'.join(re.escape(key) for key in tickerMapping.keys()), re.IGNORECASE)

        # Format JSON pour les transactions
        transactions = []

        # Concaténer achats et ventes pour obtenir tous les tickers uniques
        tickers = sorted(set(achats["Ticker"]).union(set(ventes["Ticker"])))

        for ticker in tickers:
            # Filtrer les transactions pour ce ticker
            achatDf = achats[achats["Ticker"] == ticker]
            venteDf = ventes[ventes["Ticker"] == ticker]

            # Corriger les tickers à partir des noms d'entreprises
            match = pattern.search(ticker.lower())
            if match:
                matchedKey = match.group(0).lower()
                ticker = tickerMapping[matchedKey]
            
            transaction = None
            for tr in transactions:
                if tr["ticker"] == ticker:
                    transaction = tr
                    break

            if not transaction:
                transaction = {"ticker": ticker, "achats": [], "ventes": []}
                transactions.append(transaction)

            # Ajouter les achats au ticker
            for _, row in achatDf.iterrows():
                transaction["achats"].append({
                    "date": str(row["Date d'exécution"]),
                    "price": abs(row["Montant investi"])
                })

            # Ajouter les ventes au ticker
            for _, row in venteDf.iterrows():
                transaction["ventes"].append({
                    "date": str(row["Date d'exécution"]),
                    "price": abs(row["Montant gagné"])
                })

        # Sauvegarde du fichier JSON
        with open(pathJson, 'w', encoding='utf-8') as f:
            json.dump({"transactions": transactions}, f, ensure_ascii=False, indent=4)
    
    ########## ANNEXES ##########
    @staticmethod
    def MapTickerValue(ticker: str, mappingTickers: dict, pattern: re.Pattern) -> str:
        """
        Remplace la valeur du ticker si elle correspond à un nom d'entreprise dans le dictionnaire.

        Args:
            ticker (str): La valeur du ticker à analyser.
            mappingTickers (dict): Le dictionnaire de correspondance des tickers (clefs normalisées en minuscules).
            pattern (re.Pattern): Le motif regex pour rechercher les noms d'entreprises dans les tickers.

        Returns:
            str: La valeur du ticker mappée ou l'original si aucune correspondance n'est trouvée.
        """
        assert isinstance(ticker, str), f"ticker doit être une chaîne: {type(ticker)}."
        assert isinstance(mappingTickers, dict), f"mappingTickers doit être un dictionnaire: {type(mappingTickers)}."
        assert isinstance(pattern, re.Pattern), f"pattern doit être une instance de re.Pattern: {type(pattern)}."

        match = pattern.search(ticker.lower())
        if match:
            matchedKey = match.group(0).lower()
            return mappingTickers[matchedKey]
        return ticker
    
    @staticmethod
    def ConvertDateColumnsFromFirstRow(dataFrame: pd.DataFrame, dateFormat: str="%Y-%m-%d") -> pd.DataFrame:
        """
        Identifie les colonnes contenant des objets datetime dans la première ligne,
        puis convertit ces colonnes en chaînes de caractères dans le DataFrame.

        Args:
            dataFrame (pd.DataFrame): Le DataFrame contenant les données.
            dateFormat (str): Le format souhaité pour les chaînes de caractères. Par défaut "%Y-%m-%d".

        Returns:
            pd.DataFrame: Le DataFrame avec les colonnes de dates converties en chaînes.
        """
        assert isinstance(dataFrame, pd.DataFrame), "dataFrame doit être un objet pandas DataFrame"
        assert isinstance(dateFormat, str), "dateFormat doit être une chaîne de caractères"

        # Identifier les colonnes à partir de la première ligne
        dateColumns = [
            column for column in dataFrame.columns
            if isinstance(dataFrame.iloc[0][column], date)
        ]

        # Convertir les colonnes identifiées en chaînes
        for column in dateColumns:
            # Convertir en datetime64 avant de formater si nécessaire
            dataFrame[column] = pd.to_datetime(dataFrame[column])
            dataFrame[column] = dataFrame[column].dt.strftime(dateFormat)

        return dataFrame
    #############################
    ###################################################



    #################### EXCEL ####################
    @staticmethod
    def PutDataFrameSheetExcel(workbook: Workbook, nomFeuille: str, df: pd.DataFrame, colonnesEuros=[], colonnesPourcentages=[], colonnesDates=[], appliquerTableau=False):
        """
        Formate les colonnes du DataFrame dans le classeur Excel en fonction des listes de colonnes à formater et ajoute les données dans la feuille spécifiée.

        Args:
            workbook (Workbook): Le classeur Excel où les données seront ajoutées.
            nomFeuille (str): Le nom de la feuille où les données seront ajoutées.
            df (pd.DataFrame): Le DataFrame contenant les données à formater.
            colonnesEuros (list): Liste des colonnes à formater en euros.
            colonnesPourcentages (list): Liste des colonnes à formater en pourcentage.
            colonnesDates (list): Liste des colonnes à formater en dates.
            appliquerTableau (bool): Indique s'il faut appliquer un tableau autour des données.
        """
        # Crée une nouvelle feuille dans le classeur Excel avec le nom spécifié
        feuille = workbook.create_sheet(title=nomFeuille)

        # Qui est la première colonne qui correspond aux dates de la plus récente à la plus ancienne
        df = df.sort_values(by=df.columns[0], ascending=False)

        # Ajoute les données du DataFrame à la feuille
        for row in dataframe_to_rows(df, index=False, header=True):
            feuille.append(row)

        # Obtenir les indices des colonnes à formater
        colonnesEurosIndices = [df.columns.get_loc(col) + 1 for col in colonnesEuros if col in df.columns]
        colonnesPourcentagesIndices = [df.columns.get_loc(col) + 1 for col in colonnesPourcentages if col in df.columns]
        colonnesDatesIndices = [df.columns.get_loc(col) + 1 for col in colonnesDates if col in df.columns]

        # Formater les colonnes en euros
        for col in colonnesEurosIndices:
            for cellule in feuille.iter_rows(min_col=col, max_col=col, min_row=2):
                for cell in cellule:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '0.00€'

        # Formater les colonnes en pourcentage
        for col in colonnesPourcentagesIndices:
            for cellule in feuille.iter_rows(min_col=col, max_col=col, min_row=2):
                for cell in cellule:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '0.00%'

        # Formater les colonnes en dates
        for col in colonnesDatesIndices:
            for cellule in feuille.iter_rows(min_col=col, max_col=col, min_row=2):
                for cell in cellule:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = 'DD/MM/YYYY'

        # Appliquer le tableau si le booléen est à True
        if appliquerTableau:
            ref = f"A1:{get_column_letter(feuille.max_column)}{feuille.max_row}"
            tableau = Table(displayName=("TableauDonnees" + nomFeuille).replace(" ", "").replace("'", ""), ref=ref)
            style = TableStyleInfo(
                name="TableStyleMedium6", showFirstColumn=False,
                showLastColumn=False, showRowStripes=True, showColumnStripes=True
            )
            tableau.tableStyleInfo = style
            feuille.add_table(tableau)

        # Ajuster automatiquement la largeur des colonnes en fonction du contenu
        for column in feuille.columns:
            max_length = 0
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
                cell.alignment = Alignment(horizontal="center", vertical="center")
            feuille.column_dimensions[get_column_letter(column[0].column)].width = max_length + 3
    ###############################################



    #################### PDF ####################
    ########## RENAME AND MOVE FILES ##########
    def ProcessPdf(self):
        """
        Parcourt les fichiers PDF dans un répertoire, puis utilise une fonction spécifiée pour traiter chaque fichier PDF et le déplacer.
        """
        
        directory = self.directoryData
        operations = [
            {"directory": (directory + "Ordres d'achats/data"), "directoryRename": (directory + "Ordres d'achats"), "createFunction": self.RenameAndMoveOrdresAchats},
            {"directory": (directory + "Dépôts d'argents/data"), "directoryRename": (directory + "Dépôts d'argents"), "createFunction": self.RenameAndMoveDepotRetraitArgentInteret},
            {"directory": (directory + "Dividendes/data"), "directoryRename": (directory + "Dividendes"), "createFunction": self.RenameAndMoveDividendes},
            {"directory": (directory + "Interets/data"), "directoryRename": (directory + "Interets"), "createFunction": self.RenameAndMoveDepotRetraitArgentInteret},
            {"directory": (directory + "Retraits d'argents/data"), "directoryRename": (directory + "Retraits d'argents"), "createFunction": self.RenameAndMoveDepotRetraitArgentInteret},
            {"directory": (directory + "Ordres de ventes/data"), "directoryRename": (directory + "Ordres de ventes/FacturesVentes"), "createFunction": self.RenameAndMoveOrdresVentes},
        ]

        for operation in operations:
            nameDirectory = operation["directory"]
            cheminDossierRenommer = operation["directoryRename"]
            nameFunction = operation["createFunction"]

            # Créer le répertoire de destination s'il n'existe pas
            if not os.path.exists(cheminDossierRenommer):
                os.makedirs(cheminDossierRenommer)

            # Parcourir tous les fichiers dans le dossier spécifié
            for fileName in os.listdir(nameDirectory):
                # Vérifier si le fichier est un fichier PDF
                if fileName.lower().endswith('.pdf'):
                    filePath = os.path.join(nameDirectory, fileName)
                    # Appeler la fonction spécifiée pour traiter et déplacer le fichier
                    try:
                        nameFunction(filePath, cheminDossierRenommer)
                    except:
                        nameFunction(filePath)

    def RenameAndMoveDepotRetraitArgentInteret(self, filePath: str, DossierDestination: str):
        """
        Renomme et déplace un fichier PDF en fonction des informations extraites de son texte.

        Args:
            filePath: Le chemin complet du fichier PDF à renommer et déplacer.
            DossierDestination: Le dossier où le fichier PDF sera déplacé après avoir été renommé.
        """
        assert isinstance(filePath, str) and os.path.isfile(filePath), f"Le fichier '{filePath}' n'existe pas ou n'est pas une extension .pdf"
        assert isinstance(DossierDestination, str) and os.path.isdir(DossierDestination), f"Le dossier '{DossierDestination}' n'existe pas"

        # Extraire le texte du fichier PDF
        text = self.ExtractInformationPdf(filePath)
        date = datetime.strptime(self.ExtraireDonnee(text, r'DATE ((\d{2}/\d{2}/\d{4})|(\d{2}.\d{2}.\d{4}))', 1).replace(".", "/"), "%d/%m/%Y").strftime("%Y-%m-%d")

        numero = 1
        # Si on vend plusieurs fois la même action le même jour alors on rajoute un nombre +1
        while os.path.exists(os.path.join(DossierDestination, date + " - " + str(numero) + ".pdf")):
            numero += 1
        NomFichierModifier = date + " - " + str(numero) + ".pdf"

        try:
            # Déplacer et renommer le fichier dans le dossier Details
            os.rename(filePath, os.path.join(DossierDestination, NomFichierModifier))
        except Exception as e:
            print(f"Une erreur s'est produite lors de l'enregistrement du fichier : {e}")

    def RenameAndMoveDividendes(self, filePath, DossierDestination):
        """
        Renomme et déplace un fichier PDF en fonction des informations extraites du texte.

        Args:
            filePath: Chemin complet vers le fichier PDF à renommer.
            DossierDestination: Dossier de destination où le fichier sera déplacé.
        """
        assert isinstance(filePath, str) and os.path.isfile(filePath), f"Le fichier '{filePath}' n'existe pas ou n'est pas une extension .pdf"
        assert isinstance(DossierDestination, str) and os.path.isdir(DossierDestination), f"Le dossier '{DossierDestination}' n'existe pas"

        # Extraire le texte du fichier PDF
        text = self.ExtractInformationPdf(filePath)

        date = datetime.strptime(self.ExtraireDonnee(text, r'DATE ((\d{2}/\d{2}/\d{4})|(\d{2}.\d{2}.\d{4}))', 1).replace(".", "/"), "%d/%m/%Y").strftime("%Y-%m-%d")
        ticker = self.ExtraireDonnee(text, r"(POSITION|QUANTITÉ)[^\n]*\n([A-Za-zàâäéèêëîïôöùûü'\s&.-]+)", 2)

        NomFichierModifier = f"{ticker} ({date}).pdf"

        try:
            # Déplacer et renommer le fichier dans le dossier de destination
            os.rename(filePath, os.path.join(DossierDestination, NomFichierModifier))
        except Exception as e:
            print("Une erreur s'est produite lors de l'enregistrement du fichier :", e)

    def RenameAndMoveOrdresAchats(self, filePath: str, DossierDestination: str):
        """
        Extrait des informations d'un fichier PDF pour renommer et déplacer le fichier dans un autre répertoire.

        Args:
            filePath: Chemin complet du fichier PDF à renommer et déplacer.
            DossierDestination: Chemin du répertoire où le fichier renommé sera déplacé.
        """
        assert isinstance(filePath, str) and os.path.isfile(filePath), f"Le fichier '{filePath}' n'existe pas ou n'est pas une extension .pdf"
        assert isinstance(DossierDestination, str) and os.path.isdir(DossierDestination), f"Le dossier '{DossierDestination}' n'existe pas"

        # Extraire le texte du fichier PDF
        text = self.ExtractInformationPdf(filePath)

        date = datetime.strptime(self.ExtraireDonnee(text, r'DATE ((\d{2}/\d{2}/\d{4})|(\d{2}.\d{2}.\d{4}))', 1).replace(".", "/"), "%d/%m/%Y").strftime("%Y-%m-%d")
        ticker = self.ExtraireDonnee(text, r"(POSITION|QUANTITÉ)[^\n]*\n([A-Za-zàâäéèêëîïôöùûü'\s&.-]+)", 2)

        NomFichierModifier = f"{ticker} ({date}).pdf"
        try:
            # Déplacer et renommer le fichier dans le répertoire DossierDestination
            os.rename(filePath, os.path.join(DossierDestination, NomFichierModifier))
        except Exception as e:
            print(f"Une erreur s'est produite lors du déplacement du fichier {filePath}: {e}")

    def RenameAndMoveOrdresVentes(self, filePath: str):
        """
        Renomme et déplace un fichier PDF basé sur les informations extraites du texte du fichier.
        Le nouveau nom de fichier est basé sur la date et le nom du ticker extraits du texte.
        Le fichier est déplacé vers un dossier spécifique en fonction du type d'information.

        Args:
            filePath (str): Le chemin complet du fichier PDF à renommer et déplacer.
        """
        assert isinstance(filePath, str) and os.path.isfile(filePath), f"Le fichier '{filePath}' n'existe pas ou n'est pas une extension .pdf"

        # Extraire le texte du fichier PDF
        text = self.ExtractInformationPdf(filePath)
        date = datetime.strptime(self.ExtraireDonnee(text, r'DATE (\d{2}/\d{2}/\d{4})', 1), "%d/%m/%Y").strftime("%Y-%m-%d")
        ticker = self.ExtraireDonnee(text, r'(?:POSITION QUANTITÉ PRIX MONTANT|TITRE ORDRE / QUANTITÉ VALEUR)\s+([^\n]+)', 1)


        # Déterminer le dossier de destination et le nom du fichier
        if re.search(re.escape("INFORMATIONS SUR LES COÛTS EX-ANTE DE LA VENTE DE TITRES"), text):
            debutNomFichier = f"{ticker} ({date}) Informations sur les coûts "
            DossierDestination = os.path.join('Bilan/Archives/Bourse/Fichiers pdf/Ordres de ventes/InformationsCoûts')
        else:
            debutNomFichier = f"{ticker} ({date}) Facture Vente "
            DossierDestination = os.path.join('Bilan/Archives/Bourse/Fichiers pdf/Ordres de ventes/FacturesVentes')

        # Assurer l'existence du dossier de destination
        os.makedirs(DossierDestination, exist_ok=True)

        numero = 1
        # Si on vend plusieurs fois la même action le même jour alors on rajoute un nombre +1
        while os.path.exists(os.path.join(DossierDestination, debutNomFichier + str(numero) + ".pdf")):
            numero += 1
        NomFichierModifier = debutNomFichier + str(numero) + ".pdf"

        # Déplacer et renommer le fichier
        newFilePath = os.path.join(DossierDestination, NomFichierModifier)
        try:
            os.rename(filePath, newFilePath)
        except FileNotFoundError as e:
            print(f"Erreur : {filePath} non trouvé. Détails : {e}")
        except PermissionError as e:
            print(f"Erreur : Permission refusée pour {filePath}. Détails : {e}")
        except Exception as e:
            print(f"Erreur inattendue lors du déplacement de {filePath}. Détails : {e}")
    ###########################################

    ########## EXTRACTION DES DONNEES ##########
    def DataDepotArgent(self, pdfFiles: list, chemin: str) -> pd.DataFrame:
        """
        Extrait les données des fichiers PDF spécifiés et les retourne sous forme de DataFrame.

        Args:
            pdfFiles: Liste des noms de fichiers PDF à traiter.
            chemin: Chemin du répertoire contenant les fichiers PDF.

        Returns:
            pd.DataFrame: DataFrame contenant les données extraites de chaque fichier PDF.
        """
        assert isinstance(pdfFiles, list) and all(fichier.endswith(".pdf") for fichier in pdfFiles), \
            f"pdfFiles doit être une liste dont Le contenu contient des chaînes de caractères avec l'extension .pdf: {pdfFiles}"
        assert isinstance(chemin, str) and os.path.isdir(chemin), f"Le dossier '{chemin}' n'existe pas"

        df = pd.DataFrame()
        
        for fichier in pdfFiles:
            cheminComplet = os.path.join(chemin, fichier)

            try:
                with pdfplumber.open(cheminComplet) as pdf:
                    page = pdf.pages[0]
                    text = page.extract_text()

                compteTitre = int(self.ExtraireDonnee(text, r"COMPTE-TITRES (\d+)", 1))
                montantBrut = float(self.ExtraireDonnee(text, r"(.+)(?=\s*DÉTAIL)", 1)[6:-4].replace(',', '.'))
                frais = float(self.ExtraireDonnee(text, r"(?<=Frais de paiements par carte de crédit ).+", 0)[:-4].replace(',', '.'))

                data = self.ExtraireDonnee(text, r"(?<=COMPTE-ESPÈCES DATE DE VALEUR MONTANT\n).+", 0).split()
                compteEspece = data[0]
                dateValeur = datetime.strptime(data[1], "%d/%m/%Y").date()
                montant = float(data[2].replace(',', '.'))

                # Ajouter les données au DataFrame
                newRow = {
                    "Date de valeur": dateValeur,
                    "Prix dépôt brut": montantBrut,
                    "Frais": frais,
                    "Prix dépôt net": montant,
                    "Compte-titres": compteTitre,
                    "Compte-espèces": compteEspece
                }

                # Ajouter la nouvelle ligne au DataFrame
                df = pd.concat([df, pd.DataFrame([newRow])], ignore_index=True)

            except AssertionError as e:
                print(f"AssertionError: {e}")
            except Exception as e:
                print(f"Erreur lors du traitement du fichier {fichier}: {e}")


        nomFeuille = "Dépôts Argents"
        colonnesEuros = ["Prix dépôt brut", "Frais", "Prix dépôt net"]
        colonnesPourcentages = []
        colonnesDates = ["Date de valeur"]
        appliquerTableau = True

        return df, nomFeuille, colonnesEuros, colonnesPourcentages, colonnesDates, appliquerTableau

    def DataDividendes(self, pdfFiles: list, chemin: str) -> pd.DataFrame:
        """
        Extrait les données des fichiers PDF spécifiés et les retourne sous forme de DataFrame.

        Args:
            pdfFiles: Liste des noms de fichiers PDF à traiter.
            chemin: Chemin du répertoire contenant les fichiers PDF.

        Returns:
            pd.DataFrame: DataFrame contenant les données extraites de chaque fichier PDF.
        """
        # Assertions pour vérifier les types des entrées
        assert isinstance(pdfFiles, list) and (fichier.endswith(".pdf") for fichier in pdfFiles), \
            f"pdfFiles doit être une liste dont Le contenu contient des chaînes de caractères avec l'extension .pdf: {pdfFiles}"
        assert isinstance(chemin, str) and os.path.isdir(chemin), f"Le dossier '{chemin}' n'existe pas"

        df = pd.DataFrame()

        for fichier in pdfFiles:
            cheminComplet = os.path.join(chemin, fichier)

            with pdfplumber.open(cheminComplet) as pdf:
                page = pdf.pages[0]
                text = page.extract_text()

            dateValeur = datetime.strptime((self.ExtraireDonnee(text, r'DE\d{20}\s(\d{2}[./]\d{2}[./]\d{4})', 1)).replace(".", "/"), "%d/%m/%Y").date()
            dateDividende = datetime.strptime((self.ExtraireDonnee(text, r'(?:au -|à la date du)\s*(\d{2}[./]\d{2}[./]\d{4})', 1)).replace(".", "/"), "%d/%m/%Y").date()
            compteEspece = self.ExtraireDonnee(text, r"DE\d{20}", 0)
            isin = self.ExtraireDonnee(text, r"\b[A-Z]{2}[A-Z0-9]{9}[0-9]\b", 0)
            titreDetenue = float(self.ExtraireDonnee(text, r'([-+]?\d+[\.,]?\d*)\s+(titre\(s\)|unit\.)', 1).replace(",", "."))
            ticker = self.ExtraireDonnee(text, r"(POSITION|QUANTITÉ)[^\n]*\n([A-Za-zàâäéèêëîïôöùûü'\s&.-]+)", 2)
            montantGagne = float(self.ExtraireDonnee(text, r"(?<=COMPTE-ESPÈCES DATE DE VALEUR MONTANT\n).+|COMPTE DE RÈGLEMENT DATE DE PAIEMENT MONTANT\n.+", 0).split()[-2].replace(',', '.'))

            compteTitre = self.ExtraireDonnee(text, r"COMPTE-TITRES\s+(\d+)", 1)
            if compteTitre == None:
                compteTitre = self.ExtraireDonnee(text, r"(\d+)\s+DIVIDENDE EN ESPÈCES", 1)
            if compteTitre == None:
                raise ValueError("compteTitre non trouvée dans le texte.")

            montantAvantFrais = self.ExtraireDonnee(text, r"(.+)(?=\s*DÉTAIL)|(.+)(?=\s*RELEVÉ)", 0)
            if not montantAvantFrais:
                montantAvantFrais = 0
            else:
                montantAvantFrais = float(montantAvantFrais.split()[-2].replace(",", "."))

            newImpot = self.ExtraireDonnee(text, r"(?<=Impôt à la source ).+|(?<=Impôt à la source pour).+", 0)
            newImpot = float(newImpot.split()[-2].replace(",", ".")) if newImpot else 0.0

            newData = {
                "Date de valeur": dateValeur,
                "Ticker": ticker,
                "Dividendes brut": montantAvantFrais,
                "Impôt à la source": newImpot,
                "Dividendes net": montantGagne,
                "Titre(s) détenue(s)": titreDetenue,
                "Dividende à la date du ...": dateDividende,
                "Isin": isin,
                "COMPTE-TITRE": compteTitre,
                "COMPTE ESPÈCES": compteEspece
            }

            # Ajouter la nouvelle ligne au DataFrame
            df = pd.concat([df, pd.DataFrame([newData])], ignore_index=True)


        nomFeuille = "Dividendes"
        colonnesEuros = ["Dividendes brut", "Impôt à la source", "Dividendes net"]
        colonnesPourcentages = []
        colonnesDates = ["Date de valeur", "Dividende à la date du ..."]
        appliquerTableau = True
        
        return df, nomFeuille, colonnesEuros, colonnesPourcentages, colonnesDates, appliquerTableau

    def DataInteret(self, pdf: list, chemin: str) -> pd.DataFrame:
        """
        Extrait les données des fichiers PDF spécifiés et les retourne sous forme de DataFrame.

        Args:
            pdfFiles: Liste des noms de fichiers PDF à traiter.
            chemin: Chemin du répertoire contenant les fichiers PDF.

        Returns:
            pd.DataFrame: DataFrame contenant les données extraites de chaque fichier PDF.
        """
        # Assertions pour vérifier les types des entrées
        assert isinstance(pdf, list) and (fichier.endswith(".pdf") for fichier in pdf), \
            f"pdfFiles doit être une liste dont Le contenu contient des chaînes de caractères avec l'extension .pdf: {pdf}"
        assert isinstance(chemin, str) and os.path.isdir(chemin), f"Le dossier '{chemin}' n'existe pas"

        df = pd.DataFrame()

        for fichier in pdf:
            cheminComplet = os.path.join(chemin, fichier)  # Obtenez le chemin complet du fichier PDF

            # Assurez-vous que le fichier PDF existe
            assert os.path.isfile(cheminComplet), f"Le fichier {cheminComplet} n'existe pas."

            with pdfplumber.open(cheminComplet) as pdf:
                page = pdf.pages[0]
                text = page.extract_text()

            compteEspece = self.ExtraireDonnee(text, r"COMPTE-TITRES\s+(\d+)", 1)
            if compteEspece == None:
                compteEspece = self.ExtraireDonnee(text, r"(\d+)\s+RAPPORT D'INTÉRÊTS", 1)
            if compteEspece == None:
                raise ValueError("compteTitre non trouvée dans le texte.")

            iban = self.ExtraireDonnee(text, r"DE\d{20}", 0)

            listeMotifs = [
                r"(?<=ACTIFS NATURE DES REVENUS TAUX D'INTÉRÊTS TOTAL\n).+", 
                r"(?<=ACTIFS NATURE DES REVENUS TAUX D'INTÉRÊTS DATE TOTAL\n).+", 
                r"(?<=ACTIFS NATURE DES REVENUS TAUX D'INTÉRÊT DATE TOTAL\n).+"
            ]
            ligne = self.ExtraireDonnee(text, self.CreerExpressionReguliere(listeMotifs), 0).split()
            espece = ligne[0]
            interet = ligne[1]
            pourcentage = float(ligne[2][:-1].replace(',', '.')) / 100
            prix = float(ligne[-2].replace(',', '.'))
            plageDate = f"{ligne[3]} - {ligne[5]}" if len(ligne) > 5 else None

            dateEffet = datetime.strptime(self.ExtraireDonnee(text, r"(?<=IBAN DATE D'EFFET TOTAL\n).*\b\d{2}/\d{2}/\d{4}\b", 0).split()[1], "%d/%m/%Y").date()

            newData = {
                "Date d'effet": dateEffet,
                "Taux d'interêts": pourcentage,
                "Interêts net": prix,
                "Plage de date": plageDate,
                "Nature revenus": interet,
                "Actifs": espece,
                "Iban": iban,
                "Compte-espèces": compteEspece
            }

            # Ajouter la nouvelle ligne au DataFrame
            df = pd.concat([df, pd.DataFrame([newData])], ignore_index=True)


        nomFeuille = "Interêts"
        colonnesEuros = ["Interêts net"]
        colonnesPourcentages = ["Taux d'interêts"]
        colonnesDates = ["Date d'effet"]
        appliquerTableau = True

        return df, nomFeuille, colonnesEuros, colonnesPourcentages, colonnesDates, appliquerTableau

    def DataOrdresAchats(self, pdfFiles: list, chemin: str) -> pd.DataFrame:
        """
        Extrait les données des fichiers PDF spécifiés et les retourne sous forme de DataFrame.

        Args:
            pdfFiles: Liste des noms de fichiers PDF à traiter.
            chemin: Chemin du répertoire contenant les fichiers PDF.

        Returns:
            pd.DataFrame: DataFrame contenant les données extraites de chaque fichier PDF.
        """
        # Assertions pour vérifier les types des entrées
        assert isinstance(pdfFiles, list) and (fichier.endswith(".pdf") for fichier in pdfFiles), \
            f"pdfFiles doit être une liste dont Le contenu contient des chaînes de caractères avec l'extension .pdf: {pdfFiles}"
        assert isinstance(chemin, str) and os.path.isdir(chemin), f"Le dossier '{chemin}' n'existe pas"

        df = pd.DataFrame()

        for fichier in pdfFiles:
            cheminComplet = os.path.join(chemin, fichier)  # Obtenez le chemin complet du fichier PDF

            with pdfplumber.open(cheminComplet) as pdf:
                page = pdf.pages[0]
                text = page.extract_text()
            
            compteTitre = self.ExtraireDonnee(text, r"COMPTE-TITRES\s+(\d+)", 1)
            if compteTitre == None:
                compteTitre = self.ExtraireDonnee(text, r"(\d+)\s+CONFIRMATION DE L'INVESTISSEMENT PROGRAMMÉ", 1)
            if compteTitre == None:
                raise ValueError("compteTitre non trouvée dans le texte.")
            
            iban = self.ExtraireDonnee(text, r"DE\d{20}", 0)
            isin = self.ExtraireDonnee(text, r"(?<=ISIN : ).+", 0)
            dateExecution = datetime.strptime(self.ExtraireDonnee(text,r"(?<=DATE\s)(\d{2}/\d{2}/\d{4})", 0), "%d/%m/%Y").date()
            ticker = self.ExtraireDonnee(text, r"(POSITION|QUANTITÉ)[^\n]*\n([A-Za-zàâäéèêëîïôöùûü'\s&.-]+)", 2)
            titreDetenue = float(self.ExtraireDonnee(text, r'([-+]?\d+[\.,]?\d*)\s+(titre\(s\)|unit\.)', 1).replace(",", "."))
            montantInvesti = float(self.ExtraireDonnee(text, r"COMPTE-ESPÈCES DATE DE VALEUR MONTANT\s+.*\s+([-\d,]+)\s+EUR", 1).replace(",", "."))
            coursMoyen = float(self.ExtraireDonnee(text, r"POSITION QUANTITÉ COURS MOYEN MONTANT\s+.*?\s([\d,]+)\sEUR", 1).replace(",", "."))
            
            dateValeur = self.ExtraireDonnee(text, r"(?<=COMPTE-ESPÈCES DATE DE VALEUR MONTANT\n)\S+\s+(\d{2}/\d{2}/\d{4}|\d{4}-\d{2}-\d{2})", 1)
            if "/" in dateValeur:  # Pour le format DD/MM/YYYY
                dateValeur = datetime.strptime(dateValeur, "%d/%m/%Y").date()
            elif "-" in dateValeur:  # Pour le format YYYY-MM-DD
                dateValeur = datetime.strptime(dateValeur, "%Y-%m-%d").date()

            newData = {
                "Date d'exécution": dateExecution,
                "Date de valeur": dateValeur,
                "Ticker": ticker,
                "Montant investi": montantInvesti,
                "Titre(s) détenue(s)": titreDetenue,
                "Cours moyen": coursMoyen,
                "Isin": isin,
                "COMPTE-TITRE": compteTitre,
                "Iban": iban
            }

            # Ajouter la nouvelle ligne au DataFrame
            df = pd.concat([df, pd.DataFrame([newData])], ignore_index=True)


        # Un fichier PDF n'a pas pu être téléchargé sur Trade République c'est l'investissement d'Apple donc il faut le rajouter (supprimer pour votre portefeuille)*
        newData = {
            "Date d'exécution": datetime.strptime("2023/08/16", "%Y/%m/%d").date(),
            "Date de valeur": datetime.strptime("2023/08/18", "%Y/%m/%d").date(),
            "Ticker": "Apple Inc.",
            "Montant investi": -4.00,
            "Titre(s) détenue(s)": 0.02457,
            "Cours moyen": 162.80,
            "Isin": "US0378331005",
            "COMPTE-TITRE": "0700305101",
            "Iban": "DE83502109007011655011",
        }
        df = pd.concat([df, pd.DataFrame([newData])], ignore_index=True)

        nomFeuille = "Ordres d'Achats"
        colonnesEuros = ["Montant investi", "Cours moyen"]
        colonnesPourcentages = []
        colonnesDates = ["Date d'exécution", "Date de valeur"]
        appliquerTableau = True

        return df, nomFeuille, colonnesEuros, colonnesPourcentages, colonnesDates, appliquerTableau

    def DataOrdresVentes(self, pdfFiles: list, chemin: str) -> pd.DataFrame:
        """
        Extrait les données des fichiers PDF spécifiés et les retourne sous forme de DataFrame.

        Args:
            pdfFiles: Liste des noms de fichiers PDF à traiter.
            chemin: Chemin du répertoire contenant les fichiers PDF.

        Returns:
            pd.DataFrame: DataFrame contenant les données extraites de chaque fichier PDF.
        """
        # Assertions pour vérifier les types des entrées
        assert isinstance(pdfFiles, list) and (fichier.endswith(".pdf") for fichier in pdfFiles), \
            f"pdfFiles doit être une liste dont Le contenu contient des chaînes de caractères avec l'extension .pdf: {pdfFiles}"
        assert isinstance(chemin, str) and os.path.isdir(chemin), f"Le dossier '{chemin}' n'existe pas"

        df = pd.DataFrame()

        for fichier in pdfFiles:
            cheminComplet = os.path.join(chemin, fichier)  # Obtenez le chemin complet du fichier PDF
            with pdfplumber.open(cheminComplet) as pdf:
                page = pdf.pages[0]
                text = page.extract_text()

            
            compteTitre = self.ExtraireDonnee(text, r"COMPTE-TITRES\s+(\d+)", 1)
            if compteTitre == None:
                compteTitre = self.ExtraireDonnee(text, r"(\d+)\s+CONFIRMATION DE L'INVESTISSEMENT PROGRAMMÉ", 1)
            if compteTitre == None:
                raise ValueError("compteTitre non trouvée dans le texte.")

            dateValeur = self.ExtraireDonnee(text, r"(?<=COMPTE-ESPÈCES DATE DE VALEUR MONTANT\n)\S+\s+(\d{2}/\d{2}/\d{4}|\d{4}-\d{2}-\d{2})", 1)
            if "/" in dateValeur:  # Pour le format DD/MM/YYYY
                dateValeur = datetime.strptime(dateValeur, "%d/%m/%Y").date()
            elif "-" in dateValeur:  # Pour le format YYYY-MM-DD
                dateValeur = datetime.strptime(dateValeur, "%Y-%m-%d").date()
            
            iban = self.ExtraireDonnee(text, r"DE\d{20}", 0)
            isin = self.ExtraireDonnee(text, r"(?<=ISIN : ).+", 0)
            dateExecution = datetime.strptime(self.ExtraireDonnee(text,r"(?<=DATE\s)(\d{2}/\d{2}/\d{4})", 0), "%d/%m/%Y").date()
            ticker = self.ExtraireDonnee(text, r"(POSITION|QUANTITÉ)[^\n]*\n([A-Za-zàâäéèêëîïôöùûü'\s&.-]+)", 2)
            titreDetenue = float(self.ExtraireDonnee(text, r'([-+]?\d+[\.,]?\d*)\s+(titre\(s\)|unit\.)', 1).replace(",", "."))
            montantVendu = float(self.ExtraireDonnee(text, r"POSITION QUANTITÉ PRIX MONTANT\s+.*\s+([-\d,]+)\s+EUR", 1).replace(",", "."))
            prixTickerLorsDeLaVente = float(self.ExtraireDonnee(text, r"POSITION QUANTITÉ PRIX MONTANT\s+.*?\s([\d,]+)\sEUR", 1).replace(",", "."))
            montantGagne = float(self.ExtraireDonnee(text, r"COMPTE-ESPÈCES DATE DE VALEUR MONTANT\s+.*\s+([-\d,]+)\s+EUR", 1).replace(",", "."))

            frais = self.ExtraireDonnee(text, r"POSITION MONTANT\s+.*\s+([-\d,]+)\s+EUR", 1)
            frais = float(frais.replace(",", ".")) if frais else 0


            newData = {
                "Date d'exécution": dateExecution, 
                "Date de valeur": dateValeur,
                "Ticker": ticker,
                "Montant vendu": montantVendu,
                "Frais": frais,
                "Montant gagné": montantGagne,
                "Titre(s) vendu(s)": titreDetenue,
                "Prix du ticker lors de la vente": prixTickerLorsDeLaVente,
                "Isin": isin,
                "COMPTE-TITRE": compteTitre,
                "Iban": iban
            }

            # Ajouter la nouvelle ligne au DataFrame
            df = pd.concat([df, pd.DataFrame([newData])], ignore_index=True)


        nomFeuille = "Ordres de Ventes"
        colonnesEuros = ["Montant vendu", "Frais", "Montant gagné", "Prix du ticker lors de la vente"]
        colonnesPourcentages = []
        colonnesDates = ["Date d'exécution", "Date de valeur"]
        appliquerTableau = True

        return df, nomFeuille, colonnesEuros, colonnesPourcentages, colonnesDates, appliquerTableau

    def DataRetraitArgent(self, pdfFiles: list, chemin: str) -> pd.DataFrame:
        """
        Extrait les données des fichiers PDF spécifiés et les retourne sous forme de DataFrame.

        Args:
            pdfFiles: Liste des noms de fichiers PDF à traiter.
            chemin: chemin du répertoire contenant les fichiers PDF.

        Returns:
            pd.DataFrame: DataFrame contenant les données extraites de chaque fichier PDF.
        """
        # Assertions pour vérifier les types des entrées
        assert isinstance(pdfFiles, list) and (fichier.endswith(".pdf") for fichier in pdfFiles), \
            f"pdfFiles doit être une liste dont Le contenu contient des chaînes de caractères avec l'extension .pdf: {pdfFiles}"
        assert isinstance(chemin, str) and os.path.isdir(chemin), f"Le dossier '{chemin}' n'existe pas"

        df = pd.DataFrame()

        for fichier in pdfFiles:
            cheminComplet = os.path.join(chemin, fichier)
            with pdfplumber.open(cheminComplet) as pdf:
                page = pdf.pages[0]
                text = page.extract_text()


            nomAchat = self.ExtraireDonnee(text, r"POSITION COMMANDÉ LE QUANTITÉ\s+(.*)\s\d{2}\.\d{2}\.\d{4}", 1)
            dateCommande = datetime.strptime(self.ExtraireDonnee(text, r"POSITION COMMANDÉ LE QUANTITÉ\s+.*?\s(\d{2}\.\d{2}\.\d{4})", 1).replace(".", "/"), "%d/%m/%Y").date()
            quantite = float(self.ExtraireDonnee(text, r"POSITION COMMANDÉ LE QUANTITÉ\s+.*?\d{2}\.\d{2}\.\d{4}\s+(\d+)", 1).replace(",", "."))
            dateDuFichierPdf = datetime.strptime(self.ExtraireDonnee(text, r"(?<=DATE\s)(\d{2}\.\d{2}\.\d{4})", 0).replace(".", "/"), "%d/%m/%Y").date()
            numeroFacture = self.ExtraireDonnee(text, r"(?<=FACTURE N. ).+", 0)
            prixAchat = float(self.ExtraireDonnee(text, r"POSITION MONTANT\s+.*?\s([-–]?\d+,\d+)\s?€", 1).replace(",", "."))
            
            newData = {
                "Date de la commande": dateCommande,
                "Ce que l'on à acheté": nomAchat,
                "Prix d'achat": prixAchat,
                "Quantite": quantite,
                "Date de facture": dateDuFichierPdf,
                "Numéro de facture": numeroFacture
            }

            # Ajouter la nouvelle ligne au DataFrame
            df = pd.concat([df, pd.DataFrame([newData])], ignore_index=True)


        nomFeuille = "Retrait d'argent"
        colonnesEuros = ["Prix d'achat"]
        colonnesPourcentages = []
        colonnesDates = ["Date de la commande", "Date de facture"]
        appliquerTableau = True

        return df, nomFeuille, colonnesEuros, colonnesPourcentages, colonnesDates, appliquerTableau
    ############################################

    ########## ANNEXES ##########
    @staticmethod
    def ExtraireDonnee(texte: str, pattern: str, group: int) -> str|None:
        """
        Extrait une donnée spécifique d'un texte donné à l'aide d'un pattern regex.

        Args:
            texte (str): Le texte brut contenant les informations à extraire.
            pattern (str): Le pattern regex permettant de capturer la donnée souhaitée.
            group (int): Le numéro du groupe capturé à retourner.

        Returns:
            str: La donnée extraite correspondant au groupe spécifié dans le pattern.
        """
        assert isinstance(texte, str), "Le texte doit être une chaîne de caractères."
        assert isinstance(pattern, str), "Le pattern doit être une chaîne de caractères."
        assert isinstance(group, int) and group >= 0, "Le numéro de groupe doit être un entier positif ou zéro."

        match = re.search(pattern, texte)
        if not match:
            return None

        try:
            return match.group(group)
        except IndexError:
            raise IndexError(f"Le numéro de groupe {group} est invalide pour le pattern donné.")

    @staticmethod
    def CreerExpressionReguliere(listeMotifs: list[str]) -> str:
        """
        Construit une expression régulière combinée avec des 'OU' (|) à partir
        d'une liste de motifs donnés.

        Args:
            listeMotifs (list[str]): Liste de motifs d'expressions régulières.

        Returns:
            str: Une expression régulière combinée avec des 'OU'.
        """
        assert isinstance(listeMotifs, list), "listeMotifs doit être une liste."
        assert all(isinstance(motif, str) for motif in listeMotifs), "Tous les éléments de listeMotifs doivent être des chaînes de caractères."

        # Concaténer les motifs avec des 'OU' (|)
        expressionReguliereCombinee = '|'.join(listeMotifs)

        return expressionReguliereCombinee

    @staticmethod
    def ExtractInformationPdf(filePath: str) -> str:
        """
        Extrait le texte de chaque page d'un fichier PDF.

        Args:
            filePath: Chemin complet du fichier PDF dont le texte doit être extrait.

        Returns:
            str: Texte extrait de toutes les pages du fichier PDF. Retourne None si le fichier ne peut pas être lu.
        """
        # Assertions pour vérifier les types des entrées
        assert isinstance(filePath, str), f"filePath doit être une chaîne de caractères: {filePath}"

        # Initialiser une variable pour stocker le texte extrait
        text = ''

        try:
            with pdfplumber.open(filePath) as pdf:
                page = pdf.pages[0]
                text = page.extract_text()
            return text
        except Exception as e:
            print(f"Une erreur s'est produite lors de l'extraction du texte: {e}")
            return None
    
    @staticmethod
    def GetPdfFilesFromFolder(folderPath: str) -> list:
        """
        Retourne une liste de tous les fichiers PDF dans le dossier spécifié.

        Args:
            folderPath (str): Le chemin du dossier à explorer.

        Returns:
            list: Liste des noms de fichiers PDF (en str) présents dans le dossier.
        """
        # Vérification du type d'entrée
        assert isinstance(folderPath, str) and os.path.isdir(folderPath), f"Le dossier '{folderPath}' n'existe pas."

        # Récupération de la liste des fichiers .pdf dans le dossier
        filePdf = [file for file in os.listdir(folderPath) if file.lower().endswith('.pdf') and os.path.isfile(os.path.join(folderPath, file))]

        return filePdf
    #############################
    ############################################
