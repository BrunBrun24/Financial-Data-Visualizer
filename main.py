from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import xlrd

import tkinter as tk
from tkinter import filedialog, messagebox

from datetime import datetime, timedelta
import pandas as pd
import os
import re
import copy
import json

import plotly.graph_objects as go
import plotly.express as px
from plotly.subplots import make_subplots



def create_charts(data:dict, filename:str, compteCheque:bool = True) -> None:
    def diagramme_bar(df, title):
        # Regrouper les données par catégorie et type, puis sommer les montants
        df_grouped = df.groupby(['category', 'Type']).agg({'MONTANT': 'sum'}).reset_index()
        # Trier les données par ordre croissant des montants
        df_grouped = df_grouped.sort_values(by='MONTANT', ascending=False)
        # Calculer les pourcentages
        df_grouped['Total_Category'] = df_grouped.groupby('category')['MONTANT'].transform('sum')
        df_grouped['Percentage'] = (df_grouped['MONTANT'] / df_grouped['Total_Category']) * 100
        # Convertir les pourcentages en chaîne de caractères pour l'affichage
        df_grouped['Text'] = df_grouped.apply(lambda row: f"{row['Type']} ({row['Percentage']:.1f}%)", axis=1)

        # Créer le graphique en barres empilées
        fig = px.bar(
            df_grouped,
            x='category',
            y='MONTANT',
            color='Type',
            title=title,
            labels={'category': 'Catégories', 'MONTANT': 'Prix €'},
            text='Text'
        )

        # Ajuster les étiquettes pour qu'elles soient bien visibles
        fig.update_traces(texttemplate='%{text}', textposition='inside')

        fig.update_layout(
            showlegend=True,
            width=1800,
            height=900,
            margin=dict(
                l=200,  # marge gauche
                r=100,  # marge droite
                t=50,  # marge supérieure
                b=50   # marge inférieure
            )
        )
        
        return fig

    def diagramme_circulaire(df, name):
        # Préparation des données pour le graphique Sunburst
        labels = [name]
        parents = ['']
        values = [df['MONTANT'].sum()]

        # Ajouter les catégories principales
        for category in df['category'].unique():
            labels.append(category)
            parents.append(name)
            values.append(df[df['category'] == category]['MONTANT'].sum())

            # Ajouter les types sous chaque catégorie
            for type_op in df[df['category'] == category]['Type'].unique():
                labels.append(type_op)
                parents.append(category)
                values.append(df[(df['category'] == category) & (df['Type'] == type_op)]['MONTANT'].sum())

        # Création du graphique Sunburst
        fig = go.Figure(go.Sunburst(
            labels=labels,
            parents=parents,
            values=values,
            branchvalues='total',  # Utilisez 'total' ou 'remainder'
            textinfo='label+percent entry'  # Ajouter le texte pour les pourcentages
        ))

        # Mettre à jour la mise en page
        fig.update_layout(
            showlegend=True,
            width=1800,
            height=900,
            margin=dict(
                l=200,  # marge gauche
                r=100,  # marge droite
                t=50,  # marge supérieure
                b=50   # marge inférieure
            )
        )

        return fig

    def generate_sankey(data, title):
        def clean_transactions_and_group(data: dict, keys_to_remove: list = ["DATE D'OPÉRATION", "LIBELLÉ COURT", "TYPE OPÉRATION", "LIBELLÉ OPÉRATION"]):
            # Nettoyer les transactions
            for category, transactions_list in data.items():
                if transactions_list:
                    # Supprimer les clés non nécessaires
                    for transaction in transactions_list:
                        for key in keys_to_remove:
                            transaction.pop(key, None)  # Utilise pop avec un défaut pour éviter KeyError

            # Fonction pour calculer les sommes par type
            def calculate_totals(transactions_list):
                totals = {}
                for transaction in transactions_list:
                    type_op = transaction['Type']
                    montant = round(abs(transaction['MONTANT']), 2)
                    if type_op in totals:
                        totals[type_op] += montant
                    else:
                        totals[type_op] = montant
                return totals

            # Reformatage des données
            reformatted_data = {}
            for category, transactions_list in data.items():
                if transactions_list:
                    # Conserver les transactions distinctes
                    transactions_with_totals = [
                        {'Type': transaction['Type'], 'MONTANT': transaction['MONTANT']}
                        for transaction in transactions_list
                    ]
                    # Calculer les totaux pour chaque type
                    totals = calculate_totals(transactions_list)
                    category_total =  round(sum(totals.values()), 2)
                    reformatted_data[f'{category}: {category_total:.2f}'] = transactions_with_totals
                else:
                    reformatted_data[category] = []

            return reformatted_data

        def add_transactions(transactions, cle=None, depenses=True):
            def sum_amounts(data, depense=True):
                if not depense:
                    # Dictionnaire pour stocker les sommes par type
                    somme_par_type = {}
                    # Accumuler les montants par type
                    for transaction in data:
                        type_op = transaction['Type']
                        montant = transaction['MONTANT']
                        
                        if type_op in somme_par_type:
                            somme_par_type[type_op] += montant
                        else:
                            somme_par_type[type_op] = montant

                    return somme_par_type
                else:
                    # Initialisation d'un dictionnaire pour stocker les sommes
                    sums = {}
                    # Parcourir chaque catégorie dans les données
                    for category, transactions in data.items():
                        total = 0
                        for transaction in transactions:
                            total += transaction['MONTANT']
                        sums[category] = total
                    
                    return sums
        
            def get_label_index(label):
                nonlocal next_index
                if label not in label_dict:
                    label_dict[label] = next_index
                    labels.append(label)
                    next_index += 1
                return label_dict[label]
        
            def autre(data):
                next_index = 0
                 # Remplir les labels et les flux
                for category, transactions in data.items():
                    # Ajouter la catégorie source si elle n'existe pas encore
                    if category not in label_dict:
                        label_dict[category] = next_index
                        labels.append(category)
                        next_index += 1
                    
                    for transaction in transactions:
                        # Ajouter le type de dépense cible si il n'existe pas encore
                        transaction_type = transaction['Type']
                        if transaction_type not in label_dict:
                            label_dict[transaction_type] = next_index
                            labels.append(transaction_type)
                            next_index += 1
                        
                        # Ajouter le lien (flux)
                        sources.append(label_dict[category])
                        targets.append(label_dict[transaction_type])
                        values.append(abs(transaction['MONTANT']))


            if not depenses:
                if isinstance(transactions, list):
                    for transaction in transactions:
                        target_name = transaction['Type']
                        
                        # Rajouter le montant total au Type
                        argent = 0
                        for trans in transactions:
                            if trans['Type'] == target_name:
                                argent += trans["MONTANT"]
                        target_name += f": {str(abs(round(argent, 2)))}"


                        source_idx = get_label_index(target_name)
                        target_idx = get_label_index(cle)
                        
                        sources.append(source_idx)
                        targets.append(target_idx)
                        values.append(round(abs(transaction['MONTANT']), 2))
                else:
                    autre(transactions)
            else:
                # Catégories
                for category, montant in sum_amounts(transactions).items():
                    if montant < 0:
                        target_name = category
                        main_source_idx = 1
                        source_idx = get_label_index(target_name)
                        
                        sources.append(main_source_idx)
                        targets.append(source_idx)
                        values.append(abs(round(montant, 2)))

                # Sous-catégories
                for sous_category, transaction in transactions.items():
                    if transaction:
                        for ele in transaction:
                            target_name = ele['Type']
                            
                            # Rajouter le montant total au Type
                            argent = 0
                            for trans in transaction:
                                if trans['Type'] == target_name:
                                    argent += trans["MONTANT"]
                            target_name += f": {str(abs(round(argent, 2)))}"

                            source_idx = get_label_index(sous_category)
                            target_idx = get_label_index(target_name)
                            
                            sources.append(source_idx)
                            targets.append(target_idx)
                            values.append(abs(round(ele['MONTANT'], 2)))

        
        
        # Initialisation des listes pour les labels, les sources, les cibles et les valeurs
        labels = []
        sources = []
        targets = []
        values = []

        # Créer un dictionnaire pour l'index des labels
        label_dict = {}
        next_index = 0

        data = clean_transactions_and_group(data)

        pattern = r'Revenus: -?\d+(?:,\d{3})*(?:\.\d+)?'  # Exemple d'un pattern qui match 'Revenus: (123)'
        # Traiter la catégorie "Revenus"
        cle = ""
        for key in data.keys():
            if re.match(pattern, key):
                cle = key
                add_transactions(data[cle], cle, False)
                data.pop(cle)
                break

        if cle == "":
            add_transactions(data, None, False)
        else:
            add_transactions(data)


        # Création du graphique de Sankey
        fig = go.Figure(data=[go.Sankey(
            node=dict(
                pad=15,
                thickness=20,
                line=dict(color='black', width=0.5),
                label=labels
            ),
            link=dict(
                source=sources,
                target=targets,
                value=values
            )
        )])

        fig.update_layout(
            # title_text=title,
            font_size=10,
            width=1800,
            height=900,
            margin=dict(
                l=100,  # marge gauche
                r=100,  # marge droite
                t=50,  # marge supérieure
                b=50   # marge inférieure
            )
        )

        return fig



    # Convertir les dictionnaires en DataFrames
    dfs = {key: pd.DataFrame(value) for key, value in data.items() if value}
    df_revenus = dfs.get('Revenus', pd.DataFrame())
    if not df_revenus.empty:
        df_revenus["MONTANT"] *= -1
        
    df_depenses = pd.DataFrame()

    # Récupérer l'ensemble des dépenses
    for category, df in dfs.items():
        # Ajouter les catégories
        df["category"] = category
        df["MONTANT"] *= -1

        # Si la colonne 'Type' est absente, ajouter une colonne 'Type' avec la category
        if "Type" not in df.columns:
            df["Type"] = category

        if category != "Revenus":
            df_depenses = pd.concat([df_depenses, df], ignore_index=True)


    if not compteCheque:
        # Si on a des revenus mais pas de dépenses
        if df_depenses.empty and not df_revenus.empty:
            fig_sankey = generate_sankey(data=data, title="Diagramme de Sankey des transactions")
            fig_soleil_revenus = diagramme_circulaire(df=df_revenus, name="Revenus du Livret A")
        # Si on a des des dépenses mais pas de revenus
        elif not df_depenses.empty and df_revenus.empty:
            fig_sankey = generate_sankey(data=data, title="Diagramme de Sankey des transactions")
            fig_soleil_depenses = diagramme_circulaire(df=df_depenses, name="Virement")
        # Si on a des des dépenses et des revenus
        elif not df_depenses.empty and not df_revenus.empty:
            fig_sankey = generate_sankey(data=data, title="Diagramme de Sankey des transactions")
            fig_soleil_depenses = diagramme_circulaire(df=df_depenses, name="Virement")
            fig_soleil_revenus = diagramme_circulaire(df=df_revenus, name="Revenus du Livret A")

    elif not df_revenus.empty:
        # Il faut équilibrer l'argent mit dans le Livret A en fonction de l'rgent ajouté via  "Virement interne"
        argent_retirer_LivretA = sum(df_revenus[df_revenus['Type'] == 'Virement interne']["MONTANT"])
        argent_ajoute_LivretA = sum(df_depenses[df_depenses['Type'] == 'Livret A']["MONTANT"])

        if (argent_ajoute_LivretA - argent_retirer_LivretA) <= 0:
            df_depenses = df_depenses[df_depenses['Type'] != 'Livret A']
        elif (argent_ajoute_LivretA - argent_retirer_LivretA) >= 0:
            difference = argent_ajoute_LivretA - argent_retirer_LivretA
            # Filtrer les lignes correspondant à 'Livret A'
            livret_a_entries = df_depenses[df_depenses['Type'] == 'Livret A']
            # Réinitialiser l'index de livret_a_entries
            livret_a_entries = livret_a_entries.reset_index(drop=True)
            
            # Retirer toutes les entrées sauf la première
            df_depenses = df_depenses[df_depenses['Type'] != 'Livret A']
            
            # Ajuster le montant de la première entrée restante
            livret_a_entries.at[0, 'MONTANT'] = difference
            df_depenses = pd.concat([df_depenses, livret_a_entries.iloc[[0]]], ignore_index=True)
            

        fig_bar = diagramme_bar(df=df_depenses, title="Dépenses par catégorie")

        fig_sankey = generate_sankey(data=data, title="Diagramme de Sankey des transactions")


        df_filtré = df_revenus[df_revenus['Type'] != 'Virement interne']
        fig_soleil_revenus = diagramme_circulaire(df=df_filtré, name="Revenus gagné")
        fig_soleil_allRevenus = diagramme_circulaire(df=df_revenus, name="Revenus gagné + Virement interne")
        
        
        # Filtrer les lignes où la colonne 'category' n'est pas 'Investissement'
        df_filtré = df_depenses[df_depenses['category'] != 'Investissement']
        fig_soleil_depenses = diagramme_circulaire(df=df_filtré, name="Dépenses")
        fig_soleil = diagramme_circulaire(df=df_depenses, name="Dépenses + Investissement")
        
    else:
        fig_bar = diagramme_bar(df=df_depenses, title="Dépenses par catégorie")
        
        # Filtrer les lignes où la colonne 'category' n'est pas 'Investissement'
        df_filtré = df_depenses[df_depenses['category'] != 'Investissement']
        fig_soleil_depenses = diagramme_circulaire(df=df_filtré, name="Dépenses")
        fig_soleil = diagramme_circulaire(df=df_depenses, name="Dépenses + Investissement")


    # Crée un fichier HTML pour enregistrer les graphiques
    with open(filename, 'w') as f:
        # Sauvegarder la figure combinée dans un fichier HTML
        if not compteCheque:

            
            # Si on a des revenus mais pas de dépenses
            if df_depenses.empty and not df_revenus.empty:
                fig_sankey.write_html(f, include_plotlyjs='cdn')
                fig_soleil_revenus.write_html(f, include_plotlyjs='cdn')
            # Si on a des des dépenses mais pas de revenus
            elif not df_depenses.empty and df_revenus.empty:
                fig_sankey.write_html(f, include_plotlyjs='cdn')
                fig_soleil_depenses.write_html(f, include_plotlyjs='cdn')
             # Si on a des des dépenses et des revenus
            elif not df_depenses.empty and not df_revenus.empty:
                fig_sankey.write_html(f, include_plotlyjs='cdn')
                fig_soleil_depenses.write_html(f, include_plotlyjs='cdn')
                fig_soleil_revenus.write_html(f, include_plotlyjs='cdn')

        elif not df_revenus.empty:
            # Crée une figure avec des sous-graphes pour afficher côte à côte
            fig_combined = make_subplots(
                rows=1, cols=2,
                # subplot_titles=("Montant total par catégorie et type", "Types de dépenses"),
                specs=[[{"type": "sunburst"}, {"type": "sunburst"}]]
            )

            # Ajouter les graphiques à la figure combinée
            for trace in fig_soleil_depenses.data:
                fig_combined.add_trace(trace, row=1, col=1)
            for trace in fig_soleil.data:
                fig_combined.add_trace(trace, row=1, col=2)

            # Mettre à jour la mise en page pour la figure combinée
            fig_combined.update_layout(
                showlegend=True,
                width=1800,
                height=900,
                margin=dict(
                    l=200,  # marge gauche
                    r=100,  # marge droite
                    t=50,  # marge supérieure
                    b=50   # marge inférieure
                )
            )


            # Crée une figure avec des sous-graphes pour afficher côte à côte
            fig_combined_revenus = make_subplots(
                rows=1, cols=2,
                # subplot_titles=("Montant total par catégorie et type", "Types de dépenses"),
                specs=[[{"type": "sunburst"}, {"type": "sunburst"}]]
            )

            # Ajouter les graphiques à la figure combinée
            for trace in fig_soleil_revenus.data:
                fig_combined_revenus.add_trace(trace, row=1, col=1)
            for trace in fig_soleil_allRevenus.data:
                fig_combined_revenus.add_trace(trace, row=1, col=2)

            # Mettre à jour la mise en page pour la figure combinée
            fig_combined_revenus.update_layout(
                showlegend=True,
                width=1800,
                height=900,
                margin=dict(
                    l=200,  # marge gauche
                    r=100,  # marge droite
                    t=50,  # marge supérieure
                    b=50   # marge inférieure
                )
            )

            
            fig_sankey.write_html(f, include_plotlyjs='cdn')
            fig_combined.write_html(f, include_plotlyjs='cdn')
            fig_bar.write_html(f, include_plotlyjs='cdn')
            fig_combined_revenus.write_html(f, include_plotlyjs='cdn')

        else:
            # Crée une figure avec des sous-graphes pour afficher côte à côte
            fig_combined = make_subplots(
                rows=1, cols=2,
                # subplot_titles=("Montant total par catégorie et type", "Types de dépenses"),
                specs=[[{"type": "sunburst"}, {"type": "sunburst"}]]
            )

            # Ajouter les graphiques à la figure combinée
            for trace in fig_soleil_depenses.data:
                fig_combined.add_trace(trace, row=1, col=1)
            for trace in fig_soleil.data:
                fig_combined.add_trace(trace, row=1, col=2)

            # Mettre à jour la mise en page pour la figure combinée
            fig_combined.update_layout(
                showlegend=True,
                width=1800,
                height=900,
                margin=dict(
                    l=200,  # marge gauche
                    r=100,  # marge droite
                    t=50,  # marge supérieure
                    b=50   # marge inférieure
                )
            )

            fig_combined.write_html(f, include_plotlyjs='cdn')
            fig_bar.write_html(f, include_plotlyjs='cdn')



def choisir_et_extraire_donnees_excel():
    """
    Ouvre une fenêtre avec un message d'instruction et un bouton permettant
    de rechercher un fichier Excel. Retourne les données du fichier sélectionné.
    """
    def extraire_donnees_excel(nameFile:str):
        def excel_date_to_datetime(excel_date):
            return datetime(1899, 12, 30) + timedelta(days=excel_date)
        
        data = []  # Liste pour stocker les données de tous les fichiers
        classeur = xlrd.open_workbook(nameFile) # Ouvrir le fichier Excel avec xlrd
        feuille = classeur.sheet_by_index(0)  # Sélectionnez la première feuille de calcul

        # Liste pour stocker les données extraites d'un fichier
        donnees = []

        # Parcourez les lignes de la feuille de calcul à partir de la ligne 3
        for row_index in range(2, feuille.nrows):  # 2 correspond à la ligne 3
            ligne = feuille.row_values(row_index, start_colx=0, end_colx=5)  # Les colonnes A à E (indices 0 à 4)
            donnees.append(ligne)

        data.extend(donnees)  # Ajoutez les données extraites de ce fichier à la liste totale

        data = data[1:]

        for line in data:
            if line[2] == "FACTURE CARTE":
                # On prend la date du jour de paiment
                line[0] = line[3][17:23]
                line[0] = line[0][0:2] + "/" + line[0][2:4] + "/20" + line[0][4:]

                # Convertir la date en objet datetime
                date_obj = datetime.strptime(line[0], '%d/%m/%Y')
                # Convertir l'objet datetime en nombre de série de date Excel
                line[0] = float((date_obj - datetime(1899, 12, 30)).days)

                index_debut = 24
                index_fin = line[3].find("CARTE", index_debut)
                if index_fin != -1:
                    line[3] = line[3][index_debut:index_fin]
                    line[3] = ' '.join(line[3].split())

            elif line[2] == "VIR CPTE A CPTE EMIS":
                index_debut = 22
                index_fin = line[3].find("/", index_debut)
                if index_fin != -1:
                    line[3] = line[3][index_debut:index_fin]
                    line[3] = ' '.join(line[3].split())

            elif line[2] == "VIR CPTE A CPTE RECU":
                index_debut = 22
                index_fin = line[3].find("/REF", index_debut)
                if index_fin != -1:
                    line[3] = line[3][index_debut:index_fin]
                    line[3] = ' '.join(line[3].split())

            elif line[2] == "VIR SEPA RECU":
                index_debut = 15
                index_fin = line[3].find("/REF", index_debut)
                if index_fin != -1:
                    line[3] = line[3][index_debut:index_fin]
                    line[3] = ' '.join(line[3].split())

            elif line[2] == "REMISE CHEQUES":
                index_debut = 15
                index_fin = line[3].find("/NOPT", index_debut)
                if index_fin != -1:
                    line[3] = line[3][index_debut:index_fin]
                    line[3] = ' '.join(line[3].split())


        # Tri les données de la date la plus ancienne à la plus récente
        data.sort(key=lambda x: x[0])

        
        resultat = pd.DataFrame(data, columns=["DATE D'OPÉRATION", "LIBELLÉ COURT", "TYPE OPÉRATION", "LIBELLÉ OPÉRATION", "MONTANT"])
        # Appliquer la conversion à la colonne DATE D'OPÉRATION
        resultat["DATE D'OPÉRATION"] = resultat["DATE D'OPÉRATION"].apply(excel_date_to_datetime)
        # Convertir en format Timestamp
        resultat["DATE D'OPÉRATION"] = pd.to_datetime(resultat["DATE D'OPÉRATION"])

        return resultat

    def load_dict_from_json(file_path):
        """Lit un fichier JSON et le convertit en dictionnaire.
        
        Args:
            file_path (str): Le chemin du fichier JSON à lire.
            
        Returns:
            dict: Le dictionnaire lu à partir du fichier JSON.
        """
        with open(file_path, 'r', encoding='utf-8') as file:
            data_dict = json.load(file)

        return data_dict


    def ouvrir_fichier():
        """Fonction pour ouvrir la boîte de dialogue de sélection de fichier"""
        nonlocal fichier
        fichier = filedialog.askopenfilename(
            title="Choisissez un fichier Excel ou JSON",
            filetypes=[("Fichiers Excel et JSON", "*.xlsx *.xls *.json")],
            initialdir=dossier_initial  # Utiliser le dossier prédéfini
        )
        if not fichier:
            return
        # Ferme la fenêtre après la sélection du fichier
        root.destroy()

    def rechercher_fichier():
        """Fonction appelée lorsqu'on clique sur le bouton "Rechercher"""
        ouvrir_fichier()


    fichier = None  # Variable pour stocker le chemin du fichier sélectionné

    # Dossier prédéfini pour ouvrir la boîte de dialogue
    dossier_initial = "Bilan\Archives"  # Remplace par le chemin vers ton dossier

    # Création de la fenêtre principale
    root = tk.Tk()
    root.title("Sélection du Fichier")

    # Définir la taille de la fenêtre
    window_width = 400
    window_height = 200

    # Centrer la fenêtre sur l'écran
    center_window(root, window_width, window_height)

    # Création d'un message d'instruction
    message_label = tk.Label(root, text="Veuillez sélectionner un fichier Excel pour extraire les données.", padx=20, pady=20)
    message_label.pack()

    # Création du bouton pour rechercher le fichier
    search_button = tk.Button(root, text="Rechercher un fichier", command=rechercher_fichier)
    search_button.pack(pady=10)

    # Lancer la boucle principale de l'application
    root.mainloop()

    # Lire les données du fichier après la fermeture de la fenêtre
    if fichier:
        try:
            if fichier.lower().endswith('.xlsx') or fichier.lower().endswith('.xls'):
                data = extraire_donnees_excel(fichier)
                return data, "Excel", fichier.split('/')[-2]
            elif fichier.lower().endswith('.json'):
                data = load_dict_from_json(fichier)
                return data, "Json", fichier.split('/')[-2]
            else:
                messagebox.showerror("Erreur", "Type de fichier non supporté.")
                return None

        except Exception as e:
            messagebox.showerror("Erreur", f"Une erreur s'est produite lors de la lecture du fichier : {str(e)}")
    return None

def trier_donnees(donnees):
    """
    Trie les données en utilisant les valeurs uniques de la colonne 'LIBELLÉ COURT' comme clés pour des DataFrames distincts.
    
    Args:
        donnees: DataFrame contenant les transactions bancaires avec une colonne 'LIBELLÉ COURT'.
        
    Returns:
        tableaux: Dictionnaire de DataFrames, où chaque clé est une valeur unique de 'LIBELLÉ COURT'.
    """
    # Créer un dictionnaire vide pour stocker les DataFrames triés
    tableaux = {}

    # Obtenir les valeurs uniques de 'LIBELLÉ COURT'
    valeurs_unique = donnees["LIBELLÉ COURT"].unique()

    # Remplir les tableaux en fonction de la colonne "LIBELLÉ COURT"
    for valeur in valeurs_unique:
        tableaux[valeur] = donnees[donnees["LIBELLÉ COURT"] == valeur]
    
    return tableaux


def center_window(root, width, height):
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    x = (screen_width // 2) - (width // 2)
    y = (screen_height // 2) - (height // 2)
    root.geometry(f'{width}x{height}+{x}+{y}')

def trier_elements(dictionnaire):
    def reorder_keys(item, indice):
        # Obtenir les clés de l'élément
        keys = list(item.keys())
        # Si la clé 'Type' existe et n'est pas déjà en 3e position
        if 'Type' in keys and keys.index('Type') != indice:
            keys.remove('Type')
            keys.insert(indice, 'Type')
            # Recréer l'élément avec l'ordre des clés modifié
            return {k: item[k] for k in keys}
        else:
            return item
    
    # Parcourir chaque catégorie du dictionnaire
    for category in dictionnaire:
        # Traiter chaque élément de la liste
        dictionnaire[category] = [reorder_keys(item, 1) for item in dictionnaire[category]]

    return dictionnaire

def comparer_et_categoriser_operations(cheminFileJson, operationsActuelles):
    # Charger les opérations déjà présentes dans le fichier JSON
    if os.path.exists(cheminFileJson):
        with open(cheminFileJson, 'r', encoding="UTF-8") as fichier:
            operations_existees = json.load(fichier)
    else:
        return None
    
    
    # Initialiser une liste pour stocker les DataFrames
    dataframes = []
    # Parcourir chaque catégorie
    for categorie, operations in operations_existees.items():
        if operations:  # Vérifie si la liste n'est pas vide
            # Convertir la liste en DataFrame
            df = pd.DataFrame(operations)
            df['Catégorie'] = categorie  # Ajouter une colonne pour la catégorie
            dataframes.append(df)  # Ajouter le DataFrame à la liste

    if dataframes:
        # Concaténer tous les DataFrames en un seul
        df_complet = pd.concat(dataframes, ignore_index=True)
    else:
        return None

    df = pd.DataFrame([operationsActuelles])

    # On utilise merge pour voir si la ligne existe dans le df_source
    merged_df = pd.merge(df_complet, df, how='inner')

    if not merged_df.empty:
        # Extraire la première ligne trouvée (s'il y en a plusieurs, ajustez selon besoin)
        result_row = merged_df.iloc[0]
        return [result_row['Catégorie'], result_row['Type']]
    else:
        return None



def afficher_fenetre_avec_boutons(categories, tableaux, button_labels, cheminFileJson):
    def update_display():
        nonlocal index, category_index
        
        if category_index < len(categories):
            current_category = categories[category_index]
            current_list = listes_types_operation[current_category]
            if index < len(current_list):
                current_item = current_list[index]

                # Comparer et catégoriser les opérations avec celles du fichier JSON
                operations_categorisees = comparer_et_categoriser_operations(cheminFileJson, current_item)
                
                if operations_categorisees is not None:
                    main_category = operations_categorisees[0]
                    sous_category = operations_categorisees[1]
                    if main_category in button_labels and sous_category in button_labels[main_category]:
                        current_item['Type'] = sous_category
                        if main_category not in results:
                            results[main_category] = []
                        results[main_category].append(current_item)
                        history.append((index, category_index, current_item, main_category))  # Ajouter à l'historique
                    else:
                        print(f"La catégorie {main_category} ou la sous-catégorie {sous_category} n'existe pas.")

                    # Passer à l'entrée suivante
                    index += 1
                    if index >= len(current_list):
                        index = 0
                        category_index += 1
                    update_display()
                    return
                else:
                    current_item = current_list[index]
                    montant = current_item["MONTANT"]
                    
                    #########################################################################################################
                    libelle_court = current_item["LIBELLÉ COURT"]
                    libelle_operation = current_item["LIBELLÉ OPÉRATION"]
                    restaurant = ["BURGER KING", "KFC ANGERS", "MC DO", "O TACOS", "OTACOS"]

                    # Regarder si c'est un chèque
                    if libelle_court == "REMISE CHEQUES":
                        # Vérifier si la catégorie 'Revenus' et la sous-catégorie 'Chèque reçu' existent
                        if "Revenus" in button_labels and "Chèque reçu" in button_labels["Revenus"]:
                            # Ajouter directement à 'Revenus' sous 'Chèque reçu'
                            current_item['Type'] = "Chèque reçu"
                            if "Revenus" not in results:
                                results["Revenus"] = []
                            results["Revenus"].append(current_item)
                            history.append((index, category_index, current_item, "Revenus"))  # Ajouter à l'historique
                        else:
                            # Afficher un message ou gérer le cas où 'Revenus' ou 'Chèque reçu' n'existe pas
                            print("La catégorie 'Revenus' ou la sous-catégorie 'Chèque reçu' n'existe pas.")
                        
                        # Passer à l'entrée suivante
                        index += 1
                        if index >= len(current_list):
                            index = 0
                            category_index += 1
                        update_display()
                        return  # Sortir de la fonction après avoir traité un chèque
                    #########################################################################################################
                    elif (libelle_operation == "DE AUBRUN /MOTIF VIREMENT PAUL") or (libelle_court == "VIREMENT INSTANTANE RECU") :
                        # Vérifier si la catégorie 'Revenus' et la sous-catégorie 'Virement reçu' existent
                        if "Revenus" in button_labels and "Virement reçu" in button_labels["Revenus"]:
                            # Ajouter directement à 'Revenus' sous 'Virement reçu'
                            current_item['Type'] = "Virement reçu"
                            if "Revenus" not in results:
                                results["Revenus"] = []
                            results["Revenus"].append(current_item)
                            history.append((index, category_index, current_item, "Revenus"))  # Ajouter à l'historique
                        else:
                            # Afficher un message ou gérer le cas où 'Revenus' ou 'Virement reçu' n'existe pas
                            print("La catégorie 'Revenus' ou la sous-catégorie 'Virement reçu' n'existe pas.")
                        
                        # Passer à l'entrée suivante
                        index += 1
                        if index >= len(current_list):
                            index = 0
                            category_index += 1
                        update_display()
                        return  # Sortir de la fonction après avoir traité un chèque
                    #########################################################################################################
                    elif bool(re.match(r'^DE AUBRUN PAUL EMIL', libelle_operation)) or bool(re.match(r'^DE MR PAUL AUBRUN', libelle_operation)):
                        main_category = "Revenus"
                        sous_category = "Virement interne"
                        if main_category in button_labels and sous_category in button_labels[main_category]:
                            current_item['Type'] = sous_category
                            if main_category not in results:
                                results[main_category] = []
                            results[main_category].append(current_item)
                            history.append((index, category_index, current_item, main_category))  # Ajouter à l'historique
                        else:
                            print(f"La catégorie '{main_category}' ou la sous-catégorie '{sous_category}' n'existe pas.")

                        # Passer à l'entrée suivante
                        index += 1
                        if index >= len(current_list):
                            index = 0
                            category_index += 1
                        update_display()
                        return  # Sortir de la fonction après avoir traité un chèque
                    #########################################################################################################
                    elif libelle_court == "COMMISSIONS":
                        main_category = "Banque"
                        sous_category = "Frais bancaires"
                        if main_category in button_labels and sous_category in button_labels[main_category]:
                            current_item['Type'] = sous_category
                            if main_category not in results:
                                results[main_category] = []
                            results[main_category].append(current_item)
                            history.append((index, category_index, current_item, main_category))  # Ajouter à l'historique
                        else:
                            print(f"La catégorie {main_category} ou la sous-catégorie {sous_category} n'existe pas.")
                        
                        # Passer à l'entrée suivante
                        index += 1
                        if index >= len(current_list):
                            index = 0
                            category_index += 1
                        update_display()
                        return
                    #########################################################################################################
                    elif bool(re.match(r'^TRADE REPUBLIC', libelle_operation)) and montant < 0:
                        main_category = "Investissement"
                        sous_category = "CTO"
                        if main_category in button_labels and sous_category in button_labels[main_category]:
                            current_item['Type'] = sous_category
                            if main_category not in results:
                                results[main_category] = []
                            results[main_category].append(current_item)
                            history.append((index, category_index, current_item, main_category))  # Ajouter à l'historique
                        else:
                            print(f"La catégorie '{main_category}' ou la sous-catégorie '{sous_category}' n'existe pas.")
                        
                        # Passer à l'entrée suivante
                        index += 1
                        if index >= len(current_list):
                            index = 0
                            category_index += 1
                        update_display()
                        return  # Sortir de la fonction après avoir traité un chèque
                    #########################################################################################################
                    elif libelle_court == "VIREMENT INTERNE" and montant < 0:
                        main_category = "Investissement"
                        sous_category = "Livret A"
                        if main_category in button_labels and sous_category in button_labels[main_category]:
                            current_item['Type'] = sous_category
                            if main_category not in results:
                                results[main_category] = []
                            results[main_category].append(current_item)
                            history.append((index, category_index, current_item, main_category))  # Ajouter à l'historique
                        else:
                            print(f"La catégorie '{main_category}' ou la sous-catégorie '{sous_category}' n'existe pas.")
                        
                        # Passer à l'entrée suivante
                        index += 1
                        if index >= len(current_list):
                            index = 0
                            category_index += 1
                        update_display()
                        return
                    #########################################################################################################
                    elif any(restaurant_item in libelle_operation for restaurant_item in restaurant):
                        main_category = "Loisir et sorties"
                        sous_category = "Restaurants - Bars"
                        if main_category in button_labels and sous_category in button_labels[main_category]:
                            current_item['Type'] = sous_category
                            if main_category not in results:
                                results[main_category] = []
                            results[main_category].append(current_item)
                            history.append((index, category_index, current_item, main_category))  # Ajouter à l'historique
                        else:
                            print(f"La catégorie '{main_category}' ou la sous-catégorie '{sous_category}' n'existe pas.")

                        # Passer à l'entrée suivante
                        index += 1
                        if index >= len(current_list):
                            index = 0
                            category_index += 1
                        update_display()
                        return
                    #########################################################################################################
                    elif any(stationCarburant in libelle_operation for stationCarburant in ["DAC", "STATION U ST GEORGES S"]):
                        main_category = "Transports et véhicules"
                        sous_category = "Carburant"
                        if main_category in button_labels and sous_category in button_labels[main_category]:
                            current_item['Type'] = sous_category
                            if main_category not in results:
                                results[main_category] = []
                            results[main_category].append(current_item)
                            history.append((index, category_index, current_item, main_category))  # Ajouter à l'historique
                        else:
                            print(f"La catégorie '{main_category}' ou la sous-catégorie '{sous_category}' n'existe pas.")

                        # Passer à l'entrée suivante
                        index += 1
                        if index >= len(current_list):
                            index = 0
                            category_index += 1
                        update_display()
                        return
                    #########################################################################################################
                    elif libelle_operation == "IZLY SMONEY PARIS":
                        main_category = "Vie quotidienne"
                        sous_category = "Alimentation - Supermarché"
                        if main_category in button_labels and sous_category in button_labels[main_category]:
                            current_item['Type'] = sous_category
                            if main_category not in results:
                                results[main_category] = []
                            results[main_category].append(current_item)
                            history.append((index, category_index, current_item, main_category))  # Ajouter à l'historique
                        else:
                            print(f"La catégorie {main_category} ou la sous-catégorie {sous_category} n'existe pas.")
                        
                        # Passer à l'entrée suivante
                        index += 1
                        if index >= len(current_list):
                            index = 0
                            category_index += 1
                        update_display()
                        return
                    #########################################################################################################
                    elif libelle_court == "VIREMENT PERMANENT":
                        main_category = "Revenus"
                        sous_category = "Virement reçu"
                        if main_category in button_labels and sous_category in button_labels[main_category]:
                            current_item['Type'] = sous_category
                            if main_category not in results:
                                results[main_category] = []
                            results[main_category].append(current_item)
                            history.append((index, category_index, current_item, main_category))  # Ajouter à l'historique
                        else:
                            print(f"La catégorie {main_category} ou la sous-catégorie {sous_category} n'existe pas.")
                        
                        # Passer à l'entrée suivante
                        index += 1
                        if index >= len(current_list):
                            index = 0
                            category_index += 1
                        update_display()
                        return


                    
                    display_label.config(text=current_item["LIBELLÉ OPÉRATION"] + "\n\n" + str(montant) + "€")

                    # Afficher les boutons selon la valeur du montant
                    if montant >= 0:
                        filtered_buttons = {"Revenus": button_labels["Revenus"]}
                    else:
                        filtered_buttons = {key: value for key, value in button_labels.items() if key != "Revenus"}
                        
                    update_buttons(filtered_buttons)
            else:
                index = 0
                category_index += 1
                update_display()
        else:
            display_label.config(text="Toutes les chaînes ont été affichées")
            root.destroy()


    def button_clicked(button_name):
        if isinstance(button_labels[button_name], list) and button_labels[button_name]:
            for widget in root.winfo_children():
                if widget != display_label:
                    widget.destroy()
            for i, sub_label in enumerate(button_labels[button_name]):
                button = tk.Button(root, text=sub_label, command=lambda l=sub_label: sub_button_clicked(l, button_name))
                button.grid(row=i // buttons_per_row, column=i % buttons_per_row, padx=10, pady=10, sticky='ew')
            # Ajouter le bouton "Retour" et "Suivant"
            return_button = tk.Button(root, text="Retour", command=reset_to_main_buttons)
            return_button.grid(row=(len(button_labels[button_name]) // buttons_per_row) + 10, column=0, columnspan=buttons_per_row-2, pady=10, sticky='ew')
            skip_button = tk.Button(root, text="Suivant", command=skip_entry)
            skip_button.grid(row=(len(button_labels[button_name]) // buttons_per_row) + 10, column=buttons_per_row-1, pady=10, sticky='ew')
        else:
            process_button(button_name)

    def sub_button_clicked(sub_button_name, parent_button_name):
        process_sub_button(sub_button_name, parent_button_name)
        root.after(100, reset_to_main_buttons)

    def process_sub_button(sub_button_name, parent_button_name):
        nonlocal index, category_index
        current_category = categories[category_index]
        current_list = listes_types_operation[current_category]
        if index < len(current_list):
            current_row = current_list[index]
            if parent_button_name not in results:
                results[parent_button_name] = []
            current_row['Type'] = sub_button_name
            results[parent_button_name].append(current_row)
            history.append((index, category_index, current_row, parent_button_name))  # Ajouter à l'historique
            index += 1
            if index >= len(current_list):
                index = 0
                category_index += 1
        update_display()

    def process_button(button_name):
        nonlocal index, category_index
        current_category = categories[category_index]
        current_list = listes_types_operation[current_category]
        if index < len(current_list):
            current_row = current_list[index]
            if button_name not in results:
                results[button_name] = []
            results[button_name].append(current_row)
            history.append((index, category_index, current_row, button_name))  # Ajouter à l'historique
            index += 1
            if index >= len(current_list):
                index = 0
                category_index += 1
        update_display()

    def skip_entry():
        nonlocal index, category_index
        index += 1
        if index >= len(listes_types_operation[categories[category_index]]):
            index = 0
            category_index += 1
        update_display()

    def go_back():
        nonlocal index, category_index
        if history:
            index, category_index, current_row, button_name = history.pop()
            results[button_name].remove(current_row)
            update_display()

    def reset_to_main_buttons():
        update_display()  # Reinitialiser l'affichage et les boutons

    def update_buttons(filtered_buttons):
        for widget in root.winfo_children():
            if widget != display_label:
                widget.destroy()
        for i, (label, sub_labels) in enumerate(filtered_buttons.items()):
            row = i // buttons_per_row
            column = i % buttons_per_row
            button = tk.Button(root, text=label, command=lambda l=label: button_clicked(l))
            button.grid(row=row, column=column, padx=10, pady=10, sticky='ew')
        back_button = tk.Button(root, text="Retour", command=go_back)
        back_button.grid(row=(len(filtered_buttons) // buttons_per_row) + 10, column=0, pady=10, sticky='ew')
        skip_button = tk.Button(root, text="Suivant", command=skip_entry)
        skip_button.grid(row=(len(filtered_buttons) // buttons_per_row) + 10, column=buttons_per_row-1, pady=10, sticky='ew')
        for i in range(buttons_per_row):
            root.grid_columnconfigure(i, weight=1)


    index = 0
    category_index = 0
    history = []
    results = {label: [] for label in button_labels}
    listes_types_operation = {cat: tableaux[cat].to_dict('records') for cat in categories}

    # Transformer les dates pour ne garder que l'année, le mois et le jour
    for cat, transactions in listes_types_operation.items():
        for transaction in transactions:
            transaction["DATE D'OPÉRATION"] = transaction["DATE D'OPÉRATION"].strftime('%Y-%m-%d')


    root = tk.Tk()
    root.title("Fenêtre avec Boutons et Affichage de Chaînes")

    window_width = 800
    window_height = 600

    # Centrer la fenêtre sur l'écran
    center_window(root, window_width, window_height)

    buttons_per_row = 5

    display_label = tk.Label(root, text="", wraplength=window_width - 50)
    display_label.grid(row=(len(button_labels) // buttons_per_row) + 1, column=0, columnspan=buttons_per_row, pady=20, sticky='ew')

    update_display()
    root.mainloop()

    return trier_elements(results)



def convertir_dates(data_dict):
    """
    Convertit les dates dans les données en objets datetime.
    """
    for key in data_dict:
        for entry in data_dict[key]:
            for k, v in entry.items():
                if isinstance(v, str) and len(v) == 10 and v.count('-') == 2:
                    try:
                        entry[k] = datetime.strptime(v, "%Y-%m-%d")
                    except ValueError:
                        pass

def ajouter_feuille_bilan_revenus_depenses(wb, data_dict, name_sheet_revenus, name_sheet_depenses):
    def initialiser_base_donnees(colonne, mois):
        """
        Initialise une base de données sous forme de DataFrame avec les revenus en lignes
        et les mois en colonnes.

        :param colonne: Liste des noms des revenus
        :param mois: Liste des noms des mois
        :return: DataFrame initialisé
        """
        # Créer un dictionnaire avec les mois comme colonnes et les revenus comme index
        data = {mois: [0] * len(colonne) for mois in mois}
        
        # Créer le DataFrame
        df = pd.DataFrame(data, index=colonne)
        df[''] = ''

        return df
    
    def revenus_sous_categories(revenus):
        mois = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec', '']
        type_revenus = ['Revenus']

        # Creer la base pour les revenus
        for transaction in revenus:
            type_categorie = transaction.get('Type', 'Revenus')

            if type_categorie not in type_revenus:
                type_revenus.append(type_categorie)

        df_revenus = initialiser_base_donnees(colonne=type_revenus, mois=mois)

        for transaction in revenus:
            type_categorie = transaction.get('Type', 'Revenus')
            montant = transaction['MONTANT']
            date = transaction["DATE D'OPÉRATION"]
            mois = date.strftime('%b')  # Extrait le mois sous forme de chaîne de caractères abrégée (Jan, Feb, Mar, etc.)
            df_revenus.loc[type_categorie, mois] += montant  # Ajoute le montant au mois correspondant

        # Ajouter une ligne 'Revenus' pour les totaux par colonne (mois)
        df_revenus.loc['Revenus'] = df_revenus.sum(axis=0)
        # Ajouter une colonne 'Total' pour les totaux par ligne (revenu)
        df_revenus['Total'] = df_revenus.select_dtypes(include=['number']).sum(axis=1)
        # Ajouter une colonne pourcentage en fonction de la colonne 'Total' en une seule ligne
        df_revenus['Pourcentage'] = round(df_revenus['Total'] / df_revenus.loc['Revenus', 'Total'], 4)


        # Trier le DataFrame par la colonne 'Pourcentage' en ordre décroissant
        df_revenus = df_revenus.sort_values(by='Pourcentage', ascending=False)

        # Ajouter le DataFrame à la feuille Excel
        ajouter_base_donnees_avec_format(df_revenus, wb, name_sheet_revenus, 1)

    def depenses_sous_categories(depenses):
        ############################ DEPENSES CATEGORIES ###########################
        mois = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec', '']
        type_depenses = ['Dépenses']

        # Calculer le montant total et regrouper les montants par type pour les dépenses=
        for categorie, transactions in depenses.items():
            for transaction in transactions:
                montant = transaction['MONTANT']
                if categorie not in type_depenses:
                    type_depenses.append(categorie)

        df_depense = initialiser_base_donnees(colonne=type_depenses, mois=mois)

        for categorie, transactions in depenses.items():
            for transaction in transactions:
                montant = transaction['MONTANT']
                date = transaction["DATE D'OPÉRATION"]
                mois = date.strftime('%b')  # Extrait le mois sous forme de chaîne de caractères abrégée (Jan, Feb, Mar, etc.)
                df_depense.loc[categorie, mois] += montant  # Ajoute le montant au mois correspondant

        # Ajouter une ligne 'Dépenses' pour les totaux par colonne (mois)
        df_depense.loc['Dépenses'] = df_depense.sum(axis=0)
        # Ajouter une colonne 'Total' pour les totaux par ligne (revenu)
        df_depense['Total'] = df_depense.select_dtypes(include=['number']).sum(axis=1)
        # Ajouter une colonne pourcentage en fonction de la colonne 'Total' en une seule ligne
        df_depense['Pourcentage'] = round(df_depense['Total'] / df_depense.loc['Dépenses', 'Total'], 4)


        # Trier le DataFrame par la colonne 'Pourcentage' en ordre décroissant
        df_depense = df_depense.sort_values(by='Pourcentage', ascending=False)

        
        ajouter_base_donnees_avec_format(df_depense, wb, name_sheet_depenses, 1)


        ############################ DEPENSES SOUS CATEGORIES ###########################
        mois = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec', '']
        type_depenses = ['Dépenses']

        # Calculer le montant total et regrouper les montants par type pour les dépenses=
        for categorie, transactions in depenses.items():
            for transaction in transactions:
                montant = transaction['MONTANT']
                type_categorie = transaction.get('Type', categorie)
                if type_categorie not in type_depenses:
                    type_depenses.append(type_categorie)

        df_depense = initialiser_base_donnees(colonne=type_depenses, mois=mois)

        for categorie, transactions in depenses.items():
            for transaction in transactions:
                type_categorie = transaction.get('Type', categorie)
                montant = transaction['MONTANT']
                date = transaction["DATE D'OPÉRATION"]
                mois = date.strftime('%b')  # Extrait le mois sous forme de chaîne de caractères abrégée (Jan, Feb, Mar, etc.)
                df_depense.loc[type_categorie, mois] += montant  # Ajoute le montant au mois correspondant

        # Ajouter une ligne 'Dépenses' pour les totaux par colonne (mois)
        df_depense.loc['Dépenses'] = df_depense.sum(axis=0)
        # Ajouter une colonne 'Total' pour les totaux par ligne (revenu)
        df_depense['Total'] = df_depense.select_dtypes(include=['number']).sum(axis=1)
        # Ajouter une colonne pourcentage en fonction de la colonne 'Total' en une seule ligne
        df_depense['Pourcentage'] = round(df_depense['Total'] / df_depense.loc['Dépenses', 'Total'], 4)


        # Trier le DataFrame par la colonne 'Pourcentage' en ordre décroissant
        df_depense = df_depense.sort_values(by='Pourcentage', ascending=False)



        start_row_next = wb[name_sheet_depenses].max_row + 2
        # Ajouter le DataFrame à la feuille Excel
        ajouter_base_donnees_avec_format(df_depense, wb, name_sheet_depenses, start_row_next)

    
    revenus = data_dict.get('Revenus', [])
    revenus_sous_categories(revenus)

    depenses = {cat: transactions for cat, transactions in data_dict.items() if cat != 'Revenus'}
    depenses_sous_categories(depenses)

def ajouter_base_donnees_avec_format(df, wb, sheet_name, start_row, spacing=5):
    # Créer ou obtenir la feuille
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        start_row -= 1
    else:
        ws = wb.create_sheet(title=sheet_name)
    
    # Ajouter le DataFrame à la feuille Excel
    for r in dataframe_to_rows(df, index=True, header=True):
        ws.append(r)
    
    # Définir les styles
    font_mois = Font(color="FFFFFF", bold=True)
    fill_mois = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    alignment_center = Alignment(horizontal='center', vertical='center')
    font_revenus = Font(color="000000", bold=True)
    fill_revenus = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    
    border_style_gras = Border(
        right=Side(style='thin'),
        top=Side(style='medium'),
        bottom=Side(style='medium')
    )
    border_style_gras_gauche_droite = Border(
        left=Side(style='medium'),
        right=Side(style='medium')
    )
    border_style_gras_droite = Border(
        right=Side(style='medium'),
        top=Side(style='medium'),
        bottom=Side(style='medium')
    )
    border_style_standard = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Changer le formatage de la ligne des mois
    for row in ws.iter_rows(min_row=start_row, max_row=start_row, min_col=2, max_col=ws.max_column):
        for index, cell in enumerate(row):
            if index != ws.max_column-4:
                cell.font = font_mois
                cell.fill = fill_mois
                cell.alignment = alignment_center
                cell.border = border_style_gras

    normal = False
    # Appliquer les formatages
    for index_row, row in enumerate(ws.iter_rows(min_row=start_row+2, max_row=ws.max_row, min_col=0, max_col=ws.max_column)):
        if normal:
            for index_column, cell in enumerate(row):
                # Pour les cases du tableau
                if (index_column != ws.max_column-3) and (index_column != ws.max_column-4) and (index_column != ws.max_column-1):
                    if (index_column == 0):
                        cell.border = Border(
                            right=Side(style='medium'),
                            bottom=Side(style='thin'),
                        )
                    else:
                        cell.border = border_style_standard
                # Pour la bordure à droite des tableaux
                elif (index_column == ws.max_column-4) or (index_column == ws.max_column-1):
                    cell.border = Border(
                        right=Side(style='medium'),
                        bottom=Side(style='thin')
                    )
                    
                # Pour la ligne du bas
                if (index_row == ws.max_row-start_row-2) and (index_column != ws.max_column-2):
                    if (index_column == 0):
                        cell.border = Border(
                            right=Side(style='medium'),
                            bottom=Side(style='medium'),
                        )
                    elif (index_column == ws.max_column-4) or (index_column == ws.max_column-1):
                        cell.border = Border(
                            right=Side(style='medium'),
                            bottom=Side(style='medium')
                        )
                    else:
                        cell.border = Border(
                            right=Side(style='thin'),
                            bottom=Side(style='medium')
                        )

                if (index_column == ws.max_column-3):
                    cell.border = Border(
                        right=Side(style='medium')
                    )

        # Appliquer le formatage pour la ligne 'Revenus'
        else:
            for index_column, cell in enumerate(row):
                if (index_column == 0):
                    cell.font = font_revenus
                    cell.fill = fill_revenus
                    cell.border =  Border(
                        right=Side(style='medium'),
                        top=Side(style='medium'),
                        bottom=Side(style='medium')
                    )
                elif (index_column != ws.max_column-3):
                    cell.font = font_revenus
                    cell.fill = fill_revenus
                    cell.border = border_style_gras
                else:
                    cell.border = border_style_gras_gauche_droite

            row[-1].border = border_style_gras_droite
            normal = True
    


    # Appliquer la largeur de la colonne A (environ 200 pixels)
    column_width = 250 / 7  # Conversion approximative de pixels à largeur de colonne
    ws.column_dimensions['A'].width = column_width

    column_width = 100 / 7
    ws.column_dimensions['P'].width = column_width

    for cell in ws['P']:
        cell.alignment = alignment_center

    # Ajouter l'espacement après le dernier bloc de données
    ws.append([''] * ws.max_column)  # Ajouter une ligne vide pour l'espacement
    for _ in range(spacing-1):  # Ajouter le nombre de lignes d'espacement spécifié
        ws.append([''] * ws.max_column)


    
    # Appliquer le style de bordure à la cellule de la dernière ligne et de l'avant-dernière colonne
    last_row_idx = ws.max_row-spacing
    second_last_col_idx = ws.max_column - 2
    specific_cell = ws.cell(row=last_row_idx, column=second_last_col_idx + 1)  # Ajuster pour l'index 1-based
    specific_cell.border = Border(
        right=Side(style='thin'),
        bottom=Side(style='medium')
    )

def creer_fichier_excel(data_dict: dict, ouvrir:bool = False) -> None:
    # Convertir les dates en objets datetime
    convertir_dates(data_dict)

    # Dossier prédéfini pour enregistrer le fichier
    dossier_initial = "Bilan"  # Remplace par le chemin vers ton dossier

    # Demander à l'utilisateur où enregistrer le fichier
    root = tk.Tk()
    root.withdraw()  # Cache la fenêtre principale de Tkinter
    fichier_sortie = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        initialdir=dossier_initial  # Utiliser le dossier prédéfini
    )
    if not fichier_sortie:
        return  # L'utilisateur a annulé l'enregistrement

    # Créer le fichier Excel
    with pd.ExcelWriter(fichier_sortie, engine='openpyxl') as writer:
        # Ajouter les feuilles de chaque catégorie
        for sheet_name, data in data_dict.items():
            if data:
                # Convertir la liste de dictionnaires en DataFrame
                df = pd.DataFrame(data)
                # Assurer que les en-têtes sont des chaînes de caractères
                df.columns = [str(col) for col in df.columns]
                # Écrire le DataFrame dans une feuille avec le nom correspondant à la clé
                df.to_excel(writer, sheet_name=sheet_name, index=False)

    # Charger le fichier Excel pour le formatage
    wb = load_workbook(fichier_sortie)

    ajouter_feuille_bilan_revenus_depenses(wb, data_dict, "Bilan Revenus", "Bilan Dépenses")

    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # Formater les feuilles de données
    for sheet_name in wb.sheetnames:
        if sheet_name != "Bilan Revenus" and sheet_name != "Bilan Dépenses":
            ws = wb[sheet_name]
            min_col, min_row, max_col, max_row = 1, 1, ws.max_column, ws.max_row
            table_style = TableStyleInfo(
                name='TableStyleMedium2', showFirstColumn=False,
                showLastColumn=False, showRowStripes=True, showColumnStripes=True
            )
            safe_table_name = f'Table_{sheet_name}'.replace(" ", "_")
            table = Table(displayName=safe_table_name, ref=f'A1:{chr(64 + max_col)}{max_row}')
            table.tableStyleInfo = table_style
            ws.add_table(table)

            for row in ws.iter_rows(min_row=2, max_row=max_row):
                for cell in row:
                    if isinstance(cell.value, datetime):
                        cell.number_format = 'DD/MM/YYYY'
                    cell.alignment = center_alignment

            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 10)
                ws.column_dimensions[column].width = adjusted_width

    # Sauvegarder le fichier Excel avec le format appliqué
    wb.save(fichier_sortie)
    if ouvrir:
        os.startfile(fichier_sortie)  # Ouvre le fichier après l'enregistrement



def show_buttons(options:list=["Compte Chèques", "Livret A"]) -> str:
    """
    Affiche une fenêtre avec des boutons dont les textes sont les chaînes de la liste `options`.
    Retourne la chaîne de caractères correspondant au bouton sélectionné.

    :param options: Liste de chaînes de caractères pour les textes des boutons.
    :return: Chaîne de caractères correspondant au bouton sélectionné.
    """
    def on_button_click(value):
        nonlocal selected_value
        selected_value = value
        root.destroy()  # Ferme la fenêtre

    # Créer la fenêtre principale
    root = tk.Tk()
    root.title("Choisissez une option")

    # Définir la taille de la fenêtre
    window_width = 800
    window_height = 600
    root.geometry(f"{window_width}x{window_height}")

    # Centrer la fenêtre
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    x = (screen_width // 2) - (window_width // 2)
    y = (screen_height // 2) - (window_height // 2)
    root.geometry(f"{window_width}x{window_height}+{x}+{y}")

    # Variable pour stocker la valeur sélectionnée
    selected_value = None

    # Frame pour contenir les boutons
    frame = tk.Frame(root)
    frame.pack(expand=True, fill=tk.BOTH)

    # Ajouter les boutons à la frame en utilisant grid
    max_buttons_per_row = 5
    row = 0
    col = 0

    for index, option in enumerate(options):
        button = tk.Button(frame, text=option, command=lambda opt=option: on_button_click(opt))
        button.grid(row=row, column=col, padx=10, pady=10, sticky='ew')

        col += 1
        if col >= max_buttons_per_row:
            col = 0
            row += 1

    # Ajuster les colonnes pour qu'elles s'étendent également
    for i in range(max_buttons_per_row):
        frame.columnconfigure(i, weight=1)

    # Lancer la boucle principale
    root.mainloop()

    return selected_value

def name_file(data:dict) -> str: 
    # Convertir les dictionnaires en DataFrames
    dfs = {key: pd.DataFrame(value) for key, value in data.items() if value}
    years = set()
    
    for df in dfs.values():
        if 'DATE D\'OPÉRATION' in df.columns:
            # Convertir la colonne en datetime si ce n'est pas déjà le cas
            df['DATE D\'OPÉRATION'] = pd.to_datetime(df['DATE D\'OPÉRATION'])
            
            # Extraire l'année de la date minimale et maximale
            min_year = df['DATE D\'OPÉRATION'].min().year
            max_year = df['DATE D\'OPÉRATION'].max().year
            
            years.add(min_year)
            years.add(max_year)
        
    years = sorted(years)
    
    if len(years) == 1:
        return years[0]
    return f"{years[0]}-{years[-1]}"
        


def diviser_mois(data: dict) -> dict:
    result = {}
    
    for category, operations in data.items():
        for operation in operations:
            date_str = operation["DATE D'OPÉRATION"]
            date_obj = datetime.strptime(date_str, '%Y-%m-%d')

            year_month = date_obj.strftime('%Y-%m')
            
            if year_month not in result:
                result[year_month] = {}
                
            if category not in result[year_month]:
                result[year_month][category] = []
            
            result[year_month][category].append(operation)
    
    return result



def save_dict_to_json(data_dict, file_path):
    """Sauvegarde un dictionnaire dans un fichier JSON en UTF-8.
    
    Args:
        data_dict (dict): Le dictionnaire à sauvegarder.
        file_path (str): Le chemin du fichier où sauvegarder les données.
    """
    def convert_timestamp_to_string(obj):
        """Convertit les objets Timestamp en chaînes de caractères."""
        if isinstance(obj, pd.Timestamp):
            return obj.isoformat()
        raise TypeError("Type non serialisable")

    with open(file_path, 'w', encoding='utf-8') as file:
        json.dump(data_dict, file, indent=4, default=convert_timestamp_to_string, ensure_ascii=False)



def extraire_annee(data):
    df = pd.DataFrame(data)

    # S'assurer que la colonne 'DATE D\'OPÉRATION' est au format datetime
    df['DATE D\'OPÉRATION'] = pd.to_datetime(df['DATE D\'OPÉRATION'])

    # Extraire l'année de la première date de la colonne
    annee = df['DATE D\'OPÉRATION'].dt.year.iloc[0]

    # Retourner l'année sous forme de chaîne de caractères
    return str(annee)



def main():
    data, extension, dossier = choisir_et_extraire_donnees_excel()
    
    if (data is not None) and (extension == "Excel"):
        tableaux = trier_donnees(data)
        
        year = extraire_annee(data)
        cheminFileJson = f"Bilan/Archives/{dossier}/{year}.json"

        categories = data["LIBELLÉ COURT"].unique()
        button_labels = {
            'Investissement': ["CTO", "Livret A", "Investissement - Autres"],
            'Revenus': ["Aides et allocations", "Salaires et revenus d'activité", "Revenus de placement", "Pensions", "Intérêts", "Loyers", "Dividendes", "Remboursement", "Chèque reçu", "Déblocage emprunt", "Virement reçu", "Virement interne", "Cashback", "Revenus - Autres"],
            'Abonnement': ["Téléphone", "Internet", "Streaming", "Logiciels"],
            'Impôts': ["Impôt sur taxes", "Impôt sur le revenu", "Impôt sur la fortune", "Taxe foncière", "Taxe d'habitation", "Contribution sociales (CSG / CRDS)"],
            'Banque': ["Epargne", "Remboursement emprumt", "Frais bancaires", "Prélèvement carte débit différé", "Banques - Autres"],
            "Logement": ["Logement - Autres", "Electricité, gaz", "Eau", "Chauffage", "Loyer", "Prêt Immobiler", "Bricolage et jardinage", "Assurance habitation", "Logement - Autres", "Mobilier, électroménager, déco"],
            'Loisir et sorties': ["Voyages, vacances", "Restaurants - Bars", "Diversements, sortie culturelles", "Sports", "Soirée - Sortie", "Loisirs et sorties - Autres"],
            'Santé': ["Medecin", "Pharmacie", "Dentiste", "Mutuelle", "Opticien", "Hôpital"],
            'Transports et véhicules': ["Assurance véhicule", "Crédit auto", "Carburant", "Entretient véhicule", "Transport en commun", "Billet d'avion, Billet de train", "Taxi, VTC", "Location de véhicule", "Péage", "Stationnement"],
            'Vie quotidienne': ["Alimentation - Supermarché", "Frais animaux", "Coiffeur, soins", "Habillement", "Achat, shopping", "Jeux Vidéo", "Frais postaux", "Achat multimédias - Hight tech", "Autres", "Aide-à-domicile", "Cadeaux", "Vie quotidienne - Autres"],
            'Enfant(s)': ["Pension alimentaire", "Crèche, baby-sitter", "Scolarité, études", "Argent de poche", "Activités enfants"],
        }

        # Trier les clés du dictionnaire
        sorted_keys = sorted(button_labels.keys())
        # Créer un nouveau dictionnaire avec les clés triées et les valeurs triées
        sorted_button_labels = {key: sorted(button_labels[key]) for key in sorted_keys}

        
        
        results = afficher_fenetre_avec_boutons(categories, tableaux, sorted_button_labels, cheminFileJson)

        if dossier is not None:
            if not os.path.exists(f"Bilan"):
                os.makedirs(f"Bilan")
            if not os.path.exists(f"Bilan/{dossier}"):
                os.makedirs(f"Bilan/{dossier}")
            if not os.path.exists(f"Bilan/{dossier}/{year}"):
                os.makedirs(f"Bilan/{dossier}/{year}")

            save_dict_to_json(results, f"Bilan/Archives/{dossier}/{year}.json")

            chemin_file = f"Bilan/{dossier}/{year}/Bilan {year}.html"

            if dossier == "Compte Chèques":
                results_copy = copy.deepcopy(results)
                results_copy1 = copy.deepcopy(results)

                create_charts(results_copy, chemin_file)
                # Graphique pour chaques mois
                allmonth = diviser_mois(data=results_copy1)
                for year_month, transaction in allmonth.items():
                    chemin_file = f"Bilan/{dossier}/{year}/{year_month}.html"
                    create_charts(transaction, chemin_file)

                creer_fichier_excel(results)
            else:
                results["Virement sur les comptes"] = results.pop("Investissement")
                for tab in results.values():
                    for ele in tab: 
                        if ele["Type"] == "Livret A":
                            ele["Type"] = "Compte Courant"

                results_copy = copy.deepcopy(results)
                results_copy1 = copy.deepcopy(results)


                create_charts(results_copy, chemin_file, False)
                # Graphique pour chaques mois
                allmonth = diviser_mois(data=results_copy1)
                for year_month, transaction in allmonth.items():
                    chemin_file = f"Bilan/{dossier}/{year}/{year_month}.html"
                    create_charts(transaction, chemin_file, False)

    elif (data is not None) and (extension == "Json"):
        year = name_file(data)

        if dossier is not None:
            if not os.path.exists(f"Bilan"):
                os.makedirs(f"Bilan")
            if not os.path.exists(f"Bilan/{dossier}"):
                os.makedirs(f"Bilan/{dossier}")
            if not os.path.exists(f"Bilan/{dossier}/{year}"):
                os.makedirs(f"Bilan/{dossier}/{year}")

            chemin_file = f"Bilan/{dossier}/{year}/Bilan {year}.html"
            results_copy = copy.deepcopy(data)
            results_copy1 = copy.deepcopy(data)
            
            
            if dossier == "Compte Chèques":
                create_charts(results_copy, chemin_file)
                # Graphique pour chaques mois
                allmonth = diviser_mois(data=results_copy1)
                for year_month, transaction in allmonth.items():
                    chemin_file = f"Bilan/{dossier}/{year}/{year_month}.html"
                    create_charts(transaction, chemin_file)

            else:
                for tab in results_copy.values():
                    for ele in tab: 
                        if ele["Type"] == "Livret A":
                            ele["Type"] = "Compte Courant"


                create_charts(results_copy, chemin_file, False)
                # Graphique pour chaques mois
                allmonth = diviser_mois(data=results_copy1)
                for year_month, transaction in allmonth.items():
                    chemin_file = f"Bilan/{dossier}/{year}/{year_month}.html"
                    create_charts(transaction, chemin_file, False)



if __name__ == "__main__":
    main()
