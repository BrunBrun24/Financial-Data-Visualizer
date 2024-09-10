import pandas as pd
import os
import re
from datetime import datetime
from PyPDF2 import PdfReader
from openpyxl import Workbook
import pdfplumber
import json
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter



class TradeRepublicFileExcelJson:
    """
    La classe TradeRepublicFileExcelJson est conçue pour analyser les données du courtier Trade République via les fichiers PDF téléchargés,
    et créer un fichier Excel et des fichiers JSON avec ces données.

    Attributs :
        - directoryData : Chemin du répertoire de données contenant les fichiers PDF.
        - tickerMapping : Dictionnaire pour mapper les tickers avec des informations supplémentaires.

    Méthodes :
        - __init__(self, directoryData: str, tickerMapping: dict) : Initialise la classe avec un répertoire de données.
        - DownloadDataAndCreateFileExcel(self, cheminEnregistrementFichier: str, ouvrir=False) : Crée un fichier Excel récapitulatif des gains.
        - SaveDataFrameAsJson(self, dataFrame: pd.DataFrame, pathJson: str) : Enregistre un DataFrame sous forme de fichier JSON.
        - SaveTransactionsToJson(self, dataFrames: list, pathJson: str) : Enregistre les transactions d'achats et de ventes dans un fichier JSON.
        - ConvertirColonnesDate(dataFrame: pd.DataFrame) -> pd.DataFrame : Convertit les colonnes contenant "date" en dates formatées.
        - PutDataFrameSheetExcel(workbook: Workbook, nomFeuille: str, df: pd.DataFrame, colonnesEuros=[], colonnesPourcentages=[], colonnesDates=[], appliquerTableau=False) -> None : Formate les colonnes du DataFrame dans le classeur Excel.
        - GetPdfFilesFromFolder(folderPath: str) -> list : Retourne une liste de tous les fichiers PDF dans le dossier spécifié.
        - RenameAndMoveDepotRetraitArgentInteret(self, filePath: str, detailsFolder: str) -> None : Renomme et déplace un fichier PDF en fonction des informations extraites.
        - RenameAndMoveDividendes(self, filePath, detailsFolder) : Renomme et déplace un fichier PDF en fonction des informations extraites.
        - RenameAndMoveOrdresAchats(self, filePath: str, detailsFolder: str) -> None : Renomme et déplace un fichier PDF basé sur les informations extraites.
        - RenameAndMoveOrdresVentes(self, filePath: str) : Renomme et déplace un fichier PDF basé sur les informations extraites.
        - DataDepotArgent(pdfFiles: list, chemin: str) -> pd.DataFrame : Extrait les données des fichiers PDF spécifiés et les retourne sous forme de DataFrame.
        - DataDividendes(pdfFiles: list, chemin: str) -> pd.DataFrame : Extrait les données des fichiers PDF spécifiés et les retourne sous forme de DataFrame.
        - DataInteret(pdf: list, chemin: str) -> pd.DataFrame : Extrait les données des fichiers PDF spécifiés et les retourne sous forme de DataFrame.
        - DataOrdresAchats(pdfFiles: list, chemin: str) -> pd.DataFrame : Extrait les données des fichiers PDF spécifiés et les retourne sous forme de DataFrame.
        - DataOrdresVentes(pdfFiles: list, chemin: str) -> pd.DataFrame : Extrait les données des fichiers PDF spécifiés et les retourne sous forme de DataFrame.
        - DataRetraitArgent(pdfFiles: list, chemin: str) -> pd.DataFrame : Extrait les données des fichiers PDF spécifiés et les retourne sous forme de DataFrame.
        - ProcessPdf(nameDirectory: str, cheminDossierRenommer: str, nameFunction) -> None : Parcourt les fichiers PDF dans un répertoire et les traite.
        - ExtractInformationPdf(filePath: str) -> str : Extrait le texte de chaque page d'un fichier PDF.
    """

    def __init__(self, directoryData: str, tickerMapping: dict) -> None:
        """
        Initialise une instance de TradeRepublicFileExcelJson avec un répertoire de données.

        Args:
            directoryData (str): Chemin du répertoire de données contenant les fichiers PDF.
            tickerMapping (dict): Un dictionnaire pour mapper les tickers avec des informations supplémentaires.

        Raises:
            AssertionError: Si directoryData n'est pas une chaîne, n'existe pas ou n'est pas un répertoire.
        """
        assert isinstance(directoryData, str), f"directoryData doit être une chaîne, mais c'est {type(directoryData).__name__}."
        assert os.path.exists(directoryData), f"Le fichier ou le dossier '{directoryData}' n'existe pas."
        assert os.path.isdir(directoryData), f"'{directoryData}' n'est pas un dossier."
        
        self.directoryData = directoryData
        self.tickerMapping = tickerMapping
        

    def DownloadDataAndCreateFileExcel(self, cheminEnregistrementFichier: str, ouvrir=False) -> None:
        """
        Crée un récapitulatif des gains en générant un classeur Excel avec plusieurs feuilles pour différents types de données financières.
        Les feuilles sont créées à partir de fichiers PDF spécifiques dans le dossier spécifié.

        Args:
            cheminEnregistrementFichier (str): Chemin où le fichier Excel sera enregistré.
            ouvrir (bool, optionnel): Si True, le fichier Excel sera automatiquement ouvert après sa création. Par défaut False.

        Raises:
            AssertionError: Si cheminEnregistrementFichier n'est pas une chaîne ou si ouvrir n'est pas un booléen.
        """
        assert isinstance(cheminEnregistrementFichier, str), \
            f"cheminEnregistrementFichier doit être une chaîne, mais c'est {type(cheminEnregistrementFichier).__name__}."
        assert isinstance(ouvrir, bool), \
            f"ouvrir doit être un booléen, mais c'est {type(ouvrir).__name__}."

        dossier = self.directoryData
        buySell = []

        # Créez un classeur Excel
        workbook = Workbook()
        
        # Supprime la feuille par défaut "Sheet"
        sheetToDelete = workbook["Sheet"]
        workbook.remove(sheetToDelete)

        # L'ordre dans le dictionnaire définit la continuité des feuilles dans le fichier Excel
        operations = [
            {"nomFichier": "Ordres d'achats", "data": self.DataOrdresAchats, "fileJson": "Bilan/Archives/Bourse/Argents investis.json"},
            {"nomFichier": "Ordres de ventes/FacturesVentes", "data": self.DataOrdresVentes, "fileJson": "Bilan/Archives/Bourse/Argents vendus.json"},
            {"nomFichier": "Dépôts d'argents", "data": self.DataDepotArgent, "fileJson": "Bilan/Archives/Bourse/Dépôts d'espèces.json"},
            {"nomFichier": "Dividendes", "data": self.DataDividendes, "fileJson": "Bilan/Archives/Bourse/Dividendes.json"},
            {"nomFichier": "Interets", "data": self.DataInteret, "fileJson": "Bilan/Archives/Bourse/Interêts.json"},
            {"nomFichier": "Retraits d'argents", "data": self.DataRetraitArgent, "fileJson": "Bilan/Archives/Bourse/Retraits.json"},
        ]

        # Boucle pour traiter chaque type de donnée et créer les feuilles Excel correspondantes
        for operation in operations:
            nomFichier = dossier + operation["nomFichier"]
            filesPdf = self.GetPdfFilesFromFolder(nomFichier)

            dataFilePdf, nomFeuille, colonnesEuros, colonnesPourcentages, colonnesDates, appliquerTableau = operation["data"](filesPdf, nomFichier)
            dataFilePdf.columns = [col.capitalize() for col in dataFilePdf.columns]
            self.PutDataFrameSheetExcel(workbook=workbook, nomFeuille=nomFeuille, df=dataFilePdf, colonnesEuros=colonnesEuros, colonnesPourcentages=colonnesPourcentages, colonnesDates=colonnesDates, appliquerTableau=appliquerTableau)

            
            dataFilePdfJson = self.ConvertirColonnesDate(dataFilePdf)
            self.SaveDataFrameAsJson(dataFilePdfJson, operation["fileJson"])
            
            if operation["nomFichier"] in ["Ordres d'achats", "Ordres de ventes/FacturesVentes"]:
                buySell.append(dataFilePdf)
                

        self.SaveTransactionsToJson(buySell, "Bilan/Archives/Bourse/Transactions.json")
        # Enregistrez le fichier Excel
        workbook.save(cheminEnregistrementFichier)
        
        if ouvrir:
            os.system(f'start excel "{cheminEnregistrementFichier}"')

    def SaveDataFrameAsJson(self, dataFrame: pd.DataFrame, pathJson: str):
        """
        Enregistre un DataFrame sous forme de fichier JSON et applique un mapping des tickers si une colonne 'ticker' existe.

        Args:
            dataFrame (pd.DataFrame): Le DataFrame à sauvegarder.
            pathJson (str): Le chemin où sauvegarder le fichier JSON.

        Returns:
            None
        """
        mappingTickers = self.tickerMapping
        
        # Vérifications des types des arguments
        assert isinstance(dataFrame, pd.DataFrame), f"dataFrame doit être un pd.DataFrame, mais c'est {type(dataFrame).__name__}."
        assert isinstance(pathJson, str) and pathJson.endswith(".json"), \
            f"pathJson doit être une chaîne se terminant par '.json', mais c'est {type(pathJson).__name__}."
        
        # Normalisation des clés du dictionnaire de mapping en minuscules
        mappingTickers = {key.lower(): value for key, value in mappingTickers.items()}

        # Ne pas modifier le nom des colonnes mais les comparer en minuscules
        columnsLower = [col.lower() for col in dataFrame.columns]

        # Vérifier si une colonne 'ticker' existe (insensible à la casse)
        if 'ticker' in columnsLower:
            tickerCol = dataFrame.columns[columnsLower.index('ticker')]  # Récupérer le nom original de la colonne 'ticker'

            # Création d'un motif pour rechercher les noms d'entreprises dans les tickers
            pattern = re.compile('|'.join(re.escape(key) for key in mappingTickers.keys()), re.IGNORECASE)

            def mapTickerValue(ticker):
                """Remplace la valeur du ticker si elle correspond à un nom d'entreprise dans le dictionnaire."""
                match = pattern.search(ticker.lower())
                if match:
                    matchedKey = match.group(0).lower()
                    return mappingTickers[matchedKey]
                return ticker

            # Appliquer le mappage des tickers à la colonne 'ticker'
            dataFrame[tickerCol] = dataFrame[tickerCol].apply(mapTickerValue)

        # Convertir le DataFrame en JSON
        jsonData = dataFrame.to_json(orient="records", force_ascii=False, indent=4)

        # Enregistrer le JSON dans le fichier
        with open(pathJson, 'w', encoding="utf-8") as f:
            f.write(jsonData)

    def SaveTransactionsToJson(self, dataFrames: list, pathJson: str):
        """
        Enregistre les transactions d'achats et de ventes dans un fichier JSON structuré, avec correction des tickers.

        Args:
            dataFrames (list): Liste contenant deux DataFrames, l'un pour les achats, l'autre pour les ventes.
            pathJson (str): Chemin où sauvegarder le fichier JSON.
            mappingTickers (dict): Dictionnaire de mapping des noms d'entreprises vers leurs tickers.

        Returns:
            None
        """
        mappingTickers = self.tickerMapping
        # Vérifications des types des arguments
        assert isinstance(dataFrames, list) and len(dataFrames) == 2, \
            f"dataFrames doit être une liste contenant deux DataFrames, mais c'est {type(dataFrames).__name__}."
        assert isinstance(pathJson, str) and pathJson.endswith(".json"), \
            f"pathJson doit être une chaîne se terminant par '.json', mais c'est {type(pathJson).__name__}."
        assert isinstance(mappingTickers, dict), f"mappingTickers doit être un dictionnaire: ({type(mappingTickers).__name__})"

        achats, ventes = dataFrames

        # Normalisation des clés du dictionnaire de mapping en minuscules
        mappingTickers = {key.lower(): value for key, value in mappingTickers.items()}

        # Création d'un motif pour rechercher les noms d'entreprises dans les tickers
        pattern = re.compile('|'.join(re.escape(key) for key in mappingTickers.keys()), re.IGNORECASE)

        # Format JSON pour les transactions
        transactions = []

        def findTransactionByTicker(transactions, ticker):
            """Recherche un ticker dans la liste de transactions."""
            for transaction in transactions:
                if transaction["ticker"] == ticker:
                    return transaction
            return None

        # Concaténer achats et ventes pour obtenir tous les tickers uniques
        tickers = sorted(set(achats["Ticker"]).union(set(ventes["Ticker"])))

        for ticker in tickers:
            # Filtrer les transactions pour ce ticker
            achatDf = achats[achats["Ticker"] == ticker]
            venteDf = ventes[ventes["Ticker"] == ticker]

            # Corriger les tickers à partir des noms d'entreprises
            match = pattern.search(ticker.lower())
            if match:
                matchedKey = match.group(0).lower()
                ticker = mappingTickers[matchedKey]

            # Rechercher si le ticker est déjà dans transactions
            transaction = findTransactionByTicker(transactions, ticker)
            if not transaction:
                transaction = {"ticker": ticker, "achats": [], "ventes": []}
                transactions.append(transaction)

            # Ajouter les achats au ticker
            for _, row in achatDf.iterrows():
                transaction["achats"].append({
                    "date": (datetime(1899, 12, 30) + pd.Timedelta(days=int(row["Date de valeur"]))).strftime("%Y-%m-%d"),
                    "price": abs(row["Montant investi"])
                })

            # Ajouter les ventes au ticker
            for _, row in venteDf.iterrows():
                transaction["ventes"].append({
                    "date": (datetime(1899, 12, 30) + pd.Timedelta(days=int(row["Date de valeur"]))).strftime("%Y-%m-%d"),
                    "price": abs(row["Montant gagné"])
                })

        # Sauvegarde du fichier JSON
        with open(pathJson, 'w', encoding='utf-8') as f:
            json.dump({"transactions": transactions}, f, ensure_ascii=False, indent=4)

    @staticmethod
    def ConvertirColonnesDate(dataFrame: pd.DataFrame) -> pd.DataFrame:
        """
        Convertit toutes les colonnes du DataFrame contenant "date" dans leur nom en dates formatées (YYYY-MM-DD).
        
        Args:
            dataFrame (pd.DataFrame): Le DataFrame contenant les données avec des colonnes à traiter.
        
        Returns:
            pd.DataFrame: Un nouveau DataFrame avec les colonnes de date converties.
        """
        # Vérification du type du paramètre
        assert isinstance(dataFrame, pd.DataFrame), f"dataFrame doit être un DataFrame, mais c'est {type(dataFrame).__name__}."
        
        df = dataFrame.copy()
        
        # Parcours de toutes les colonnes pour détecter celles contenant 'date' (insensible à la casse)
        for col in df.columns:
            if 'date' in col.lower() and isinstance(df[col].iloc[-1], int):
                # Conversion des valeurs de cette colonne en dates au format YYYY-MM-DD
                df[col] = df[col].apply(lambda x: (datetime(1899, 12, 30) + pd.Timedelta(days=int(x))).strftime("%Y-%m-%d") if pd.notnull(x) else x)
        
        return df

    @staticmethod
    def PutDataFrameSheetExcel(workbook: Workbook, nomFeuille: str, df: pd.DataFrame, colonnesEuros=[], colonnesPourcentages=[], colonnesDates=[], appliquerTableau=False) -> None:
        """
        Formate les colonnes du DataFrame dans le classeur Excel en fonction des listes de colonnes à formater et ajoute les données dans la feuille spécifiée.
        
        Args:
            workbook (Workbook): Le classeur Excel où les données seront ajoutées.
            df (pd.DataFrame): Le DataFrame contenant les données à formater.
            colonnesEuros (list): Liste des colonnes à formater en euros.
            colonnesPourcentages (list): Liste des colonnes à formater en pourcentage.
            colonnesDates (list): Liste des colonnes à formater en dates.
            appliquerTableau (bool): Indique s'il faut appliquer un tableau autour des données.
            nomFeuille (str): Le nom de la feuille où les données seront ajoutées.
        """
        # Crée une nouvelle feuille dans le classeur Excel avec le nom spécifié
        feuille = workbook.create_sheet(title=nomFeuille)
        
        # Qui est la première colonne qui correspond aux dates de la plus récente à la plus ancienne
        df = df.sort_values(by=df.columns[0], ascending=False)
        
        # Ajoute les données du DataFrame à la feuille
        for row in dataframe_to_rows(df, index=False, header=True):
            feuille.append(row)

        # Obtenir les indices des colonnes à formater
        colonnesEurosIndices = [df.columns.get_loc(col) + 1 for col in colonnesEuros if col in df.columns]
        colonnesPourcentagesIndices = [df.columns.get_loc(col) + 1 for col in colonnesPourcentages if col in df.columns]
        colonnesDatesIndices = [df.columns.get_loc(col) + 1 for col in colonnesDates if col in df.columns]

        # Formater les colonnes en euros
        for col in colonnesEurosIndices:
            for cellule in feuille.iter_rows(min_col=col, max_col=col, min_row=2):
                for cell in cellule:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '0.00€'

        # Formater les colonnes en pourcentage
        for col in colonnesPourcentagesIndices:
            for cellule in feuille.iter_rows(min_col=col, max_col=col, min_row=2):
                for cell in cellule:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '0.00%'

        # Formater les colonnes en dates
        for col in colonnesDatesIndices:
            for cellule in feuille.iter_rows(min_col=col, max_col=col, min_row=2):
                for cell in cellule:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = 'DD/MM/YYYY'

        # Appliquer le tableau si le booléen est à True
        if appliquerTableau:
            ref = f"A1:{get_column_letter(feuille.max_column)}{feuille.max_row}"
            tableau = Table(displayName=("TableauDonnees" + nomFeuille).replace(" ", "").replace("'", ""), ref=ref)
            style = TableStyleInfo(
                name="TableStyleMedium6", showFirstColumn=False,
                showLastColumn=False, showRowStripes=True, showColumnStripes=True
            )
            tableau.tableStyleInfo = style
            feuille.add_table(tableau)

        # Ajuster automatiquement la largeur des colonnes en fonction du contenu
        for column in feuille.columns:
            max_length = 0
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
                cell.alignment = Alignment(horizontal="center", vertical="center")
            feuille.column_dimensions[get_column_letter(column[0].column)].width = max_length + 3


    @staticmethod
    def GetPdfFilesFromFolder(folderPath: str) -> list:
        """
        Retourne une liste de tous les fichiers PDF dans le dossier spécifié.

        Args:
            folderPath (str): Le chemin du dossier à explorer.

        Returns:
            list: Liste des noms de fichiers PDF (en str) présents dans le dossier.
        """
        # Vérification du type d'entrée
        assert isinstance(folderPath, str) and os.path.isdir(folderPath), f"Le dossier '{folderPath}' n'existe pas."

        # Récupération de la liste des fichiers .pdf dans le dossier
        filePdf = [file for file in os.listdir(folderPath) if file.lower().endswith('.pdf') and os.path.isfile(os.path.join(folderPath, file))]
        
        return filePdf


    def RenameAndMoveDepotRetraitArgentInteret(self, filePath: str, detailsFolder: str) -> None:
        """
        Renomme et déplace un fichier PDF en fonction des informations extraites de son texte.

        Args:
            filePath: Le chemin complet du fichier PDF à renommer et déplacer.
            detailsFolder: Le dossier où le fichier PDF sera déplacé après avoir été renommé.

        Returns:
            None
        """

        # Assertions pour vérifier les types des arguments
        assert isinstance(filePath, str) and os.path.isfile(filePath), f"Le fichier '{filePath}' n'existe pas ou n'est pas une extension .pdf"
        assert isinstance(detailsFolder, str) and os.path.isdir(detailsFolder), f"Le dossier '{detailsFolder}' n'existe pas"
        
        # Extraire le texte du fichier PDF
        text = self.ExtractInformationPdf(filePath)
        
        # Assertions pour vérifier que le texte a été extrait correctement
        assert text is not None, "Le texte extrait du fichier PDF est None."

        # Extraire les informations nécessaires pour le renommage du fichier
        dateMatch = re.search(r'DATE (\d{2}/\d{2}/\d{4})', text)
        
        if dateMatch:
            dateStr = dateMatch.group(1)
            # Convertir la chaîne en objet datetime
            dateObj = datetime.strptime(dateStr, "%d/%m/%Y")
            # Reformatter la date dans le format "année/mois/jour"
            nouvelleDateStr = dateObj.strftime("%Y/%m/%d")

            # Créer le nouveau nom de fichier
            newFileName = f"{nouvelleDateStr.replace('/', '-')}.pdf"
            
            # Vérifier si le fichier existe déjà dans le dossier de destination
            counter = 0
            while os.path.exists(os.path.join(detailsFolder, newFileName)):
                counter += 1
                newFileName = f"{nouvelleDateStr.replace('/', '-')}_({counter}).pdf"

            try:
                # Déplacer et renommer le fichier dans le dossier Details
                os.rename(filePath, os.path.join(detailsFolder, newFileName))
            except Exception as e:
                print(f"Une erreur s'est produite lors de l'enregistrement du fichier : {e}")
        else:
            raise ValueError(f"Impossible de renommer le fichier {filePath}. Informations de date manquantes.")

    def RenameAndMoveDividendes(self, filePath, detailsFolder):
        """
        Renomme et déplace un fichier PDF en fonction des informations extraites du texte.

        Args:
            filePath: Chemin complet vers le fichier PDF à renommer.
            detailsFolder: Dossier de destination où le fichier sera déplacé.
        """

        # Vérifiez que les chemins sont des chaînes de caractères
        assert isinstance(filePath, str) and os.path.isfile(filePath), f"Le fichier '{filePath}' n'existe pas ou n'est pas une extension .pdf"
        assert isinstance(detailsFolder, str) and os.path.isdir(detailsFolder), f"Le dossier '{detailsFolder}' n'existe pas"

        # Extraire le texte du fichier PDF
        text = self.ExtractInformationPdf(filePath)
        assert text, "Le texte extrait du fichier PDF est vide"

        # Extraire les informations nécessaires pour le renommage du fichier
        dateMatch = re.search(r'DATE (\d{2}/\d{2}/\d{4})', text)
        dateMatch1 = re.search(r'DATE (\d{2}\.\d{2}\.\d{4})', text)
        
        if dateMatch:
            dateStr = dateMatch.group(1)
            # Convertir la chaîne en objet datetime
            dateObj = datetime.strptime(dateStr, "%d/%m/%Y")
            # Reformatter la date dans le format "année/mois/jour"
            nouvelleDateStr = dateObj.strftime("%Y/%m/%d")
        elif dateMatch1:
            dateStr = dateMatch1.group(1).replace(".", "/")
            # Convertir la chaîne en objet datetime
            dateObj = datetime.strptime(dateStr, "%d/%m/%Y")
            # Reformatter la date dans le format "année/mois/jour"
            nouvelleDateStr = dateObj.strftime("%Y/%m/%d")
        else:
            raise ValueError(f"Date non trouvée dans le fichier {filePath}")

        # Recherche de la phrase spécifique
        pattern = r'POSITION QUANTITÉ REVENU MONTANT\n(.*?)\n'
        pattern1 = r'POSITION QUANTITÉ TAUX MONTANT\n(.*?)\n'
        match = re.search(pattern, text, re.DOTALL)
        match1 = re.search(pattern1, text, re.DOTALL)
        
        if match:
            nameTicker = match.group(1).strip()
            if re.search(re.escape(" Inc."), nameTicker):
                nameTicker = nameTicker[:-5]
        elif match1:
            nameTicker = match1.group(1).strip()
            if re.search(re.escape(" Inc."), nameTicker):
                nameTicker = nameTicker[:-5]
        else:
            raise ValueError(f"Nom du ticker non trouvé dans le fichier {filePath}")

        # Vérifier si les informations nécessaires ont été trouvées
        if not nouvelleDateStr:
            raise ValueError(f"Date reformattée non trouvée pour le fichier {filePath}")
        if not nameTicker:
            raise ValueError(f"Nom du ticker non trouvé pour le fichier {filePath}")

        # Créer le nouveau nom de fichier
        newFileName = f"{nameTicker}_{nouvelleDateStr.replace('/', '-')}.pdf"
        
        # Vérifier si le fichier existe déjà dans le dossier de destination
        counter = 0
        while os.path.exists(os.path.join(detailsFolder, newFileName)):
            counter += 1
            newFileName = f"{nameTicker}_{nouvelleDateStr.replace('/', '-')}_({counter}).pdf"

        try:
            # Déplacer et renommer le fichier dans le dossier de destination
            os.rename(filePath, os.path.join(detailsFolder, newFileName))
        except Exception as e:
            print("Une erreur s'est produite lors de l'enregistrement du fichier :", e)

    def RenameAndMoveOrdresAchats(self, filePath: str, detailsFolder: str) -> None:
        """
        Extrait des informations d'un fichier PDF pour renommer et déplacer le fichier dans un autre répertoire.

        Args:
            filePath: Chemin complet du fichier PDF à renommer et déplacer.
            detailsFolder: Chemin du répertoire où le fichier renommé sera déplacé.

        Raises:
            ValueError: Si la date ou le nom du ticker ne peut pas être trouvé dans le texte du PDF.
        """

        # Assertions pour vérifier les types des entrées
        assert isinstance(filePath, str) and os.path.isfile(filePath), f"Le fichier '{filePath}' n'existe pas ou n'est pas une extension .pdf"
        assert isinstance(detailsFolder, str) and os.path.isdir(detailsFolder), f"Le dossier '{detailsFolder}' n'existe pas"
        
        # Extraire le texte du fichier PDF
        text = self.ExtractInformationPdf(filePath)
        
        # Extraire la date du texte
        dateMatch = re.search(r'DATE (\d{2}/\d{2}/\d{4})', text)
        if dateMatch:
            dateStr = dateMatch.group(1)
            # Convertir la chaîne en objet datetime
            dateObj = datetime.strptime(dateStr, "%d/%m/%Y")
            # Reformatter la date dans le format "année-mois-jour"
            nouvelleDateStr = dateObj.strftime("%Y-%m-%d")
        else:
            raise ValueError(f"Date non trouvée dans le fichier {filePath}.")
        
        # Extraire le nom du ticker
        pattern = r'POSITION QUANTITÉ COURS MOYEN MONTANT\n(.*?)\n'
        match = re.search(pattern, text, re.DOTALL)
        if match:
            nameTicker = match.group(1).strip()
            # Supprimer " Inc." du nom du ticker s'il est présent
            if " Inc." in nameTicker:
                nameTicker = nameTicker.replace(" Inc.", "")
        else:
            raise ValueError(f"Nom du ticker non trouvé dans le fichier {filePath}.")
        
        # Créer le nouveau nom de fichier
        newFileName = f"{nameTicker}_{nouvelleDateStr}.pdf"
        try:
            # Déplacer et renommer le fichier dans le répertoire detailsFolder
            os.rename(filePath, os.path.join(detailsFolder, newFileName))
        except Exception as e:
            print(f"Une erreur s'est produite lors du déplacement du fichier {filePath}: {e}")

    def RenameAndMoveOrdresVentes(self, filePath: str):
        """
        Renomme et déplace un fichier PDF basé sur les informations extraites du texte du fichier.
        Le nouveau nom de fichier est basé sur la date et le nom du ticker extraits du texte.
        Le fichier est déplacé vers un dossier spécifique en fonction du type d'information.

        Args:
            filePath (str): Le chemin complet du fichier PDF à renommer et déplacer.
        """

        # Assertions pour vérifier les types des arguments
        assert isinstance(filePath, str) and os.path.isfile(filePath), f"Le fichier '{filePath}' n'existe pas ou n'est pas une extension .pdf"

        # Extraire le texte du fichier PDF
        text = self.ExtractInformationPdf(filePath)

        # Extraire la date du texte
        dateMatch = re.search(r'DATE (\d{2}/\d{2}/\d{4})', text)
        assert dateMatch, f"Date non trouvée dans le texte du fichier {filePath}"
        dateStr = dateMatch.group(1)
        dateObj = datetime.strptime(dateStr, "%d/%m/%Y")
        nouvelleDateStr = dateObj.strftime("%Y-%m-%d")

        # Extraire le nom du ticker
        targetPhrase = "TITRE ORDRE / QUANTITÉ VALEUR"
        pattern = re.escape(targetPhrase) + r'(.*?)ISIN'
        match = re.search(pattern, text, re.DOTALL)
        if match:
            nameTicker = match.group(1).strip()
        else:
            targetPhrase = "POSITION QUANTITÉ PRIX MONTANT"
            pattern = re.escape(targetPhrase) + r'\s+(\b\w+\b)'
            match = re.search(pattern, text, re.DOTALL)
            assert match, f"Nom du ticker non trouvé dans le texte du fichier {filePath}"
            nameTicker = match.group(1).strip()

        # Vérifier si les informations nécessaires ont été trouvées
        assert nameTicker, f"Nom du ticker manquant dans le fichier {filePath}"
        assert nouvelleDateStr, f"Date de renommage manquante pour le fichier {filePath}"

        # Déterminer le dossier de destination et le nom du fichier
        if re.search(re.escape("INFORMATIONS SUR LES COÛTS EX-ANTE DE LA VENTE DE TITRES"), text):
            newFileName = f"{nameTicker}_{nouvelleDateStr}_Informations_sur_les_coûts.pdf"
            detailsFolder = os.path.join('Code Python/Annexes/Ventes en Bourse/InformationsCoûts')
        else:
            newFileName = f"{nameTicker}_{nouvelleDateStr}_Facture_Vente.pdf"
            detailsFolder = os.path.join('Code Python/Annexes/Ventes en Bourse/FacturesVentes')

        # Assurer l'existence du dossier de destination
        os.makedirs(detailsFolder, exist_ok=True)

        # Déplacer et renommer le fichier
        newFilePath = os.path.join(detailsFolder, newFileName)
        try:
            os.rename(filePath, newFilePath)
        except FileNotFoundError as e:
            print(f"Erreur : {filePath} non trouvé. Détails : {e}")
        except PermissionError as e:
            print(f"Erreur : Permission refusée pour {filePath}. Détails : {e}")
        except Exception as e:
            print(f"Erreur inattendue lors du déplacement de {filePath}. Détails : {e}")

        print(f"Fichier renommé et déplacé de {filePath} vers {newFilePath}")


    @staticmethod
    def DataDepotArgent(pdfFiles: list, chemin: str) -> pd.DataFrame:
        """
        Extrait les données des fichiers PDF spécifiés et les retourne sous forme de DataFrame.
        
        Args:
            pdfFiles: Liste des noms de fichiers PDF à traiter.
            chemin: Chemin du répertoire contenant les fichiers PDF.
            
        Returns:
            pd.DataFrame: DataFrame contenant les données extraites de chaque fichier PDF.
        """
        # Assertions pour vérifier les types des entrées
        assert isinstance(pdfFiles, list) and all(fichier.endswith(".pdf") for fichier in pdfFiles), \
            f"pdfFiles doit être une liste dont Le contenu contient des chaînes de caractères avec l'extension .pdf: {pdfFiles}"
        assert isinstance(chemin, str) and os.path.isdir(chemin), f"Le dossier '{chemin}' n'existe pas"

        df = pd.DataFrame(columns=["Date de valeur", "Prix dépôt brut", "Frais", "Prix dépôt net", "Compte-titres", "Compte-espèces"])

        for fichier in pdfFiles:
            cheminComplet = os.path.join(chemin, fichier)  # Chemin complet du fichier PDF
            
            try:
                with pdfplumber.open(cheminComplet) as pdf:
                    page = pdf.pages[0]
                    text = page.extract_text()

                # Extraction des données
                compteTitrePattern = re.compile(r"(?<=49170 Saint-Georges-sur-Loire COMPTE-TITRES ).+")
                compteTitreMatch = compteTitrePattern.search(text)
                assert compteTitreMatch is not None, f"Compte titre non trouvé dans {fichier}"
                compteTitre = int(compteTitreMatch.group(0))
                assert isinstance(compteTitre, int), f"Le compte titre doit être un entier dans {fichier}"

                achatPattern = re.compile(r"(.+)(?=\s*DÉTAIL)")
                achatMatch = achatPattern.search(text)
                assert achatMatch is not None, f"Achat non trouvé dans {fichier}"
                achat = achatMatch.group(0)[6:-4].replace(',', '.')
                achat = float(achat)
                assert isinstance(achat, float), f"L'achat doit être un nombre flottant dans {fichier}"

                fraisPattern = re.compile(r"(?<=Frais de paiements par carte de crédit ).+")
                fraisMatch = fraisPattern.search(text)
                assert fraisMatch is not None, f"Frais non trouvés dans {fichier}"
                frais = fraisMatch.group(0)[:-4].replace(',', '.')
                frais = float(frais)
                assert isinstance(frais, float), f"Les frais doivent être un nombre flottant dans {fichier}"

                dataPattern = re.compile(r"(?<=COMPTE-ESPÈCES DATE DE VALEUR MONTANT\n).+")
                dataMatch = dataPattern.search(text)
                assert dataMatch is not None, f"Informations sur le compte espèces non trouvées dans {fichier}"
                titre = dataMatch.group(0).split()
                
                compteEspece = titre[0]
                assert isinstance(compteEspece, str), f"Le compte espèces doit être une chaîne de caractères dans {fichier}"
                
                dateStr = titre[1]
                # Convertir la chaîne de date en objet datetime
                dateObj = datetime.strptime(dateStr, "%d/%m/%Y")
                # Définir la date de référence
                dateReference = datetime(1900, 1, 1)
                # Calculer le nombre de jours écoulés depuis la date de référence
                dateValeur = (dateObj - dateReference).days
                assert isinstance(dateValeur, int), f"La date valeur doit être un entier dans {fichier}"

                montant = titre[2].replace(',', '.')
                montant = float(montant)
                assert isinstance(montant, float), f"Le montant doit être un nombre flottant dans {fichier}"

                # Ajouter les données au DataFrame
                newData = {
                    "Date de valeur": dateValeur,
                    "Prix dépôt brut": achat,
                    "Frais": frais,
                    "Prix dépôt net": montant,
                    "Compte-titres": compteTitre,
                    "Compte-espèces": compteEspece
                }
                
                # Ajouter la nouvelle ligne au DataFrame
                df = pd.concat([df, pd.DataFrame([newData])], ignore_index=True)
                

            except AssertionError as e:
                print(f"AssertionError: {e}")
            except Exception as e:
                print(f"Erreur lors du traitement du fichier {fichier}: {e}")
            

        nomFeuille = "Dépôts Argents"
        colonnesEuros = ["Prix dépôt brut", "Frais", "Prix dépôt net"]
        colonnesPourcentages = []
        colonnesDates = ["Date de valeur"]
        appliquerTableau = True
        
        return df, nomFeuille, colonnesEuros, colonnesPourcentages, colonnesDates, appliquerTableau

    @staticmethod
    def DataDividendes(pdfFiles: list, chemin: str) -> pd.DataFrame:
        """
        Extrait les données des fichiers PDF spécifiés et les retourne sous forme de DataFrame.
        
        Args:
            pdfFiles: Liste des noms de fichiers PDF à traiter.
            chemin: Chemin du répertoire contenant les fichiers PDF.
            
        Returns:
            pd.DataFrame: DataFrame contenant les données extraites de chaque fichier PDF.
        """
        # Assertions pour vérifier les types des entrées
        assert isinstance(pdfFiles, list) and (fichier.endswith(".pdf") for fichier in pdfFiles), \
            f"pdfFiles doit être une liste dont Le contenu contient des chaînes de caractères avec l'extension .pdf: {pdfFiles}"
        assert isinstance(chemin, str) and os.path.isdir(chemin), f"Le dossier '{chemin}' n'existe pas"

        df = pd.DataFrame(columns=[
            "Date de valeur", "TICKER", "Dividendes brut", "Impôt à la source",
            "Dividendes net", "Titre(s) détenue(s)", "Dividende à la date du ...", "ISIN",
            "COMPTE-TITRE", "COMPTE ESPÈCES"
        ])

        for fichier in pdfFiles:
            cheminComplet = os.path.join(chemin, fichier)  # Obtenez le chemin complet du fichier PDF

            # Assurez-vous que le fichier PDF existe
            assert os.path.isfile(cheminComplet), f"Le fichier {cheminComplet} n'existe pas."

            with pdfplumber.open(cheminComplet) as pdf:
                page = pdf.pages[0]
                text = page.extract_text()

            # Définir des expressions régulières pour extraire les données
            dateValeurRe = re.compile(r"(?<=COMPTE-ESPÈCES DATE DE VALEUR MONTANT\n).+")
            dateValeurRe1 = re.compile(r"(?<=COMPTE DE RÈGLEMENT DATE DE PAIEMENT MONTANT\n).+")

            titre = dateValeurRe.search(text)
            if titre:
                titre = titre.group(0)
                dateValeurReFinal = re.search(r"\d{2}/\d{2}/\d{4}", titre).group(0)
            elif dateValeurRe1:
                titre = dateValeurRe1.search(text)
                assert titre is not None, "Date de valeur non trouvée dans le texte."
                dateValeurReFinal = re.search(r"\b\d{2}\.\d{2}\.\d{4}\b", text).group(0).replace(".", "/")


            # Convertir la chaîne de date en objet datetime
            dateObj = datetime.strptime(dateValeurReFinal, "%d/%m/%Y")
            # Définir la date de référence
            dateReference = datetime(1900, 1, 1)
            # Calculer le nombre de jours écoulés depuis la date de référence
            dateValeur = (dateObj - dateReference).days

            dateDividendeRe = re.compile(r"(?<=Dividende à la date du ).+")
            dateDividendeReText = dateDividendeRe.search(text)
            if dateDividendeReText:
                dateDividendeRe = dateDividendeReText.group(0)[:-1]
                # Convertir la chaîne de date en objet datetime
                dateObj = datetime.strptime(dateDividendeRe, "%d/%m/%Y")
                # Calculer le nombre de jours écoulés depuis la date de référence
                dateDividende = (dateObj - dateReference).days
            else:
                # Nouvelle option : extraire la date sous le format "Dividende en espèces avec la date d'exécution au -16.08.2024"
                dateExecutionRe = re.compile(r"date d'exécution au\s*-\s*(\d{2}\.\d{2}\.\d{4})")
                dateExecutionMatch = dateExecutionRe.search(text)
                if dateExecutionMatch:
                    dateDividendeStr = dateExecutionMatch.group(1)
                    # Convertir la chaîne de date au format "dd.mm.yyyy" en objet datetime
                    dateObj = datetime.strptime(dateDividendeStr, "%d.%m.%Y")
                    # Calculer le nombre de jours écoulés depuis la date de référence
                    dateDividende = (dateObj - dateReference).days
                else:
                    dateDividende = None


            compteTitreRe1 = re.compile(r"(?<=49170 Saint-Georges-sur-Loire COMPTE-TITRES ).+")
            compteTitreRe2 = re.compile(r"(?<=49170 Saint-Georges-sur-Loire).+")
            if (compteTitreRe1.search(text) is not None):
                compteTitre = compteTitreRe1.search(text)
                assert (compteTitre is not None), "Compte titre non trouvé dans le texte."
            else:
                compteTitre = compteTitreRe2.search(text)
                assert (compteTitre is not None), "Compte titre non trouvé dans le texte."

            compteTitre = compteTitre.group(0)
            compteTitre = compteTitre.replace(',', '.')
            compteTitre = float(compteTitre)

            compteEspeceRe = re.compile(r"(?<=COMPTE-ESPÈCES DATE DE VALEUR MONTANT\n).+")
            compteEspece = compteEspeceRe.search(text)
            if compteEspece:
                compteEspece = compteEspece.group(0)
                compteEspece = re.search(r"DE\d{20}", compteEspece).group(0)
            else:
                compteEspeceRe = re.compile(r"(?<=COMPTE DE RÈGLEMENT DATE DE PAIEMENT MONTANT\n).+")
                compteEspece = compteEspeceRe.search(text)
                assert compteEspece is not None, "Compte espèces non trouvé dans le texte."
                compteEspece = compteEspece.group(0)
                compteEspece = re.search(r"DE\d{20}", compteEspece).group(0)

            # Extraire ISIN
            isinRe = re.compile(r"(?<=ISIN : ).+")
            isinMatch = isinRe.search(text)
            if isinMatch:
                isin = isinMatch.group(0)
            else:
                isinRe = re.compile(r"(?<=\n)([A-Z0-9]{12})\s+")
                isinMatch = isinRe.search(text)
                isin = isinMatch.group(0).strip() if isinMatch else None


            # Extraire Titre
            titreRe = re.compile(r"POSITION QUANTITÉ REVENU MONTANT\n(.+)")
            titreRe1 = re.compile(r"POSITION QUANTITÉ TAUX MONTANT\n(.+)")
            titreMatch = titreRe.search(text)
            titre = titreMatch.group(1) if titreMatch else titreRe1.search(text).group(1) if titreRe1.search(text) else None

            # Extraire Ticker et Titre Detenue
            ticker = None
            titreDetenue = None
            if titre:
                tickerMatch = re.search(r"([^0-9]+)(?=\d)", titre)
                ticker = tickerMatch.group(0) if tickerMatch else None

                titreDetenueMatch = re.search(r"\d,\d+", titre)
                if titreDetenueMatch:
                    titreDetenue = titreDetenueMatch.group(0).replace(',', '.')
                    titreDetenue = float(titreDetenue)
                else:
                    titreRe1 = re.compile(r"\d+,\d+\n(.+)")
                    titreDetenueMatch = titreRe1.search(text)
                    if titreDetenueMatch:
                        titreDetenue = float(titreDetenueMatch.group(1).split()[0])
                    else:
                        # Si les autres méthodes échouent, chercher la quantité sous la forme "0.856462 unit"
                        quantiteMatch = re.search(r"(\d+\.\d+)\s*unit", text)
                        if quantiteMatch:
                            titreDetenue = float(quantiteMatch.group(1))

            # Extraire Dividendes net
            montantGagneRe = re.compile(r"(?<=COMPTE-ESPÈCES DATE DE VALEUR MONTANT\n).+")
            montantGagne = montantGagneRe.search(text)
            if montantGagne:
                montantGagne = montantGagne.group(0).split()
                montantGagne = montantGagne[2]
                montantGagne = montantGagne.replace(',', '.')
                montantGagne = float(montantGagne)
            else:
                montantGagneRe = re.compile(r"COMPTE DE RÈGLEMENT DATE DE PAIEMENT MONTANT\n.+")
                montantGagneLigne = montantGagneRe.search(text).group(0).split("\n")[1]
                montantGagne = montantGagneLigne.split()[2]
                montantGagne = montantGagne.replace(',', '.')
                montantGagne = float(montantGagne)

            # Extraire Montant Avant Frais
            montantAvantFraisRe = re.compile(r"(.+)(?=\s*DÉTAIL)")
            montantAvantFraisText = montantAvantFraisRe.search(text)
            if montantAvantFraisText:
                montantAvantFrais = montantAvantFraisText.group(0).split()
                if len(montantAvantFrais) > 2:
                    montantAvantFrais = float(montantAvantFrais[1].replace(',', '.'))
            else:
                montantAvantFrais = montantGagne

            # Extraire Impôt
            impotRe = re.compile(r"(?<=Impôt à la source ).+")
            impotRe2 = re.compile(r"(?<=Impôt à la source pour les émetteurs américains ).+")
            impotText = impotRe2.search(text) or impotRe.search(text)
            if impotText:
                impot = impotText.group(0)[:-4]
                if ',' in impot:
                    impot = impot.replace(',', '.')
                newImpot = float(impot)
            else:
                newImpot = 0.0


            newData = {
                "Date de valeur": dateValeur,
                "TICKER": ticker,
                "Dividendes brut": montantAvantFrais,
                "Impôt à la source": newImpot,
                "Dividendes net": montantGagne,
                "Titre(s) détenue(s)": titreDetenue,
                "Dividende à la date du ...": dateDividende,
                "ISIN": isin,
                "COMPTE-TITRE": compteTitre,
                "COMPTE ESPÈCES": compteEspece
            }

            # Ajouter la nouvelle ligne au DataFrame
            df = pd.concat([df, pd.DataFrame([newData])], ignore_index=True)


        nomFeuille = "Dividendes"
        colonnesEuros = ["Dividendes brut", "Impôt à la source", "Dividendes net"]
        colonnesPourcentages = []
        colonnesDates = ["Date de valeur", "Dividende à la date du ..."]
        appliquerTableau = True
        
        return df, nomFeuille, colonnesEuros, colonnesPourcentages, colonnesDates, appliquerTableau

    @staticmethod
    def DataInteret(pdf: list, chemin: str) -> pd.DataFrame:
        """
        Extrait les données des fichiers PDF spécifiés et les retourne sous forme de DataFrame.
        
        Args:
            pdfFiles: Liste des noms de fichiers PDF à traiter.
            chemin: Chemin du répertoire contenant les fichiers PDF.
            
        Returns:
            pd.DataFrame: DataFrame contenant les données extraites de chaque fichier PDF.
        """
        # Assertions pour vérifier les types des entrées
        assert isinstance(pdf, list) and (fichier.endswith(".pdf") for fichier in pdf), \
            f"pdfFiles doit être une liste dont Le contenu contient des chaînes de caractères avec l'extension .pdf: {pdf}"
        assert isinstance(chemin, str) and os.path.isdir(chemin), f"Le dossier '{chemin}' n'existe pas"

        df = pd.DataFrame(columns=["Date d'effet", "Taux d'interêts", "Interêts net", "Plage de date", "Nature revenus", "Actifs", "Iban", "Compte-espèces"])

        for fichier in pdf:
            cheminComplet = os.path.join(chemin, fichier)  # Obtenez le chemin complet du fichier PDF

            # Assurez-vous que le fichier PDF existe
            assert os.path.isfile(cheminComplet), f"Le fichier {cheminComplet} n'existe pas."

            with pdfplumber.open(cheminComplet) as pdf:
                page = pdf.pages[0]
                text = page.extract_text()

            # Extraire compte espèces
            compteEspeceRe = re.compile(r"(?<=49170 Saint-Georges-sur-Loire COMPTE ESPÈCES ).+")
            compteEspeceMatch = compteEspeceRe.search(text)
            assert compteEspeceMatch is not None, "Compte espèces non trouvé dans le texte."
            compteEspece = int(compteEspeceMatch.group(0))

            # Extraire les données principales
            dataRe = re.compile(r"(?<=ACTIFS NATURE DES REVENUS TAUX D'INTÉRÊTS TOTAL\n).+")
            dataMatch = dataRe.search(text)
            if dataMatch:
                titre = dataMatch.group(0).split()
                assert len(titre) >= 4, "Le format des données est incorrect."
                espece = titre[0]
                interet = titre[1]
                pourcentage = float(titre[2][:-1].replace(',', '.')) / 100
                prix1 = float(titre[3].replace(',', '.'))
                plageDate = None
            else:
                dataRe = re.compile(r"(?<=ACTIFS NATURE DES REVENUS TAUX D'INTÉRÊTS DATE TOTAL\n).+")
                dataMatch = dataRe.search(text)
                assert dataMatch is not None, "Données principales non trouvées dans le texte."
                titre = dataMatch.group(0).split()
                assert len(titre) >= 7, "Le format des données est incorrect."
                espece = titre[0]
                interet = titre[1]
                pourcentage = float(titre[2][:-1].replace(',', '.')) / 100
                prix1 = float(titre[6].replace(',', '.'))
                plageDate = f"{titre[3]} - {titre[5]}"

            # Extraire IBAN et autres détails
            ibanRe = re.compile(r"(?<=IBAN DATE D'EFFET TOTAL\n).+")
            ibanMatch = ibanRe.search(text)
            assert ibanMatch is not None, "IBAN non trouvé dans le texte."
            titre = ibanMatch.group(0).split()
            assert len(titre) >= 3, "Le format des données IBAN est incorrect."
            iban = titre[0]

            dateStr = titre[1]
            # Convertir la chaîne de date en objet datetime
            dateObj = datetime.strptime(dateStr, "%d/%m/%Y")
            # Définir la date de référence
            dateReference = datetime(1900, 1, 1)
            # Calculer le nombre de jours écoulés depuis la date de référence
            dateEffet = (dateObj - dateReference).days

            prix2 = float(titre[2].replace(',', '.'))

            newData = {
                "Date d'effet": dateEffet,
                "Taux d'interêts": pourcentage,
                "Interêts net": prix2,
                "Plage de date": plageDate,
                "Nature revenus": interet,
                "Actifs": espece,
                "Iban": iban,
                "Compte-espèces": compteEspece
            }

            compteEspece, iban, plageDate, dateEffet, espece, interet, pourcentage, prix2

            # Ajouter la nouvelle ligne au DataFrame
            df = pd.concat([df, pd.DataFrame([newData])], ignore_index=True)


        nomFeuille = "Interêts"
        colonnesEuros = ["Interêts net"]
        colonnesPourcentages = ["Taux d'interêts"]
        colonnesDates = ["Date d'effet"]
        appliquerTableau = True
        
        return df, nomFeuille, colonnesEuros, colonnesPourcentages, colonnesDates, appliquerTableau

    @staticmethod
    def DataOrdresAchats(pdfFiles: list, chemin: str) -> pd.DataFrame:
        """
        Extrait les données des fichiers PDF spécifiés et les retourne sous forme de DataFrame.
        
        Args:
            pdfFiles: Liste des noms de fichiers PDF à traiter.
            chemin: Chemin du répertoire contenant les fichiers PDF.
            
        Returns:
            pd.DataFrame: DataFrame contenant les données extraites de chaque fichier PDF.
        """
        # Assertions pour vérifier les types des entrées
        assert isinstance(pdfFiles, list) and (fichier.endswith(".pdf") for fichier in pdfFiles), \
            f"pdfFiles doit être une liste dont Le contenu contient des chaînes de caractères avec l'extension .pdf: {pdfFiles}"
        assert isinstance(chemin, str) and os.path.isdir(chemin), f"Le dossier '{chemin}' n'existe pas"

        df = pd.DataFrame(columns=["Date de valeur", "TICKER", "Montant investi", "Titre(s) détenue(s)", "COURS MOYEN", "COMPTE-TITRE", "ISIN", "COMPTE ESPÈCES"])

        for fichier in pdfFiles:
            cheminComplet = os.path.join(chemin, fichier)  # Obtenez le chemin complet du fichier PDF
            
            with pdfplumber.open(cheminComplet) as pdf:
                page = pdf.pages[0]
                text = page.extract_text() or ""

            # Extraction de la date de valeur
            dateValeurRe = re.compile(r"(?<=COMPTE-ESPÈCES DATE DE VALEUR MONTANT\n).+")
            titreMatch = dateValeurRe.search(text)
            if titreMatch:
                titre = titreMatch.group(0)
                dateValeurRe = re.search(r"\d{2}/\d{2}/\d{4}", titre)
                if dateValeurRe:
                    dateValeurRe = dateValeurRe.group(0)
                else:
                    dateValeurRe = re.search(r"\d{4}-\d{2}-\d{2}", titre)
                    if dateValeurRe:
                        dateValeurRe = dateValeurRe.group(0).replace('-', '/')
                        dateValeurRe = dateValeurRe[-2:] + dateValeurRe[-6:-2] + dateValeurRe[:4]
            else:
                dateValeurRe = ""
            
            # Convertir la chaîne de date en objet datetime
            try:
                dateObj = datetime.strptime(dateValeurRe, "%d/%m/%Y")
                dateValeur = (dateObj - datetime(1900, 1, 1)).days
            except ValueError:
                dateValeur = None

            # Extraction du compte titre
            compteTitreRe = re.compile(r"COMPTE-TITRES.+")
            compteTitreMatch = compteTitreRe.search(text)
            if compteTitreMatch:
                compteTitre = compteTitreMatch.group(0)[14:]
                compteTitre = compteTitre.replace(',', '.')
                compteTitre = float(compteTitre)
            else:
                compteTitre = None

            # Extraction du compte espèces
            compteEspeceRe = re.compile(r"(?<=COMPTE-ESPÈCES DATE DE VALEUR MONTANT\n).+")
            compteEspeceMatch = compteEspeceRe.search(text)
            if compteEspeceMatch:
                compteEspece = re.search(r"DE\d{20}", compteEspeceMatch.group(0))
                compteEspece = compteEspece.group(0) if compteEspece else None
            else:
                compteEspece = None

            # Extraction de l'ISIN
            isinRe = re.compile(r"(?<=ISIN : ).+")
            isinMatch = isinRe.search(text)
            isin = isinMatch.group(0) if isinMatch else None

            # Extraction du titre détenu
            titreRe = re.compile(r"(?<=POSITION QUANTITÉ COURS MOYEN MONTANT\n).+")
            titreMatch = titreRe.search(text)
            if titreMatch:
                titre = titreMatch.group(0)
                tickerMatch = re.search(r"([^0-9]+)(?=\d)", titre)
                ticker = tickerMatch.group(0) if tickerMatch else None

                titreDetenueMatch = re.search(r"\d,\d+", titre)
                titreDetenue = titreDetenueMatch.group(0).replace(',', '.') if titreDetenueMatch else None
                titreDetenue = float(titreDetenue) if titreDetenue else None

                montantPattern = re.compile(r'\b\d{1,6}(?:,\d{2})? EUR\b')
                montantMatch = montantPattern.search(titre)
                montantAchat = float(montantMatch.group(0).replace(",", ".")[:-4]) if montantMatch else None
            else:
                ticker = None
                titreDetenue = None
                montantAchat = None

            # Extraction de l'impôt
            impotRe = re.compile(r"(?<=Impôt à la source ).+")
            impotRe2 = re.compile(r"(?<=Impôt à la source pour les émetteurs américains ).+")
            impotMatch = impotRe2.search(text) or impotRe.search(text)
            impot = impotMatch.group(0)[:-4] if impotMatch else "0.00"
            impot = float(impot.replace(',', '.'))

            # Extraction du montant gagné
            montantGagneRe = re.compile(r"(?<=COMPTE-ESPÈCES DATE DE VALEUR MONTANT\n).+")
            montantGagneMatch = montantGagneRe.search(text)
            if montantGagneMatch:
                montantGagne = montantGagneMatch.group(0).split()
                montantGagne = montantGagne[2]
                montantGagne = float(montantGagne.replace(',', '.'))
            else:
                montantGagne = None

            newData = {
                "Date de valeur": dateValeur,
                "TICKER": ticker,
                "Montant investi": montantGagne,
                "Titre(s) détenue(s)": titreDetenue,
                "COURS MOYEN": montantAchat,
                "ISIN": isin,
                "COMPTE-TITRE": compteTitre,
                "COMPTE ESPÈCES": compteEspece
            }

            # Ajouter la nouvelle ligne au DataFrame
            df = pd.concat([df, pd.DataFrame([newData])], ignore_index=True)

        
        # Un fichier PDF n'a pas pu être téléchargé sur Trade République c'est l'investissement d'Apple donc il faut le rajouter (supprimer pour votre portefeuille)*
        newData = {
            "Date de valeur": 45154,
            "TICKER": "Apple Inc.",
            "Montant investi": -4.00,
            "Titre(s) détenue(s)": 0.02457,
            "COURS MOYEN": 162.80,
            "ISIN": "US0378331005",
            "COMPTE-TITRE": 700305101,
            "COMPTE ESPÈCES": "DE83502109007011655011",
        }
        df = pd.concat([df, pd.DataFrame([newData])], ignore_index=True)


        nomFeuille = "Ordres d'Achats"
        colonnesEuros = ["Montant investi", "COURS MOYEN"]
        colonnesPourcentages = []
        colonnesDates = ["Date de valeur"]
        appliquerTableau = True
        
        return df, nomFeuille, colonnesEuros, colonnesPourcentages, colonnesDates, appliquerTableau

    @staticmethod
    def DataOrdresVentes(pdfFiles: list, chemin: str) -> pd.DataFrame:
        """
        Extrait les données des fichiers PDF spécifiés et les retourne sous forme de DataFrame.
        
        Args:
            pdfFiles: Liste des noms de fichiers PDF à traiter.
            chemin: Chemin du répertoire contenant les fichiers PDF.
            
        Returns:
            pd.DataFrame: DataFrame contenant les données extraites de chaque fichier PDF.
        """
        # Assertions pour vérifier les types des entrées
        assert isinstance(pdfFiles, list) and (fichier.endswith(".pdf") for fichier in pdfFiles), \
            f"pdfFiles doit être une liste dont Le contenu contient des chaînes de caractères avec l'extension .pdf: {pdfFiles}"
        assert isinstance(chemin, str) and os.path.isdir(chemin), f"Le dossier '{chemin}' n'existe pas"

        df = pd.DataFrame(columns=["Date de valeur", "TICKER", "Montant gagné", "Titre(s) détenue(s)", "ISIN", "COMPTE-TITRE", "COMPTE ESPÈCES"])

        for fichier in pdfFiles:
            cheminComplet = os.path.join(chemin, fichier)  # Obtenez le chemin complet du fichier PDF
            with pdfplumber.open(cheminComplet) as pdf:
                page = pdf.pages[0]
                text = page.extract_text()

            # Extraire et formater la date
            dataRe = re.search(r"(?<=COMPTE-ESPÈCES DATE DE VALEUR MONTANT\n).+", text)
            assert dataRe, f"Date non trouvée dans le fichier {fichier}"
            titre = dataRe.group(0)
            
            try:
                dateRe = re.search(r"\d{2}/\d{2}/\d{4}", titre).group(0)
            except AttributeError:
                dateRe = re.search(r"\d{4}-\d{2}-\d{2}", titre).group(0).replace("-", "/")
                dateRe = dateRe[-2:] + dateRe[4:-2] + dateRe[:4]
            
            # Extraire et formater les autres données
            compteEspece = re.search(r"DE\d{20}", titre).group(0)
            montantGagne = re.search(r"(?<=COMPTE-ESPÈCES DATE DE VALEUR MONTANT\n).+", text).group(0).split()[2]
            montantGagne = float(montantGagne.replace(',', '.'))
            
            # Convertir la chaîne de date en objet datetime
            dateObj = datetime.strptime(dateRe, "%d/%m/%Y")
            dateReference = datetime(1900, 1, 1)
            date = (dateObj - dateReference).days

            compteTitre = re.search(r"COMPTE-TITRES.+", text).group(0)[14:]
            compteTitre = float(compteTitre.replace(',', '.'))
            
            try:
                isin = re.search(r"(?<=ISIN : ).+", text).group(0)
            except AttributeError:
                isin = None
            
            titreRe = re.search(r"(?<=POSITION QUANTITÉ PRIX MONTANT\n).+", text)
            titre = titreRe.group(0) if titreRe else None

            try:
                ticker = re.search(r"([^0-9]+)(?=\d)", titre).group(0) if titre else None
            except AttributeError:
                ticker = None
            
            try:
                titreDetenue = re.search(r"\d,\d+", titre).group(0).replace(',', '.')
                titreDetenue = float(titreDetenue)
            except AttributeError:
                titreDetenue = None

            newData = {
                "Date de valeur": date,
                "TICKER": ticker,
                "Montant gagné": montantGagne,
                "Titre(s) détenue(s)": titreDetenue,
                "ISIN": isin,
                "COMPTE-TITRE": compteTitre,
                "COMPTE ESPÈCES": compteEspece
            }

            # Ajouter la nouvelle ligne au DataFrame
            df = pd.concat([df, pd.DataFrame([newData])], ignore_index=True)


        nomFeuille = "Ordres de Ventes"
        colonnesEuros = ["Montant gagné"]
        colonnesPourcentages = []
        colonnesDates = ["Date de valeur"]
        appliquerTableau = True
        
        return df, nomFeuille, colonnesEuros, colonnesPourcentages, colonnesDates, appliquerTableau

    @staticmethod
    def DataRetraitArgent(pdfFiles: list, chemin: str) -> pd.DataFrame:
        """
        Extrait les données des fichiers PDF spécifiés et les retourne sous forme de DataFrame.
        
        Args:
            pdfFiles: Liste des noms de fichiers PDF à traiter.
            chemin: chemin du répertoire contenant les fichiers PDF.
            
        Returns:
            pd.DataFrame: DataFrame contenant les données extraites de chaque fichier PDF.
        """
        # Assertions pour vérifier les types des entrées
        assert isinstance(pdfFiles, list) and (fichier.endswith(".pdf") for fichier in pdfFiles), \
            f"pdfFiles doit être une liste dont Le contenu contient des chaînes de caractères avec l'extension .pdf: {pdfFiles}"
        assert isinstance(chemin, str) and os.path.isdir(chemin), f"Le dossier '{chemin}' n'existe pas"

        df = pd.DataFrame(columns=["Date de la commande", "CE QUE L'ON A ACHETE", "Retraits net", "FRAIS", "QUANTITE", "Date de facture", "NUMERO DE FACTURE"])

        for fichier in pdfFiles:
            cheminComplet = os.path.join(chemin, fichier)  # Obtenez le chemin complet du fichier PDF
            with pdfplumber.open(cheminComplet) as pdf:
                page = pdf.pages[0]
                text = page.extract_text()

            # Extraction de la date de valeur
            dateValeurRe = re.compile(r"(?<=square Jean Gasiorowski 7 DATE ).+")
            titre = dateValeurRe.search(text).group(0).replace('.', '/')
            dateValeurReFinal = re.search(r"\d{2}/\d{2}/\d{4}", titre).group(0)

            # Convertir la chaîne de date en objet datetime
            dateObj = datetime.strptime(dateValeurReFinal, "%d/%m/%Y")
            # Définir la date de référence
            dateReference = datetime(1900, 1, 1)
            # Calculer le nombre de jours écoulés depuis la date de référence
            dateValeur = (dateObj - dateReference).days

            # Extraction du numéro de facture
            numeroFactureRe = re.compile(r"(?<=FACTURE N. ).+")
            numeroFacture = numeroFactureRe.search(text).group(0)

            # Extraction des détails de la commande
            dataRe = re.compile(r"(?<=POSITION COMMANDÉ LE QUANTITÉ\n).+")
            titre = dataRe.search(text).group(0).split()
            achat = ' '.join(titre[:-2])
            commandeDate = titre[-2].replace('.', '/')
            # Convertir la chaîne de date en objet datetime
            dateObj = datetime.strptime(commandeDate, "%d/%m/%Y")
            # Calculer le nombre de jours écoulés depuis la date de référence
            commandeDate = (dateObj - dateReference).days

            quantite = int(titre[-1])

            # Extraction du Retraits net et du type de paiement
            dataRe = re.compile(r"(?<=POSITION MONTANT\n).+")
            titre = dataRe.search(text).group(0).split()
            typePayment = titre[0]
            montant = float(titre[1].replace(',', '.'))

            # Assertions pour vérifier les types et valeurs attendus
            assert isinstance(numeroFacture, str), "Le numéro de facture doit être une chaîne de caractères"
            assert isinstance(dateValeur, int), "La date de valeur doit être un entier"
            assert isinstance(achat, str), "Le détail de l'achat doit être une chaîne de caractères"
            assert isinstance(commandeDate, int), "La date de commande doit être un entier"
            assert isinstance(quantite, int), "La quantité doit être un entier"
            assert isinstance(typePayment, str), "Le type de paiement doit être une chaîne de caractères"
            assert isinstance(montant, float), "Le montant doit être un nombre flottant"

            newData = {
                "Date de la commande": commandeDate,
                "CE QUE L'ON A ACHETE": achat,
                "Retraits net": montant,
                "FRAIS": typePayment,
                "QUANTITE": quantite,
                "Date de facture": dateValeur,
                "NUMERO DE FACTURE": numeroFacture
            }

            # Ajouter la nouvelle ligne au DataFrame
            df = pd.concat([df, pd.DataFrame([newData])], ignore_index=True)


        nomFeuille = "Retrait d'argent"
        colonnesEuros = ["Retraits net", "FRAIS"]
        colonnesPourcentages = []
        colonnesDates = ["Date de la commande", "Date de facture"]
        appliquerTableau = True
        
        return df, nomFeuille, colonnesEuros, colonnesPourcentages, colonnesDates, appliquerTableau


    @staticmethod
    def ProcessPdf(nameDirectory: str, cheminDossierRenommer: str, nameFunction) -> None:
        """
        Parcourt les fichiers PDF dans un répertoire, puis utilise une fonction spécifiée pour traiter chaque fichier PDF et le déplacer.

        Args:
            nameDirectory: Chemin du répertoire contenant les fichiers PDF à traiter.
            cheminDossierRenommer: Chemin du répertoire où les fichiers traités seront déplacés.
            nameFunction: Fonction à appeler pour chaque fichier PDF. La fonction doit accepter deux arguments :
                        - Le chemin complet du fichier PDF
                        - Le chemin du répertoire de destination

        Returns:
            None
        """
        # Assertions pour vérifier les types des entrées
        assert isinstance(nameDirectory, str), f"nameDirectory doit être une chaîne de caractères: {nameDirectory}"
        assert isinstance(cheminDossierRenommer, str), f"cheminDossierRenommer doit être une chaîne de caractères: {cheminDossierRenommer}"
        assert callable(nameFunction), f"nameFunction doit être une fonction: {nameFunction}"

        # Créer le répertoire de destination s'il n'existe pas
        if not os.path.exists(cheminDossierRenommer):
            os.makedirs(cheminDossierRenommer)

        # Parcourir tous les fichiers dans le dossier spécifié
        for fileName in os.listdir(nameDirectory):
            # Vérifier si le fichier est un fichier PDF
            if fileName.lower().endswith('.pdf'):
                filePath = os.path.join(nameDirectory, fileName)
                # Appeler la fonction spécifiée pour traiter et déplacer le fichier
                nameFunction(filePath, cheminDossierRenommer)

    @staticmethod
    def ExtractInformationPdf(filePath: str) -> str:
        """
        Extrait le texte de chaque page d'un fichier PDF.

        Args:
            filePath: Chemin complet du fichier PDF dont le texte doit être extrait.

        Returns:
            str: Texte extrait de toutes les pages du fichier PDF. Retourne None si le fichier ne peut pas être lu.
        """
        # Assertions pour vérifier les types des entrées
        assert isinstance(filePath, str), f"filePath doit être une chaîne de caractères: {filePath}"
        
        # Initialiser une variable pour stocker le texte extrait
        text = ''
        
        try:
            # Ouvrir le fichier PDF
            with open(filePath, 'rb') as file:
                reader = PdfReader(file)
                # Extraire le texte de chaque page
                for page in reader.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text += page_text
        except Exception as e:
            print(f"Une erreur s'est produite lors de l'extraction du texte: {e}")
            return None
        
        return text
