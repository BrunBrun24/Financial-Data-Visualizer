import tkinter as tk
from tkinter import filedialog
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils.dataframe import dataframe_to_rows
import os



class ExcelReportGenerator:
    """
    La classe `ExcelReportGenerator` facilite la génération de rapports Excel à partir de données
    structurées sous forme de dictionnaires. Elle permet de sauvegarder les données dans un fichier Excel,
    d'ajouter des feuilles avec un formatage personnalisé, et de créer des bilans de revenus et de dépenses
    en regroupant les transactions par catégories.

    Attributs:
        - `dataDict` (dict): Dictionnaire contenant les données à enregistrer et à analyser.
        - `wb` (Workbook): Objet Workbook représentant le fichier Excel généré.
        - `outputFile` (str): Chemin du fichier Excel généré.

    Méthodes:
        - `__init__(self, dataDict)`: Initialise l'instance avec les données fournies et démarre le processus
          de création du fichier Excel.
        - `CreateFileExcel(self)`: Créer un fichier Excel.
        - `AddInFileExcel(self)`: Formate les feuilles Excel, applique des styles et ouvre le fichier généré.
        - `SheetBilan(self)`: Crée des bilans de revenus et de dépenses en les regroupant par sous-catégories.
        - `ConvertDates(self)`: Convertit les chaînes de dates au format 'YYYY-MM-DD' en objets `datetime`.
        - `InitializeDataFrame(self, columnNames, months)`: Initialise un DataFrame avec des colonnes représentant
          les mois et des lignes pour les catégories de revenus ou de dépenses.
        - `RevenueSubCategories(self, revenues)`: Crée un tableau des revenus par catégories et par mois.
        - `ExpenseSubCategories(self, expenses)`: Crée un tableau des dépenses par catégories et par mois.
        - `AddDataFrameToSheet(self, df, sheetName, startRow, spacing=5)`: Ajoute un DataFrame dans une feuille Excel
          avec un formatage personnalisé.
    """

    def __init__(self, dataDict, outputFile):
        """
        Initialise la classe avec un dictionnaire de données.

        Args:
            dataDict (dict): Dictionnaire contenant les données pour les feuilles Excel.
        """
        assert isinstance(dataDict, dict), f"dataDict doit être un dictionnaire, mais c'est {type(dataDict).__name__}."

        self.dataDict = dataDict
        self.outputFile = outputFile
        self.ConvertDates()
        self.wb = self.CreateFileExcel()

        if self.wb is not None:
            self.AddInFileExcel()

    def CreateFileExcel(self):
        """
        Créer un fichier Excel
        
        Returns:
            Workbook: Un objet Workbook de openpyxl représentant le fichier Excel.
        """

        self.CreateDirectories()
        # Créer le fichier Excel
        with pd.ExcelWriter(self.outputFile, engine='openpyxl') as writer:
            # Ajouter les feuilles de chaque catégorie
            for sheetName, data in self.dataDict.items():
                if data:
                    # Convertir la liste de dictionnaires en DataFrame
                    df = pd.DataFrame(data)
                    # Assurer que les en-têtes sont des chaînes de caractères
                    df.columns = [str(col) for col in df.columns]
                    # Écrire le DataFrame dans une feuille avec le nom correspondant à la clé
                    df.to_excel(writer, sheet_name=sheetName, index=False)

        # Charger le fichier Excel pour le formatage ultérieur
        return load_workbook(self.outputFile)


    def AddInFileExcel(self):
        """
        Applique des formats aux feuilles Excel, ajoute des tableaux avec styles, ajuste les largeurs de colonnes et ouvre le fichier final.
        
        Formate chaque feuille de calcul (excepté "Bilan Revenus" et "Bilan Dépenses") en ajoutant un tableau et en ajustant les largeurs de colonnes.
        Enfin, le fichier Excel est sauvegardé et ouvert.
        """
        self.SheetBilan()

        centerAlignment = Alignment(horizontal='center', vertical='center')
        
        # Formater les feuilles de données
        for sheetName in self.wb.sheetnames:
            if sheetName != "Bilan Revenus" and sheetName != "Bilan Dépenses":
                ws = self.wb[sheetName]
                min_col, min_row, max_col, max_row = 1, 1, ws.max_column, ws.max_row
                table_style = TableStyleInfo(
                    name='TableStyleMedium2', showFirstColumn=False,
                    showLastColumn=False, showRowStripes=True, showColumnStripes=True
                )
                safeTableName = f'Table_{sheetName}'.replace(" ", "_")
                table = Table(displayName=safeTableName, ref=f'A1:{chr(64 + max_col)}{max_row}')
                table.tableStyleInfo = table_style
                ws.add_table(table)

                for row in ws.iter_rows(min_row=2, max_row=max_row):
                    for cell in row:
                        if isinstance(cell.value, datetime):
                            cell.number_format = 'DD/MM/YYYY'
                        cell.alignment = centerAlignment

                for col in ws.columns:
                    maxLength = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > maxLength:
                                maxLength = len(cell.value)
                        except:
                            pass
                    adjusted_width = (maxLength + 10)
                    ws.column_dimensions[column].width = adjusted_width

        # Sauvegarder le fichier Excel avec le format appliqué
        self.wb.save(self.outputFile)

    def SheetBilan(self):
        """
        Crée et ajoute des sous-catégories pour les feuilles "Revenus" et "Dépenses".
        
        Regroupe les revenus et les dépenses en sous-catégories pour chaque type de transaction.
        """
        revenus = self.dataDict.get('Revenus', [])
        self.RevenueSubCategories(revenus)

        depenses = {cat: transactions for cat, transactions in self.dataDict.items() if cat != 'Revenus'}
        self.ExpenseSubCategories(depenses)

    def ConvertDates(self):
        """
        Convertit les dates dans les données en objets datetime.
        """
        for key in self.dataDict:
            for entry in self.dataDict[key]:
                for k, v in entry.items():
                    if isinstance(v, str) and len(v) == 10 and v.count('-') == 2:
                        try:
                            entry[k] = datetime.strptime(v, "%Y-%m-%d")
                        except ValueError:
                            pass

    def InitializeDataFrame(self, columnNames, months):
        """
        Initialise un DataFrame avec les colonnes représentant les mois et les index représentant les revenus/dépenses.

        Args:
            columnNames (list): Liste des noms de colonnes (revenus/dépenses).
            months (list): Liste des noms des mois.

        Returns:
            pd.DataFrame: DataFrame initialisé avec les colonnes et index spécifiés.
        """
        assert isinstance(columnNames, list), f"columnNames doit être une liste, mais c'est {type(columnNames).__name__}."
        assert isinstance(months, list), f"months doit être une liste, mais c'est {type(months).__name__}."
        
        data = {month: [0] * len(columnNames) for month in months}
        df = pd.DataFrame(data, index=columnNames)
        df[''] = ''

        return df

    def RevenueSubCategories(self, revenues):
        """
        Crée un DataFrame des revenus sous différentes catégories et mois.

        Args:
            revenues (list): Liste de transactions de revenus.
        """
        assert isinstance(revenues, list), f"revenues doit être une liste, mais c'est {type(revenues).__name__}."
        months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec', '']
        revenueTypes = ['Revenus']

        for transaction in revenues:
            categoryType = transaction.get('Type', 'Revenus')
            if categoryType not in revenueTypes:
                revenueTypes.append(categoryType)

        dfRevenues = self.InitializeDataFrame(columnNames=revenueTypes, months=months)

        for transaction in revenues:
            categoryType = transaction.get('Type', 'Revenus')
            amount = transaction['MONTANT']
            date = transaction["DATE D'OPÉRATION"]
            month = date.strftime('%b')
            dfRevenues.loc[categoryType, month] += amount

        dfRevenues.loc['Revenus'] = dfRevenues.sum(axis=0)
        dfRevenues['Total'] = dfRevenues.select_dtypes(include=['number']).sum(axis=1)
        dfRevenues['Pourcentage'] = round(dfRevenues['Total'] / dfRevenues.loc['Revenus', 'Total'], 4)

        dfRevenues = dfRevenues.sort_values(by='Pourcentage', ascending=False)

        self.AddDataFrameToSheet(dfRevenues, "Bilan Revenus", 1)

    def ExpenseSubCategories(self, expenses):
        """
        Crée un DataFrame des dépenses sous différentes catégories et mois.

        Args:
            expenses (dict): Dictionnaire des transactions de dépenses par catégorie.
        """
        assert isinstance(expenses, dict), f"expenses doit être un dictionnaire, mais c'est {type(expenses).__name__}."
        months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec', '']
        expenseTypes = ['Dépenses']

        for category, transactions in expenses.items():
            for transaction in transactions:
                amount = transaction['MONTANT']
                if category not in expenseTypes:
                    expenseTypes.append(category)

        dfExpenses = self.InitializeDataFrame(columnNames=expenseTypes, months=months)

        for category, transactions in expenses.items():
            for transaction in transactions:
                amount = transaction['MONTANT']
                date = transaction["DATE D'OPÉRATION"]
                month = date.strftime('%b')
                dfExpenses.loc[category, month] += amount

        dfExpenses.loc['Dépenses'] = dfExpenses.sum(axis=0)
        dfExpenses['Total'] = dfExpenses.select_dtypes(include=['number']).sum(axis=1)
        dfExpenses['Pourcentage'] = round(dfExpenses['Total'] / dfExpenses.loc['Dépenses', 'Total'], 4)

        dfExpenses = dfExpenses.sort_values(by='Pourcentage', ascending=False)

        self.AddDataFrameToSheet(dfExpenses, "Bilan Dépenses", 1)

        # Sous-catégories
        expenseTypes = ['Dépenses']
        for category, transactions in expenses.items():
            for transaction in transactions:
                amount = transaction['MONTANT']
                categoryType = transaction.get('Type', category)
                if categoryType not in expenseTypes:
                    expenseTypes.append(categoryType)

        dfExpenses = self.InitializeDataFrame(columnNames=expenseTypes, months=months)

        for category, transactions in expenses.items():
            for transaction in transactions:
                categoryType = transaction.get('Type', category)
                amount = transaction['MONTANT']
                date = transaction["DATE D'OPÉRATION"]
                month = date.strftime('%b')
                dfExpenses.loc[categoryType, month] += amount

        dfExpenses.loc['Dépenses'] = dfExpenses.sum(axis=0)
        dfExpenses['Total'] = dfExpenses.select_dtypes(include=['number']).sum(axis=1)
        dfExpenses['Pourcentage'] = round(dfExpenses['Total'] / dfExpenses.loc['Dépenses', 'Total'], 4)

        dfExpenses = dfExpenses.sort_values(by='Pourcentage', ascending=False)

        startRowNext = self.wb["Bilan Dépenses"].max_row + 2
        self.AddDataFrameToSheet(dfExpenses, "Bilan Dépenses", startRowNext)

    def AddDataFrameToSheet(self, df, sheetName, startRow, spacing=5):
        """
        Ajoute un DataFrame à une feuille Excel avec un formatage spécifique.

        Args:
            df (pd.DataFrame): DataFrame à ajouter à la feuille.
            sheetName (str): Nom de la feuille Excel.
            startRow (int): Ligne de départ pour ajouter les données.
            spacing (int): Espacement après le dernier bloc de données.
        """
        assert isinstance(df, pd.DataFrame), f"df doit être un DataFrame, mais c'est {type(df).__name__}."
        assert isinstance(sheetName, str), f"sheetName doit être une chaîne de caractères, mais c'est {type(sheetName).__name__}."
        assert isinstance(startRow, int), f"startRow doit être un entier, mais c'est {type(startRow).__name__}."
        assert isinstance(spacing, int), f"spacing doit être un entier, mais c'est {type(spacing).__name__}."

        if sheetName in self.wb.sheetnames:
            ws = self.wb[sheetName]
            startRow -= 1
        else:
            ws = self.wb.create_sheet(title=sheetName)
        
        for r in dataframe_to_rows(df, index=True, header=True):
            ws.append(r)

        # Définir les styles
        fontMonth = Font(color="FFFFFF", bold=True)
        fillMonth = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        alignmentCenter = Alignment(horizontal='center', vertical='center')
        fontRevenue = Font(color="000000", bold=True)
        fillRevenue = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        
        borderStyleBold = Border(
            right=Side(style='thin'),
            top=Side(style='medium'),
            bottom=Side(style='medium')
        )
        borderStyleBoldLR = Border(
            left=Side(style='medium'),
            right=Side(style='medium')
        )
        borderStyleBoldRight = Border(
            right=Side(style='medium'),
            top=Side(style='medium'),
            bottom=Side(style='medium')
        )
        borderStyleStandard = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Changer le formatage de la ligne des mois
        for row in ws.iter_rows(min_row=startRow, max_row=startRow, min_col=2, max_col=ws.max_column):
            for index, cell in enumerate(row):
                if index != ws.max_column-4:
                    cell.font = fontMonth
                    cell.fill = fillMonth
                    cell.alignment = alignmentCenter
                    cell.border = borderStyleBold

        normal = False
        # Appliquer les formatages
        for index_row, row in enumerate(ws.iter_rows(min_row=startRow+2, max_row=ws.max_row, min_col=0, max_col=ws.max_column)):
            if normal:
                for index_column, cell in enumerate(row):
                    # Pour les cases du tableau
                    if (index_column != ws.max_column-3) and (index_column != ws.max_column-4) and (index_column != ws.max_column-1):
                        if (index_column == 0):
                            cell.border = Border(
                                right=Side(style='medium'),
                                bottom=Side(style='thin'),
                            )
                        else:
                            cell.border = borderStyleStandard
                    # Pour la bordure à droite des tableaux
                    elif (index_column == ws.max_column-4) or (index_column == ws.max_column-1):
                        cell.border = Border(
                            right=Side(style='medium'),
                            bottom=Side(style='thin')
                        )
                        
                    # Pour la ligne du bas
                    if (index_row == ws.max_row-startRow-2) and (index_column != ws.max_column-2):
                        if (index_column == 0):
                            cell.border = Border(
                                right=Side(style='medium'),
                                bottom=Side(style='medium'),
                            )
                        elif (index_column == ws.max_column-4) or (index_column == ws.max_column-1):
                            cell.border = Border(
                                right=Side(style='medium'),
                                bottom=Side(style='medium')
                            )
                        else:
                            cell.border = Border(
                                right=Side(style='thin'),
                                bottom=Side(style='medium')
                            )

                    if (index_column == ws.max_column-3):
                        cell.border = Border(
                            right=Side(style='medium')
                        )

            # Appliquer le formatage pour la ligne 'Revenus'
            else:
                for index_column, cell in enumerate(row):
                    if (index_column == 0):
                        cell.font = fontRevenue
                        cell.fill = fillRevenue
                        cell.border =  Border(
                            right=Side(style='medium'),
                            top=Side(style='medium'),
                            bottom=Side(style='medium')
                        )
                    elif (index_column != ws.max_column-3):
                        cell.font = fontRevenue
                        cell.fill = fillRevenue
                        cell.border = borderStyleBold
                    else:
                        cell.border = borderStyleBoldLR

                row[-1].border = borderStyleBoldRight
                normal = True
        


        # Appliquer la largeur de la colonne A (environ 200 pixels)
        column_width = 250 / 7  # Conversion approximative de pixels à largeur de colonne
        ws.column_dimensions['A'].width = column_width

        column_width = 100 / 7
        ws.column_dimensions['P'].width = column_width

        for cell in ws['P']:
            cell.alignment = alignmentCenter

        # Ajouter l'espacement après le dernier bloc de données
        ws.append([''] * ws.max_column)  # Ajouter une ligne vide pour l'espacement
        for _ in range(spacing-1):  # Ajouter le nombre de lignes d'espacement spécifié
            ws.append([''] * ws.max_column)


        
        # Appliquer le style de bordure à la cellule de la dernière ligne et de l'avant-dernière colonne
        last_row_idx = ws.max_row-spacing
        second_last_col_idx = ws.max_column - 2
        specific_cell = ws.cell(row=last_row_idx, column=second_last_col_idx + 1)  # Ajuster pour l'index 1-based
        specific_cell.border = Border(
            right=Side(style='thin'),
            bottom=Side(style='medium')
        )

    def CreateDirectories(self):
        """
        Vérifie l'existence des dossiers et sous-dossiers dans le chemin spécifié et les crée si nécessaire.
        """
        # Obtenez le répertoire parent du fichier pour créer les dossiers
        directory = os.path.dirname(self.outputFile)
        
        # Vérifiez si le répertoire existe, sinon créez-le
        if not os.path.exists(directory):
            os.makedirs(directory)

